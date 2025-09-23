from django.shortcuts import render, redirect, get_object_or_404
from django.views.generic import TemplateView
from django.http import HttpResponse
from .forms import OrderExcelUploadForm, ProductImportForm
from django.contrib import messages

from kseurasia_manage_app.models import *
from django.views.decorators.http import require_POST
import io
import pandas as pd
from typing import Any, Dict, List, Tuple, Optional
from dateutil import parser as dateparser  # 3.6 では fromisoformat がないため
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Q, Count

import openpyxl
from openpyxl.styles import Border, Side, Font,Alignment
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.styles.colors import COLOR_INDEX  # 関数内インポートOK
import json
from django.conf import settings
import logging
logger = logging.getLogger(__name__)
from django.http import FileResponse
from pathlib import Path
from urllib.parse import quote
from django.utils.encoding import escape_uri_path

from datetime import datetime, date, timedelta
from django.utils import timezone

import re
from collections import OrderedDict,defaultdict
from django.db.models.functions import Coalesce
from django.db.models.expressions import ExpressionWrapper
from django.db.models import Count, F, IntegerField

from decimal import Decimal, InvalidOperation, ROUND_HALF_UP,ROUND_DOWN
from django.urls import reverse
from dateutil.relativedelta import relativedelta
import zipfile

# Create your views here.
def index(request):
    # form = OrderExcelUploadForm()
    # return render(request, 'kseurasia_manage_app/index.html', {'form': form})
    return render(request, 'kseurasia_manage_app/index.html', {'message': 'OK'})

PO_TEMPLATE_TO_ROYAL_MAP = {
    # ID系
    "jan": "Jan_code",
    "jan code": "Jan_code",
    "jan_code": "Jan_code",
    "janコード": "Jan_code",
    "janｺｰﾄﾞ": "Jan_code",
    "janコード（jan code）": "Jan_code",
    "JAN": "Jan_code",

    "product number": "Product_number",
    "product_number": "Product_number",
    "商品コード": "Product_number",
    "商品ｺｰﾄﾞ": "Product_number",
    "prod no": "Product_number",

    "sku number": "SKU_number",
    "sku": "SKU_number",
    "sku_number": "SKU_number",

    # ブランド・名称
    "brand": "Brand_name",
    "brand name": "Brand_name",
    "brand_name": "Brand_name",
    "Brand name": "Brand_name",

    "product name": "Product_name",
    "produt name": "Product_name",  # 元データの綴り揺れ対策
    "description of goods": "Product_name",
    "商品名": "Product_name",
    "品名": "Product_name",
    "商品名（日本語）": "日本語名",
    "商品名": "日本語名",

    "english name (ru)": "English_name",
    "english name": "English_name",

    "rassian name (ru)": "Rassian_name",
    "russian name": "Rassian_name",
    "ロシア語名": "Rassian_name",

    # 仕様
    "contents": "Contents",
    "内容量": "Contents",

    "volume": "Volume",

    "case q'ty": "Case_qty",
    "case qty": "Case_qty",
    "ケース入数": "Case_qty",
    "ケース\n入数": "Case_qty",

    "lot": "Lot",
    "ロット": "Lot",

    # 数量・金額
    "q'ty": "Order",
    "qty": "Order",
    "order": "Order",
    "数量": "Order",
    "オーダー": "Order",
    "発注数": "Order",

    "unit price": "Unit_price",
    "単価": "Unit_price",
    "定価": "Unit_price",
    "上代\n（税抜）": "Unit_price",

    "amount": "Amount",
    "金額": "Amount",
    "小計": "Amount",

    # 体積・重量（単品/合計）
    "case volume": "Case_volume",
    "ケース容積": "Case_volume",
    "measurement": "Case_volume",

    "case weight": "Case_weight",
    "gross weight": "Case_weight",
    "ケース重量": "Case_weight",

    "ttl volume": "TTL_volume",
    "total volume": "TTL_volume",
    "合計容積": "TTL_volume",

    "ttl weight": "TTL_weight",
    "total weight": "TTL_weight",
    "合計重量": "TTL_weight",

    "unit n/w(kg)": "Unit_nw",
    "unit n/w": "Unit_nw",
    "net weight(kg)": "Unit_nw",
    "net weight": "Unit_nw",
    "ネット重量": "Unit_nw",
    "ネット重量\n中身+容器　（ｇ）": "Unit_nw",

    "ttl n/w(kg)": "TTL_nw",
    "total n/w(kg)": "TTL_nw",
    "ttl n/w": "TTL_nw",
    "total n/w": "TTL_nw",
    "合計n/w": "TTL_nw",

    "product size": "Product_size",
    "商品サイズ": "Product_size",
    "size": "Product_size",
    "サイズ": "Product_size",

    "ingredients": "Ingredients",
    "成分": "Ingredients",

    # 原価・利益（必要なら）
    "仕入値": "Purchase_price",
    "納価\n（税抜）": "Purchase_price",
    "仕入値合計": "Purchase_amount",
    "仕入合計": "Purchase_amount",
    "利益": "profit",
    "利益率": "profit_rate",

    # DS系（ロシア向けメタ）
    "реквизиты дс": "DS_details",
    "марка (бренд) дс": "DS_brandname",
    "производель дс": "DS_Manufacturer",
}

NP_PO_FIELD_MAP = {
    # 品名
    "品名": "Description_of_goods",
    "description of goods": "Description_of_goods",
    "description": "Description_of_goods",
    "商品名": "Description_of_goods",
    "商品名（日本語）":"日本語名",

    "商品コード":"Jan_code",

    "サイズ": "商品サイズ",

    "ネット重量\n中身+容器　（ｇ）": "Unit_NW",

    # 数量
    "数量": "ORDER",
    "qty": "ORDER",
    "q'ty": "ORDER",
    "order": "ORDER",
    "発注数": "ORDER",

    # 単価・金額
    "単価": "Unit_price",
    "unit price": "Unit_price",
    "金額": "Amount",
    "合計": "Amount",
    "小計": "Amount",
    "定価": "Unit_price",
    "上代\n（税抜）": "Unit_price",
    "納価\n（税抜）": "仕入値",
    "仕入値": "仕入値",
    "仕入値合計": "仕入値合計",
    "仕入合計": "仕入値合計",
    "オーダー":"ORDER",

    # 参照系
    "jan": "Jan_code",
    "jan code": "Jan_code",
    "janコード": "Jan_code",
    "JANコード": "Jan_code",
    "型番": "Артикул",
    "артикул": "Артикул",
    "hs code": "HS_CODE",
    "JAN": "Jan_code",
    "Brand name": "Brand_name",
}


INVOICE_ALIASES: Dict[str, str] = {
  "FLOUVEIL": 19,
  "リレント通常注文": 20,
  "C'BON": 21,
  "Q'1st-1": 22,
  "HIMELABO": 23,
  "SUNSORIT": 24,
  "ELEGADOLL": 25,
  "MAYURI": 26,
  "ATMORE": 27,
  "DIME HEALTH CARE": 28,
  "LAPIDEM": 29,
  "ROSY DROP": 30,
  "ESTLABO": 31,
  "MEROS": 32,
  "COSMEPRO": 33,
  "AFURA": 34,
  "PECLIA": 35,
  "LEJEU": 36,
  "AISHODO": 37,
  "Dr.MEDION": 38,
  "McCoy": 39,
  "Luxces": 40,
  "Evliss": 41,
  "Esthe Pro Labo": 42,
  "COCOCHI　発注書": 43,
  "PURE BIO": 44,
  "BEAUTY GARAGE": 45,
  "DIAMANTE": 46
}

SALES_TABLE_RC_MAP: Dict[str, str] = {
  '(FLOUVEIL)': 'FLOUVEIL',
  '"(RELENT)': 'リレント通常注文',
  "C'BON": "C'BON",
  'Q1st': "Q'1st-1",
  'CHANSON': 'CHANSON',
  'HIMELABO': 'HIMELABO',
  'SUNSORIT': 'SUNSORIT',
  'KYOTOMO': 'KYOTOMO',
  'COREIN': '',
  'ELEGADOLL': 'ELEGADOLL',
  'MAYURI': 'MAYURI',
  'ATMORE': 'ATMORE',
  'OLUPONO': '',
  'DIME HEALTH CARE': 'DIME HEALTH CARE',
  'EMU': 'EMU',
  'CHIKUHODO': 'CHIKUHODO',
  'LAPIDEM': 'LAPIDEM',
  'MARY PLATINUE': 'MARY.P',
  'POD(ROSY DROP)': 'ROSY DROP',
  'CBS(ESTLABO)': 'ESTLABO',
  'DOSHISHA': '',
  'ISTYLE': 'ISTYLE',
  'MEROS': 'MEROS',
  'STAR LAB': 'STARLAB',
  'Beauty Conexion': 'Beauty Conexion',
  'COSMEPRO': 'COSMEPRO',
  'AFURA': 'AFURA',
  'PECLIA': 'PECLIA',
  'OSATO': 'OSATO',
  'HANAKO': 'HANAKO',
  'LEJEU': 'LEJEU',
  'AISHODO': 'AISHODO',
  'CARING JAPAN (RUHAKU)': 'RUHAKU',
  'MEDION': 'Dr.MEDION',
  'McCoy': 'McCoy',
  'URESHINO': 'URESHINO',
  'Luxces': 'Luxces',
  'Evliss': 'Evliss',
  'Pro Labo': 'Esthe Pro Labo',
  'Rey Beaty': 'Rey Beauty',
  'Pure Bio': 'PURE BIO',
  'Diaasjapan': 'Diaas',
  'DIAMANTE': 'DIAMANTE',
  'FAJ': 'FAJ',
  'ＣＨＡＮＳＯＮ': 'CHANSON',
  'Kyo Tomo': 'KYOTOMO',
  'Rey.': 'Rey Beauty',
  '（FLOUVEIL）': 'FLOUVEIL',
  '（RELENT）': 'リレント通常注文',
  '(CBON)': "C'BON",
  '(Q1st)': "Q'1st-1",
  '(姫ラボ）': 'HIMELABO',
  '(SUNSORIT)': 'SUNSORIT',
  'AISEN': '',
  'MARY PL.': 'MARY.P',
  'BEAUTY CONEXION': 'Beauty Conexion',
  'Rey': 'Rey Beauty',
  'FLOUVEIL→\nセンコン': 'FLOUVEIL',
  'センコン→\nKS\n(FLOUVEIL分）': 'FLOUVEIL',
  'RELENT→\nKS': 'リレント通常注文',
  'CBON→\nセンコン': "C'BON",
  "センコン→\nKS\n(C'BON分）": "C'BON",
}

AP_BRAND_MAP: Dict[str, str] = {
    'FLOUVEIL': 'FLOUVEIL',
    "C'BON": "C'BON",
    'RELENT': 'リレント通常注文',
    'HIMELABO': 'HIMELABO',
    'UTENA': 'UTENA',
    'SUNSORIT': 'SUNSORIT',
    'CHANSON': 'CHANSON',
    'EMS': 'EMS',
    'KYO TOMO': 'KYOTOMO',
    'COREIN': 'COREIN',
    'MAYURI': 'MAYURI',
    'Q1sｔ': "Q'1st-1",
    'ELEGADOLL': 'ELEGADOLL',
    'DIME HEALTH CARE': 'DIME HEALTH CARE',
    'ATMORE': 'ATMORE',
    'ESTLABO': 'ESTLABO',
    'CBS': 'CBS',
    'MEROS  ロイソ剤': '',
    'LEJEU': 'LEJEU',
    'PECLIA': 'PECLIA',
    'COSMEPRO': 'COSMEPRO',
    "BEAUTY CONEXION": 'Beauty Conexion',
    'AFURA': 'AFURA',
    'be-10': 'be-10',
    'AISHODO': 'AISHODO',
    'CBON  SAMPLE': "C'BON",
    'CBON  mini sample': "C'BON",
    'CBON　SAMPLE': "C'BON",
    'LUXCES': 'Luxces',
    'McCoy': 'McCoy',
    'MEDION': 'Dr.MEDION',
    'Diaas JAPAN': 'Diaas',
    'REY BEAUTY': 'Rey Beauty',
    'ESTHE PRO LABO': 'Esthe Pro Labo',
    'EVLISS': 'Evliss',
    'PURE BIO': 'PURE BIO',
    'COCOCHI　発注書': "COCOCHI　発注書",
    'BEAUTY GARAGE': "BEAUTY GARAGE",
    'COCOCHI': "COCOCHI　発注書",
    'RELENT　通常注文': 'リレント通常注文',
    "EMU":"EMU",
    "LAPIDEM":"LAPIDEM",
    "MARY PLATINUE":'MARY.P',
    "DIAMANTE":"DIAMANTE",
    "CHIKUHODO":"CHIKUHODO",
    "FAJ":"FAJ",
    "ROSY DROP":"ROSY DROP",
    "DOSHISHA":"DOSHISHA",
    "ISTYLE":"ISTYLE",
    "STAR LAB":"STARLAB",
    "RUHAKU":"RUHAKU",
    "OLUPONO":"OLUPONO",
    "OSATO":"OSATO",
}

#excelインポート
HEADER_MAP: Dict[str, str] = {
    # ---- 代表列名（左：Excel側正規化名 → 右：OrderContentのフィールド） ----
    "Jan code": "Jan_code",
    "Product Number": "Product_number",
    "Brand name": "Brand_name",

    "SKU Number (RU)": "SKU_number",

    "Produt Name": "Product_name",
    "日本語名": "日本語名",
    "ORDER": "ORDER",

    "English Name (RU)": "English_name",
    "RASSIAN NAME (RU)": "Rassian_name",

    "Contents": "Contents",
    "Volume": "Volume",
    "Case Q'ty": "Case_qty",
    "LOT": "Lot",

    "Unit price": "Unit_price",
    "Amount": "Amount",
    "Case Volume": "Case_volume",
    "Case Weight": "Case_weight",
    "Case Q'ty": "Case_qty2",
    "TTL Volume": "TTL_volume",
    "TTL Weight": "TTL_weight",
    "商品サイズ": "Product_size",
    "Unit N/W(kg)": "Unit_nw",
    "TTL N/W(kg)": "TTL_nw",
    "Ingredients": "Ingredients",

    "仕入値": "Purchase_price",
    "仕入値合計": "Purchase_amount",
    "利益": "profit",
    "利益率": "profit_rate",

    # ロシア用
    "Реквизиты ДС": "DS_details",
    "Марка (бренд) ДС": "DS_brandname",
    "Производель ДС": "DS_Manufacturer",
}

NIPPONIKATRADING_HEADER_MAP: Dict[str, str] = {
    "HS CODE": "HS_CODE",
    "Jan code": "Jan_code",
    "Артикул": "Артикул",
    "Brand name": "Brand_name",
    "日本語名": "日本語名",
    "Description of goods": "Description_of_goods",
    "Наименование ДС англ": "Наименование_ДС_англ",
    "Наименование ДС рус": "Наименование_ДС_рус",
    "Contents": "Contents",
    "LOT": "LOT",
    "Case Q'ty": "Case_Qty",
    "ORDER": "ORDER",
    "Unit price": "Unit_price",
    "Amount": "Amount",
    "仕入値": "仕入値",
    "仕入値合計": "仕入値合計",
    "利益": "利益",
    "利益率": "利益率",
    "ケース容積": "ケース容積",
    "ケース重量": "ケース重量",
    "ケース数量": "ケース数量",
    "合計容積": "合計容積",
    "合計重量": "合計重量",
    "商品サイズ": "商品サイズ",
    "Unit N/W(kg)": "Unit_NW",
    "Total N/W(kg)": "Total_NW",
    "成分": "成分",
    "Марка (бренд) ДС": "Марка_бренд_ДС",
    "Производель ДС": "Производель_ДС",
}

YAMATO_TOYO_HEADER_MAP: Dict[str, str] = {
    "Brand": "Brand",
    "Order \nCode": "Order_Code",
    "Item Name": "Item_Name",
    "Quantity": "Quantity",
    "Unit \nprice\nJPY": "Unit_price_JPY",
    "Amount\nJPY": "Amount_JPY",
    "販売価格": "販売価格",
    "輸出額": "輸出額",
    "利益": "利益",
    "利益率": "利益率",
    "pcs\n/ct": "pcs_ct",
    "CTN": "CTN",
}

VENDOR_CHOICES = [
    ("yamato_toyo", "YAMATO/TOYO TRADING社向け"),
    ("royal_cosmetics", "ROYAL COSMETICS社向け"),
    ("nipponikatrading", "NIPPONIKATRADING社向け"),
]

MODEL_MAP = {
    "yamato_toyo": YAMATO_TOYO_ProductInfo,     # ← 実在モデル名に合わせてください
    "royal_cosmetics": RY_ProductInfo,
    "nipponikatrading": NIPPONIKATRADING_ProductInfo,
}

LABEL_MAP = dict(VENDOR_CHOICES)  # キー→ラベル

def _resolve_vendor(request):
    v = request.GET.get("vendor") or request.POST.get("vendor") or "nipponikatrading"
    return v if v in MODEL_MAP else "nipponikatrading"

# ====== Brand名→テンプレシート名のエイリアス（未解決はTODO） ======
BRAND_TO_SHEET_ALIAS: Dict[str, str] = {
    "CBON": "C'BON","CBON mini sample": "C'BON","CBON SAMPLE": "C'BON","CBON　SAMPLE": "C'BON","CBON　TESTER": "C'BON",
    "EST LABO": "ESTLABO", "EST LABO PRO": "ESTLABO", "ESTLABO": "ESTLABO", "ESTLABO PRO TESTER": "ESTLABO", "ESTLABO STAND": "ESTLABO", "ESTLABO TESTER": "ESTLABO", "LABO+": "ESTLABO", "LABO+ PRO": "ESTLABO", "LABO+ PRO TESTER": "ESTLABO", "LABO+ TESTER": "ESTLABO", "MOTHERMO": "ESTLABO",
    "Elega Doll": "ELEGADOLL","ELEGADOLL TESTER": "ELEGADOLL","Elega Doll PRO": "ELEGADOLL",
    "Evliss": "Evliss", "EVLISS TESTER": "Evliss",
    "LUXCES": "Luxces", "Luxces TESTER": "Luxces",
    "Lapidem": "LAPIDEM", "Lapidem PRO": "LAPIDEM","LAPIDEM": "LAPIDEM", "Lapidem TESTER": "LAPIDEM","Lapidem PRO TESTER": "LAPIDEM",
    "Cosmepro": "COSMEPRO", "Cosmepro TESTER": "COSMEPRO", "Cosmepro PRO": "COSMEPRO",
    "AFURA": "AFURA", "AFURA TESTER": "AFURA",
    "Beaty Conexion": "Beauty Conexion", "Beauty Conexion TESTER": "Beauty Conexion",
    "COCOCHI": "COCOCHI　発注書", "COCOCHI TESTER": "COCOCHI　発注書",  # 全角スペースあり
    "Hime Labo": "HIMELABO",
    "Lejeu TESTER": "LEJEU", "LEJEU": "LEJEU",
    "HANAKO": "HANAKO", "HANAKO TESTER": "HANAKO",
    "McCoy": "McCoy", "McCoy PRO": "McCoy", "McCoy mini pouch": "McCoy", "McCoy TESTER": "McCoy", "McCoy PRO TESTER": "McCoy", "McCoy SANPLE": "McCoy",
    "MEDION": "Dr.Medion", "MEDION PRO": "Dr.Medion", "MEDION sample": "Dr.Medion", "MEDION TESTER": "Dr.Medion",
    "Diaasjapan ": "Diaas", "Diaasjapan mini sample": "Diaas", "Diaasjapan TESTER": "Diaas",
    "ROSY DROP": "ROSY DROP", "ROSY DROP SAMPLE": "ROSY DROP", "ROSY DROP TESTER": "ROSY DROP",
    "Relent": "リレント通常注文", "RELENT PRO": "リレント通常注文", "Relent Sample": "リレント無料提供", "Relent TESTER": "リレント無料提供",
    "AISHODO": "AISHODO", "AISHODO TESTER": "AISHODO",
    "Ajuste": "Ajuste",
    "Atmore": "ATMORE",
    "CHANSON": "CHANSON", "CHANSON TESTER": "CHANSON",
    "Quality 1st": "Q'1st-1", "Quality 1st SAMPLE": "Q'1st-1", "Quality 1st TESTER": "Q'1st-1",
    "Sunsorit": "SUNSORIT", "Sunsorit SAMPLE": "SUNSORIT",
    "DIAMANTE": "DIAMANTE", "DIAMANTE TESTER": "DIAMANTE",
    "Chikuhodo": "CHIKUHODO",
    "SUNTREG": "SUNTREG",
    "MAYURI": "MAYURI", "MAYURI TESTER": "MAYURI",
    "Rey Beauty Studio.": "Rey Beauty",
    "URESHINO": "URESHINO", "URESHINO TESTER": "URESHINO",
    "Kyo Tomo": "KYOTOMO", "Kyo Tomo PRO": "KYOTOMO",
    "Esthe Pro Labo": "Esthe Pro Labo", "Esthe Pro Labo TESTER": "Esthe Pro Labo", "Esthe Pro Labo SAMPLE": "Esthe Pro Labo",
    "MARY PLATINUE": "MARY.P", "MARY PLATINUE TESTER": "MARY.P", "Skin Innovation": "MARY.P", "Skin Innovation PRO": "MARY.P",
    "MEROS": "MEROS", "MEROS TESTER": "MEROS",
    "RUHAKU": "RUHAKU", "RUHAKU TESTER": "RUHAKU", "RUHAKU SAMPLE": "RUHAKU",
    "Olupono": "OLUPONO",
    "Salon de Flouveil": "FLOUVEIL", "Salon de Flouveil SAMPLE": "FLOUVEIL",
    "PECLIA": "PECLIA",
    "Star Lab Cosmetics": "STARLAB",
    "OSATO": "OSATO",
    "Emu No Shizuku": "EMU",
    "BEAUTY GARAGE": "BEAUTY GARAGE",
    "BELEGA": "BELEGA",
    "DENBA": "DENBA",
    "Healing Relax": "Healing Relax",
    "PureBio": "PURE BIO", "Purebio TESTER": "PURE BIO",
    "Dime Health Care PRO": "DIME HEALTH CARE",
    "Lishan": "ISTYLE", "Lishan TESTER": "ISTYLE",
    "Tilla Caps": "FAJ",
    #以下YAMATO_TOYO


    # === TODO: 未解決ブランドの寄せ先が決まり次第ここに追記してください ===
    # "beautygarage": "Rey Beauty",           # 例
    # "dimehealthcarepro": "（対象シート名）",
}

PURCHASE_ORDER_SHEET_ALIAS: Dict[str, str] = {
    #"INV No.": "",
    "Jan code": "Jan code",
    "Brand name": "Brand name",
    "Description of goods": "Produt Name",
    "Case Q'ty": "Case Q'ty",
    "LOT": "LOT",
    "Q'ty": "ORDER",
    "仕入値": "仕入値",
    "仕入値合計": "仕入値合計",
    "ケース容積": "Case Volume",
    "ケース重量": "Case Weight",
    "ケース数量": "Case Q'ty",
    "合計容積": "商品サイズ",
    "合計重量": "Unit N/W(kg)",
    "Unit N/W(kg)": "Unit N/W(kg)",
    "Total N/W(kg)": "TTL N/W(kg)",
    "成分": "Contents",
    "商品名": "Produt Name",
    "JANコード": "Jan code",
    "数量": "ORDER",
    "単位": "Case Q'ty",
    "金額": "Unit price",
    "合計": "Amount",
    "c/s": "ORDER",
    "商品コード": "Product Number",
    "商品名（日本語）": "Produt Name",
    "サイズ": "商品サイズ",
    "ネット重量 中身+容器 （ｇ）": "Unit N/W(kg)",
    "定価": "Unit price",
    "オーダー": "ORDER",
    "仕入合計": "仕入値合計",
    "НАМИМЕНОВАНИЕ": "RASSIAN NAME (RU)",
    "JAN": "Jan code"
}

RC_INVOICE_MAP: Dict[str, str] = {
    "Description of goods":"Product_name",
    "Наименование ДС англ":"English_name",
    "Наименование ДС рус":"Rassian_name",
    "Q'ty":"Order",
    "Unit Price":"Unit_price",
    "TTL":"Amount",
    "N/W":"Unit_nw",
    "TTL N/W":"TTL_nw",
    "Артикул":"SKU_number",
    "Реквизиты ДС":"DS_details",
    "Марка (бренд) ДС":"DS_brandname",
    "Производель ДС":"DS_Manufacturer",
}
RC_PACKING_MAP: Dict[str, str] = {
    "Case":"Case Weight",
    "Q'ty":"Order",
    "Description of goods":"Product_name",
    "Наименование ДС англ":"English_name",
    "Наименование ДС рус":"Rassian_name",
    "Contents":"Contents",
    "Measurement":"Case_volume",
    "Net Weight(kg)":"Unit_nw",
    "Gross Weight":"Case Q'ty",
    "Цена за шт":"Unit_price",
    "Цена TTL":"Amount",
    "Артикула":"SKU_number",
    "Реквизиты ДС":"DS_details",
    "Марка (бренд) ДС":"DS_brandname",
    "Производель ДС":"DS_Manufacturer",
}

TESTER_KEYWORDS = ["TESTER", "Tester", "tester", "SAMPLE", "Sample", "sample"]

HEADERS_JSON_PATH = settings.BASE_DIR / "data" / "headers_no_heuristics.json"
NAMES_JSON_PATH = settings.BASE_DIR / "data" / "desc2order_by_kind_strong.json"
PURCHASE_TEMPLATE_PATH = settings.BASE_DIR / "data" / "PurchaseOrder_Template.xlsx"
INVOICE_TEMPLATE_PATH = settings.BASE_DIR / "data" / "InvoicePacking_Template.xlsx"
TEMP_PATH = settings.BASE_DIR / "data" / "tmp"
SALES_TABLE_TEMPLATE_PATH = settings.BASE_DIR / "data" / "SalesTable_Template.xlsx"
FORCAST_TEMPLATE_PATH = settings.BASE_DIR / "data" / "forecast_template.xlsx"
AP_TEMPLATE_PATH = settings.BASE_DIR / "data" / "apTable_Template.xlsx"
AR_TEMPLATE_PATH = settings.BASE_DIR / "data" / "arTable_Template.xlsx"

ORDER_VENDOR_MODELS = {
    "royal": OrderContent,
    "nipponika": NIPPONIKATRADING_OrderContent,
    "yamato": YAMATO_TOYO_OrderContent,
}
ORDER_VENDOR_LABELS = {
    "royal": "ROYAL COSMETICS",
    "nipponika": "NIPPONIKA TRADING",
    "yamato": "YAMATO / TOYO",
}

def _to_bool_required(val) -> bool:
    """
    「必要/不要」や true/false, 1/0, yes/no, on/off などを厳密に True/False に変換。
    不明な値・空は False に倒す（必要なら None を返す設計にしてもOK）。
    """
    s = str(val).strip().lower()
    true_words  = {"true", "1", "yes", "y", "on", "必要", "ひつよう", "必須", "required"}
    false_words = {"false","0","no",  "n", "off","不要", "ふよう", "not required"}

    if s in true_words:
        return True
    if s in false_words:
        return False
    return False  # 既定は False（曖昧値対策）

def _infer_vendor_from_obj(obj) -> str:
    """
    モデル or バッチのbuyers から vendor を推測
    """
    # モデル型優先
    if isinstance(obj, OrderContent):
        return "royal"
    if isinstance(obj, NIPPONIKATRADING_OrderContent):
        return "nipponika"
    if isinstance(obj, YAMATO_TOYO_OrderContent):
        return "yamato"
    # 念のため buyers でも判定（存在する場合）
    buyers = getattr(getattr(obj, "batch", None), "buyers", "") or ""
    b = buyers.upper()
    if "ROYAL" in b:
        return "royal"
    if "NIPPONIKA" in b:
        return "nipponika"
    if "YAMATO" in b or "TOYO" in b:
        return "yamato"
    # デフォルト
    return "royal"

def _get_order_obj_by_pk(pk: int, vendor_hint: Optional[str] = None):
    """
    vendor 指定があれば該当モデルから、なければ 3 モデルを順に探索して取得
    """
    if vendor_hint in ORDER_VENDOR_MODELS:
        Model = ORDER_VENDOR_MODELS[vendor_hint]
        return get_object_or_404(Model, pk=pk)

    # vendor 指定なし：各モデルを順に探索
    for Model in (OrderContent, NIPPONIKATRADING_OrderContent, YAMATO_TOYO_OrderContent):
        try:
            return Model.objects.get(pk=pk)
        except Model.DoesNotExist:
            continue
    # どれにも無ければ 404
    raise Http404("Order not found")

def _editable_field_names(Model):
    """
    編集可能フィールド名のリスト（id/created_at/updated_at/batch などは除外）
    """
    names = []
    for f in Model._meta.get_fields():
        if not getattr(f, "concrete", False):
            continue
        if getattr(f, "auto_created", False):
            continue
        if isinstance(f, models.AutoField):
            continue
        if f.name in ("id", "created_at", "updated_at", "batch", "batch_id"):
            continue
        names.append(f.name)
    return names

def _labelize(name: str) -> str:
    """
    ラベル用の軽い整形（日本語名はそのまま、英字はスネーク→タイトル）
    """
    try:
        # 日本語などはそのまま
        name.encode("ascii")
    except UnicodeEncodeError:
        return name
    # 英字は軽く変換
    return name.replace("_", " ").title()

def _fields_for_display(obj, vendor: str):
    """
    テーブル表示用に {name, label, value, multiline, mono, readonly} を生成
    - モデル全フィールドを基本対象（編集不可は created/updated/batch 等）
    - 長文っぽい列名で multiline に寄せる簡易ルール
    """
    Model = obj.__class__
    editable = set(_editable_field_names(Model))
    out = []
    for f in Model._meta.get_fields():
        if not getattr(f, "concrete", False):
            continue
        if getattr(f, "auto_created", False):
            continue

        name = f.name
        if name in ("id", "batch_id"):
            # id と batch_id は表示（読み取り専用）
            readonly = True
        elif name in ("created_at", "updated_at", "batch"):
            readonly = True
        else:
            readonly = (name not in editable)

        # 値の取得と簡易フォーマット
        val = getattr(obj, name, None)
        label = _labelize(name)
        multiline = any(k in name.lower() for k in ("detail", "details", "ingredient", "description"))
        mono = False

        out.append({
            "name": name,
            "label": label,
            "value": val,
            "readonly": readonly,
            "multiline": multiline,
            "mono": mono,
        })
    # id を先頭、作成/更新日時は末尾に寄せる軽い並べ替え
    def keyer(row):
        order = {"id": -2, "batch_id": -1, "created_at": 98, "updated_at": 99}.get(row["name"], 50)
        return (order, row["label"])
    out.sort(key=keyer)
    return out

def _last_header_col(ws, header_row: int) -> int:
    last = 0
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if v not in (None, ""):
            last = c
    # 念のためフォールバック（ヘッダーが空だった場合）
    return last or ws.max_column

def _as_decimal_money(v) -> Optional[Decimal]:
    """通貨系の文字列/数値を Decimal に正規化。空や不正値は None。"""
    if v is None:
        return None
    s = str(v).strip()
    if not s:
        return None
    # カンマ・通貨記号等の簡易除去（必要なら追加）
    s = s.replace(",", "").replace("¥", "").replace("JPY", "").replace("$", "")
    try:
        return Decimal(s)
    except InvalidOperation:
        try:
            # Excel 由来の float 文字列などの救済
            return Decimal(str(float(s)))
        except Exception:
            return None

def _norm_po_header(s: str) -> str:
    if s is None:
        return ""
    s = str(s).strip()
    s = s.replace("\u00A0", " ").replace("\u3000", " ")
    s = s.replace("’", "'").replace("‘", "'").replace("“", '"').replace("”", '"')
    while "  " in s:
        s = s.replace("  ", " ")
    return s.lower()

# 正規化済みルックアップ辞書
_PO_CANON = { _norm_po_header(k): v for k, v in PO_TEMPLATE_TO_ROYAL_MAP.items() }

def resolve_po_field(header_text: str) -> Optional[str]:
    """テンプレ列見出しから ROYAL(OrderContent) のフィールド名を返す。未ヒットは None。"""
    return _PO_CANON.get(_norm_po_header(header_text))

# 正規化済みルックアップ辞書
_PO_CANON_NIPPONIKA = { _norm_po_header(k): v for k, v in NP_PO_FIELD_MAP.items() }

def resolve_po_field_nipponika(header_text: str) -> Optional[str]:
    """テンプレ列見出しから ROYAL(OrderContent) のフィールド名を返す。未ヒットは None。"""
    return _PO_CANON_NIPPONIKA.get(_norm_po_header(header_text))


def parse_money_to_int(value) -> Optional[int]:
    """
    金額表示（例: '1,234', '¥2,500', '\3,000', '4,500円' など）を整数に変換する。
    整数に変換できない場合は None を返す。
    """
    if value is None:
        return 0

    # 数値やDecimalが来た場合はそのまま整数化
    if isinstance(value, (int, float)):
        try:
            return int(value)
        except Exception:
            return None

    # 文字列として扱う
    s = str(value).strip()

    # 数字と小数点、マイナス以外を除去（円記号・￥・,・空白など削除）
    s = re.sub(r"[^\d\.-]", "", s)

    if s == "":
        return None

    try:
        return int(float(s))  # 小数点が入っていても丸められる
    except ValueError:
        return None

def detect_header_rows(ws,b, min_hits: int = 4, max_scan_rows: int = 50) -> dict:
    """
    先頭から max_scan_rows 行まで走査し、マップに min_hits 以上ヒットした行をヘッダー候補とみなす。
    返り値: {"normal": 行番号 or None, "tester": 行番号 or None}
    """
    normal = None
    tester = None

    for r in range(1, 200):
        hits = 0
        any_texts = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if v in (None, ""):
                continue
            any_texts.append(str(v))
            if b == "rc":
                if resolve_po_field(v):
                    hits += 1
            elif b == "nipponika":
                if resolve_po_field_nipponika(v):
                    hits += 1

        if hits >= min_hits:
            # 行内テキストに tester の語を含むなら TESTER ヘッダー扱い
            if normal is None:
                normal = r
            elif normal is not None:
                tester = r

            # 両方見つかったら早期終了
            if normal is not None and tester is not None:
                break

    return {"normal": normal, "tester": tester}

def build_po_colmap(ws, b, header_row: int) -> Dict[int, str]:
    """
    指定行の列見出しを読み、{列番号: OrderContentフィールド名} を返す。
    """
    colmap = {}
    for col in range(1, ws.max_column + 1):
        hdr = ws.cell(header_row, col).value
        if b == "rc":
            field = resolve_po_field(hdr)
        elif b == "nipponika":
            field = resolve_po_field_nipponika(hdr)
        if field:
            colmap[col] = field
    return colmap

def _to_int(v):
    """数量等を int 化（'1.0' や '1,000' 吸収）"""
    if v is None or v == "":
        return None
    try:
        return int(float(str(v).replace(",", "")))
    except Exception:
        return None

# ========= 数量の数値化（空/文字列ゆれ→0） =========
def _as_qty(v: Any) -> float:
    if v is None:
        return 0.0
    s = str(v).strip()
    if not s:
        return 0.0
    # ○/●/x 等で発注を表す場合の簡易対応（必要なければ削除してOK）
    if s in {"○", "●", "◯", "x", "X"}:
        return 1.0
    # カンマ等除去
    s = s.replace(",", "")
    try:
        return float(s)
    except:
        return 0.0
    
def _to_num(v) -> float:
    """数値っぽいものを float に。None/空 は 0。'1,234' なども吸収"""
    if v is None: 
        return 0.0
    s = str(v).strip().replace(",", "")
    if s == "":
        return 0.0
    try:
        return float(s)
    except Exception:
        return 0.0

# ===== Excel→モデルの列名マッピング（正規化後キー→OrderContentフィールド） =====

def _norm(s: str) -> str:
    if s is None:
        return ""
    s = str(s).strip().replace("\u3000", " ")
    s = re.sub(r"\s+", " ", s)
    return s.lower()

def build_header_index(ws, header_row: int):
    """
    指定したヘッダー行を読み、2つの辞書を返す:
      - exact: {元の文字列そのまま: 列番号}
      - norm : {正規化した文字列: 列番号}
    """
    exact = {}
    normd = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if v is None: 
            continue
        s = str(v).strip()
        if not s:
            continue
        exact[s] = c
        normd[_norm(s)] = c
    return exact, normd

from typing import Dict
def find_col(header_index_norm: Dict[str, int], *candidates: str) -> Optional[int]:
    """
    候補名（ゆらぎ含む複数）から最初に見つかった列番号を返す
    """
    for name in candidates:
        col = header_index_norm.get(_norm(name))
        if col:
            return col
    return None

def safe_insert_blank_rows(ws, idx: int, amount: int = 1):
    """
    行 idx の直前に amount 行を挿入する。
    既存の結合セルは、挿入行にかからないように再構築する（= 挿入行は非結合に保つ）。

    ルール:
      - 完全に idx より上の結合…そのまま
      - 完全に idx より下の結合…行数ぶん下へシフト
      - idx をまたぐ結合（min_row < idx <= max_row）…上側と下側に分割して再結合
        （＝挿入行は結合に含めない）
    """
    # 1) 既存の結合を控える
    old_ranges = list(ws.merged_cells.ranges)

    # 2) いったん全部ほどく
    for rng in old_ranges:
        ws.unmerge_cells(str(rng))

    # 3) 行を挿入
    ws.insert_rows(idx, amount=amount)

    # 4) 結合を再構築
    for rng in old_ranges:
        r1, r2, c1, c2 = rng.min_row, rng.max_row, rng.min_col, rng.max_col

        if r2 < idx:
            # そのまま
            new_parts = [(r1, r2, c1, c2)]
        elif r1 >= idx:
            # まとめて下へシフト
            new_parts = [(r1 + amount, r2 + amount, c1, c2)]
        else:
            # idx をまたいでいる → 上下に分割（挿入行は含めない）
            upper = (r1, idx - 1, c1, c2) if r1 <= idx - 1 else None
            lower = (idx + amount, r2 + amount, c1, c2) if idx <= r2 else None
            new_parts = [p for p in (upper, lower) if p is not None]

        # 面積>1 のものだけ再結合（1セルは結合不要）
        for nr1, nr2, nc1, nc2 in new_parts:
            if nr1 < nr2 or nc1 < nc2:
                ws.merge_cells(start_row=nr1, start_column=nc1, end_row=nr2, end_column=nc2)

def safe_insert_blank_cols(ws, idx: int, amount: int = 1):
    """
    列 idx の直前に amount 列を挿入する。
    既存の結合セルは、挿入列にかからないように再構築する（= 挿入列は非結合に保つ）。

    ルール:
      - 完全に idx より左の結合 … そのまま
      - 完全に idx 以右の結合 … 列数ぶん右へシフト
      - idx をまたぐ結合（min_col < idx <= max_col）… 左右に分割して再結合
        （＝挿入列は結合に含めない）
    """
    # 1) 現在の結合を退避
    old_ranges = list(ws.merged_cells.ranges)

    # 2) いったん全て解除
    for rng in old_ranges:
        ws.unmerge_cells(str(rng))

    # 3) 列を挿入
    ws.insert_cols(idx, amount=amount)

    # 4) 結合を再構築
    for rng in old_ranges:
        r1, r2, c1, c2 = rng.min_row, rng.max_row, rng.min_col, rng.max_col

        if c2 < idx:
            # 挿入位置より左側に完全に存在 → そのまま
            new_parts = [(r1, r2, c1, c2)]
        elif c1 >= idx:
            # 挿入位置より右側に完全に存在 → 右へシフト
            new_parts = [(r1, r2, c1 + amount, c2 + amount)]
        else:
            # 挿入列をまたぐ → 左右に分割（挿入列は含めない）
            left  = (r1, r2, c1, idx - 1)              if c1 <= idx - 1 else None
            right = (r1, r2, idx + amount, c2 + amount) if idx <= c2   else None
            new_parts = [p for p in (left, right) if p is not None]

        # 面積が 1×1 を超えるときだけ再結合（1セルは結合不要）
        for nr1, nr2, nc1, nc2 in new_parts:
            if (nr2 - nr1) >= 1 or (nc2 - nc1) >= 1:
                ws.merge_cells(start_row=nr1, start_column=nc1, end_row=nr2, end_column=nc2)

# ========= 取り込み =========
def import_orders(request):
    """
    .xlsx を受け取り、buyers判定→商品名キーで *ProductInfo を検索→
    各社の *OrderContent に一括登録し、その内容で Invoice/Packing のExcelを
    テンプレから生成して project/data/tmp 配下に保存する。
    """
    if request.method != "POST":
        # GET: フォーム表示
        form = OrderExcelUploadForm()
        return render(request, "kseurasia_manage_app/index.html", {"form": form})


    form = OrderExcelUploadForm(request.POST, request.FILES)
    if not form.is_valid():
        messages.error(request, "フォームの入力に誤りがあります。")
        return render(request, "kseurasia_manage_app/import_orders.html", {"form": form})

    f = form.cleaned_data["file"]
    sheet_name = form.cleaned_data.get("sheet_name") or 0

    raw = io.BytesIO(f.read())

    # --- 1) ヘッダー判定（buyers 決定） ---
    hdr = 2
    try:
        top_df = pd.read_excel(io.BytesIO(raw.getvalue()), sheet_name=sheet_name,
                               engine="openpyxl", header=None, nrows=3, usecols="A:D")
    except Exception as e:
        messages.error(request, f"先頭プレビューの読み込みに失敗しました: {e}")
        return render(request, "kseurasia_manage_app/import_orders.html", {"form": form})

    a1 = top_df.iat[0, 0] if top_df.shape[0] > 0 else ""
    d3 = top_df.iat[2, 3] if top_df.shape[0] > 2 and top_df.shape[1] > 3 else ""
    if "ROYAL COSMETICS FOR DUBAI" in str(a1):
        hdr = 4; buyers = "NIPPONIKATRADING"
        logger.info("ヘッダー行・買い手を自動判定: NIPPONIKATRADING (A1)")
        price_field = "Unit_price"
        amount_field = "Amount"
        money_quant = Decimal("1")
        purchase_price_field = "仕入値"
        purchase_amount_field = "仕入値合計"
        profit_field = "利益"
        profit_rate_field = "利益率"
    elif "ROYAL COSMETICS" in str(a1):
        hdr = 2; buyers = "ROYAL COSMETICS"
        logger.info("ヘッダー行・買い手を自動判定: ROYAL COSMETICS (A1)")
        price_field = "Unit_price"
        amount_field = "Amount"
        money_quant = Decimal("1")
        purchase_price_field = "Purchase_price"
        purchase_amount_field = "Purchase_amount"
        profit_field = "profit"
        profit_rate_field = "profit_rate"
    elif "KSユーラシア株式会社様" in str(d3):
        hdr = 19; buyers = "YAMATO_TOYO"
        logger.info("ヘッダー行・買い手を自動判定: YAMATO_TOYO (D3)")
        price_field = "販売価格"
        amount_field = "総販売価格"
        money_quant = Decimal("1")
        purchase_price_field = "Unit_price_JPY"
        purchase_amount_field = "Amount_JPY"
        profit_field = "利益"
        profit_rate_field = "利益率"
    else:
        logger.warning(f"ヘッダー行・買い手の自動判定に失敗しました: A1='{a1}' D3='{d3}'")
        messages.error(request, f"ヘッダー行・買い手の自動判定に失敗しました: A1='{a1}' D3='{d3}'")
        return render(request, "kseurasia_manage_app/import_orders.html", {"form": form})

    # --- 2) 本読み込み ---
    try:
        df = pd.read_excel(io.BytesIO(raw.getvalue()), sheet_name=sheet_name,
                           engine="openpyxl", header=hdr)
    except Exception as e:
        messages.error(request, f"Excelの読み込みに失敗しました: {e}")
        return render(request, "kseurasia_manage_app/import_orders.html", {"form": form})

    if df.empty:
        messages.warning(request, "Excelにデータがありません。")
        return render(request, "kseurasia_manage_app/import_orders.html", {"form": form})

    # --- 3) 列マッピング（未定義なら列名そのまま） ---
    def norm_header(s: str) -> str:
        if s is None: return ""
        s = str(s)
        s = s.replace("\u00A0", " ").replace("\u3000", " ")  # NBSP/全角空白
        s = s.replace("’", "'").replace("‘", "'").replace("“", '"').replace("”", '"')
        s = s.replace("\r\n", "\n").replace("\r", "\n")
        # 連続空白の縮約（改行は残す）
        s = re.sub(r"[ \t]+", " ", s).strip()
        return s
    
    if buyers == "ROYAL COSMETICS":
        header_map = HEADER_MAP
    elif buyers == "YAMATO_TOYO":
        header_map = YAMATO_TOYO_HEADER_MAP
    else:
        header_map = NIPPONIKATRADING_HEADER_MAP

    mapped = []
    for oc in df.columns:
        key = norm_header(oc)
        # まずはそのままヒット
        dest = header_map.get(key)
        if dest is None and key in header_map:
            # None は“無視”の合図
            continue
        if dest:
            mapped.append((oc, dest))
        else:
            # だめ押しで未正規化キーでも探す（稀に必要）
            dest2 = header_map.get(oc)
            if dest2:
                mapped.append((oc, dest2))

    # --- 4) バッチ作成 ---
    batch = ImportBatch.objects.create(
        source_filename=(getattr(f, "name", "") or "")[:255],
        sheet_name=str(sheet_name) if sheet_name != 0 else "",
        buyers=buyers,
    )

    # ========== この関数内だけのインラインヘルパ ==========
    def _norm(v):
        if v is None or (isinstance(v, float) and pd.isna(v)): return None
        s = str(v).strip()
        return s or None
    
    def _as_int_qty(v) -> int:
        if v is None: return 0
        s = str(v).strip().replace(",", "")
        if s in {"○","●","◯","x","X"}:  # 記号→1
            return 1
        try:
            # Excel 由来の 1.0 なども吸収
            return int(float(s))
        except Exception:
            return 0
    def _qty_from(data: Dict[str, Any]) -> int:
        for key in ("Order", "ORDER", "Quantity"):
            if key in data and data[key] not in (None, ""):
                return _as_int_qty(data[key])
        return 0

    if buyers == "ROYAL COSMETICS":
        ProductInfoModel = RY_ProductInfo
        OrderModel = OrderContent
        qty_field_for_order = "Order"
        copy_fields = [
            "Jan_code","Product_number","Brand_name","SKU_number","Product_name",
            "English_name","Rassian_name","Contents","Volume","Case_qty","Lot",
            "Unit_price","Amount","Case_volume","Case_weight","Case_qty2",
            "TTL_volume","TTL_weight","Product_size","Unit_nw","TTL_nw","Ingredients",
            "Purchase_price","Purchase_amount","profit","profit_rate",
            "DS_details","DS_brandname","DS_Manufacturer",
        ]
        def _find_product(row: pd.Series):
            name_candidates = [row.get("Product name"), row.get("Product_name"), row.get("日本語名")]
            prod_no = _norm(row.get("Product number") or row.get("Product_number"))
            jan = _norm(row.get("Jan code") or row.get("JAN code") or row.get("Jan_code"))
            for nm in name_candidates:
                n = _norm(nm)
                if n:
                    hit = ProductInfoModel.objects.filter(Product_name__iexact=n).first()
                    if hit: return hit
                    hit = ProductInfoModel.objects.filter(English_name__iexact=n).first()
                    if hit: return hit
                    hit = ProductInfoModel.objects.filter(Rassian_name__iexact=n).first()
                    if hit: return hit
            if prod_no:
                hit = ProductInfoModel.objects.filter(Product_number__iexact=prod_no).first()
                if hit: return hit
            if jan:
                hit = ProductInfoModel.objects.filter(Jan_code__iexact=jan).first()
                if hit: return hit
            return None

    elif buyers == "YAMATO_TOYO":
        ProductInfoModel = YAMATO_TOYO_ProductInfo
        OrderModel = YAMATO_TOYO_OrderContent
        qty_field_for_order = "Quantity"
        copy_fields = [
            "Brand","Order_Code","Item_Name","Quantity","Unit_price_JPY","Amount_JPY",
            "販売価格","輸出額","利益","利益率","pcs_ct","CTN",
        ]
        def _find_product(row: pd.Series):
            name_candidates = [row.get("Item Name"), row.get("Item_Name")]
            code = _norm(row.get("Order Code") or row.get("Order_Code"))
            for nm in name_candidates:
                n = _norm(nm)
                if n:
                    hit = ProductInfoModel.objects.filter(Item_Name__iexact=n).first()
                    if hit: return hit
            if code:
                hit = ProductInfoModel.objects.filter(Order_Code__iexact=code).first()
                if hit: return hit
            return None

    else:  # NIPPONIKATRADING
        ProductInfoModel = NIPPONIKATRADING_ProductInfo
        OrderModel = NIPPONIKATRADING_OrderContent
        qty_field_for_order = "ORDER"
        copy_fields = [
            "HS_CODE","Jan_code","Артикул","Brand_name","日本語名","Description_of_goods",
            "Наименование_ДС_англ","Наименование_ДС_рус","Contents","LOT","Case_Qty",
            "ORDER","Unit_price","Amount","仕入値","仕入値合計","利益","利益率",
            "ケース容積","ケース重量","ケース数量","合計容積","合計重量","商品サイズ",
            "Unit_NW","Total_NW","成分","Марка_бренд_ДС","Производель_ДС",
        ]
        def _find_product(row_data: Dict[str, Any]):
            name_candidates = [
                row_data.get("日本語名"),
                row_data.get("Description_of_goods"),
                row_data.get("Наименование_ДС_англ"),
                row_data.get("Наименование_ДС_рус"),
            ]
            art = _norm(row_data.get("Артикул"))
            jan = _norm(row_data.get("Jan_code"))

            for nm in name_candidates:
                n = _norm(nm)
                if not n:
                    continue
                hit = (
                    ProductInfoModel.objects.filter(日本語名__iexact=n).first()
                    or ProductInfoModel.objects.filter(Description_of_goods__iexact=n).first()
                    or ProductInfoModel.objects.filter(Наименование_ДС_англ__iexact=n).first()
                    or ProductInfoModel.objects.filter(Наименование_ДС_рус__iexact=n).first()
                )
                if hit: return hit

            if art:
                hit = ProductInfoModel.objects.filter(Артикул__iexact=art).first()
                if hit: return hit
            if jan:
                hit = ProductInfoModel.objects.filter(Jan_code__iexact=jan).first()
                if hit: return hit

            logger.warning(f"ProductInfo の突合に失敗しました: {name_candidates} / Артикул='{art}' Jan='{jan}'")
            return None

    order_fields = {f.name for f in OrderModel._meta.get_fields() if getattr(f, "concrete", False)}
    logger.info(f"OrderContent のフィールド: {order_fields}")
    product_fields = {f.name for f in ProductInfoModel._meta.get_fields() if getattr(f, "concrete", False)}
    logger.info(f"ProductInfo のフィールド: {product_fields}")

    # --- 5) 行ループ：数量>0 だけ、ProductInfo 突合→Order に成形 ---
    royal_objects: List[OrderContent] = []
    nippon_objects: List[NIPPONIKATRADING_OrderContent] = []
    yamato_objects: List[YAMATO_TOYO_OrderContent] = []

    logger.info(f"取り込み行数: {len(df)}")
    logger.info(f"マッピングされた列: {mapped}")
    for _, row in df.iterrows():
        row_data: Dict[str, Any] = {}
        for excel_col, model_field in mapped:
            val = row.get(excel_col)
            row_data[model_field] = None if pd.isna(val) else val

        #logger.info(f"処理中の行データ: {row_data}")
        if _qty_from(row_data) <= 0:
            continue
        prod = _find_product(row_data)
        req_flg = bool(getattr(prod, "RequiredProduct_flg", False)) if prod else False

        data: Dict[str, Any] = {}

        # 1) ProductInfo からコピー
        if prod:
            for src in copy_fields:
                if src in product_fields:
                    v = getattr(prod, src, None)
                    if v not in (None, ""):
                        data[src] = str(v)

        # 2) 行側の数量を会社の数量カラムへ
        qty_int = _qty_from(row_data)          # ← int で取得
        if qty_int > 0:
            data[qty_field_for_order] = qty_int

        #合計金額だけ計算して格納
        unit_price = _as_decimal_money(row_data.get(price_field))
        if unit_price is None:
            unit_price = _as_decimal_money(data.get(price_field))

        if qty_int > 0 and unit_price is not None:
            # まず単価を通貨の粒度に丸める（JPYなら整数）
            unit_price_q = unit_price.quantize(money_quant, rounding=ROUND_HALF_UP)
            amount_val = (unit_price_q * Decimal(qty_int)).quantize(money_quant, rounding=ROUND_HALF_UP)

            # モデルにそのフィールドがある場合だけ代入（保険）
            if price_field in order_fields:
                if money_quant == Decimal("1"):
                    data[price_field] = int(unit_price_q)   # 円は整数で保存
                else:
                    data[price_field] = unit_price_q        # 小数を保持

            if amount_field in order_fields:
                if money_quant == Decimal("1"):
                    data[amount_field] = int(amount_val)
                else:
                    data[amount_field] = amount_val
        else:
            # 単価が取れない場合は金額計算をスキップ（必要ならログ）
            logger.info(f"Skip amount calc: qty={qty_int}, unit_price={unit_price}")

        if purchase_price_field and purchase_amount_field:
            pp = _as_decimal_money(row_data.get(purchase_price_field))
            if pp is None:
                # ProductInfoからコピー済みの値が data[...] に入っている場合も使う
                pp = _as_decimal_money(data.get(purchase_price_field))

            if qty_int > 0 and pp is not None:
                pp_q = pp.quantize(money_quant, rounding=ROUND_HALF_UP)
                pa_val = (pp_q * Decimal(qty_int)).quantize(money_quant, rounding=ROUND_HALF_UP)

                # 仕入れ値そのものも、整数通貨ならintで揃える（任意）
                if purchase_price_field in order_fields and data.get(purchase_price_field) is None:
                    data[purchase_price_field] = int(pp_q) if money_quant == Decimal("1") else pp_q

                if purchase_amount_field in order_fields:
                    data[purchase_amount_field] = int(pa_val) if money_quant == Decimal("1") else pa_val
            else:
                logger.info(f"Skip purchase_amount calc: qty={qty_int}, purchase_price={pp}")

        try:
            amt_val = _as_decimal_money(data.get(amount_field))
            pa_val  = _as_decimal_money(data.get(purchase_amount_field))
        except NameError:
            amt_val = _as_decimal_money(data.get(amount_field))
            pa_val  = _as_decimal_money(data.get(purchase_amount_field))

        if amt_val is not None and pa_val is not None:
            try:
                # 利益（通貨粒度で丸め：JPYなら整数）
                profit_val = (amt_val - pa_val).quantize(money_quant, rounding=ROUND_HALF_UP)
                # 利益率（= 原価率%）。小数点2桁で丸め
                rate_val = None
                if amt_val != 0:
                    rate_val = (pa_val / amt_val * Decimal("100")).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

                # モデルにフィールドがあるときだけ入れる（保険）
                if profit_field in order_fields:
                    data[profit_field] = int(profit_val) if money_quant == Decimal("1") else profit_val
                if rate_val is not None and profit_rate_field in order_fields:
                    # Decimal のままでもOK（DecimalField想定）。floatにしたければ float(rate_val)
                    data[profit_rate_field] = rate_val
            except Exception as e:
                logger.info(f"Skip profit calc: amt={amt_val}, purchase_amt={pa_val}, err={e}")

        # 3) よく使うキーを補完（OrderModel にあれば）
        for k in ("Jan_code","Product_number","SKU_number","Brand_name","Product_name",
                  "日本語名","Item_Name","Order_Code"):
            if k in order_fields:
                v = _norm(row_data.get(k))
                if v is not None and k not in data:
                    data[k] = v

        data["batch"] = batch
        data["RequiredProduct_flg"] = req_flg
        if buyers == "YAMATO_TOYO":
            data["Brand"] = "UTENA"
        cleaned = {k: v for k, v in data.items() if k in order_fields}

        if OrderModel is OrderContent:
            royal_objects.append(OrderModel(**cleaned))
        elif OrderModel is YAMATO_TOYO_OrderContent:
            yamato_objects.append(OrderModel(**cleaned))
        else:
            nippon_objects.append(OrderModel(**cleaned))

    if not (royal_objects or nippon_objects or yamato_objects):
        messages.warning(request, "ORDER（数量）が入っている行が見つかりませんでした。")
        return render(request, "kseurasia_manage_app/import_orders.html", {"form": form})

    logger.info(f"登録予定オブジェクト数: ROYAL={len(royal_objects)} NIPPONIKA={len(nippon_objects)} YAMATO={len(yamato_objects)}")

    # --- 6) 一括登録 ---
    try:
        with transaction.atomic():
            if royal_objects:
                royal_obj = OrderContent.objects.bulk_create(royal_objects, batch_size=500)
            if nippon_objects:
                nipponika_obj = NIPPONIKATRADING_OrderContent.objects.bulk_create(nippon_objects, batch_size=500)
            if yamato_objects:
                yamat_obj = YAMATO_TOYO_OrderContent.objects.bulk_create(yamato_objects, batch_size=500)

    except Exception as e:
        messages.error(request, f"登録中にエラーが発生しました: {e}")
        return render(request, "kseurasia_manage_app/import_orders.html", {"form": form})

    # =========================
    # 7) Invoice/Packing 作成
    # =========================
    from pathlib import Path
    from django.conf import settings
    from openpyxl import load_workbook, Workbook
    import os, datetime

    # ▼ ブランド→シート割当（実データに差し替えてください）
    INVOICE1 = ["FLOUVEIL","リレント通常注文","C'BON","Q'1st-1"]
    INVOICE2 = ["HIMELABO","SUNSORIT","ELEGADOLL","MAYURI","ATMORE","DIME HEALTH CARE","ROSY DROP","LAPIDEM","ESTLABO","MEROS","COSMEPRO","AFURA","PECLIA","LEJEU"]
    INVOICE3 = ["Dr.Medion","AISHODO","Luxces","McCoy","Esthe Pro Labo","Evliss","PURE BIO","COCOCHI　発注書","BEAUTY GARAGE","DIAMANTE"]
    # TESTER/SAMPLE はINVOICE4/5へ
    INVOICE4 = ["リレント通常注文","C'BON","Q'1st-1","CHANSON","LAPIDEM"]
    INVOICE5 = ["MARY.P","ROSY DROP","MEROS","B-10","ELEGADOLL","MAYURI","Mediplorer","McCoy","PURE BIO","Glow","ISTYLE"]

    NIPPONIKA_INVOICE = ["リレント通常注文","C'BON","LAPIDEM","B-10","ROSY DROP","BEAUTY GARAGE","MARY.P"]

    total_dict = {}
    qty_dict = {}

    def _sheet_for_royal(item_brand: str, item_name: str) -> str:
        name_u = (item_name or "").upper()
        if "TESTER" in name_u:
            return "INVOICE4"
        if "SAMPLE" in name_u:
            return "INVOICE5"
        b = (item_brand or "").strip()
        if b in INVOICE1: return "INVOICE1"
        if b in INVOICE2: return "INVOICE2"
        if b in INVOICE3: return "INVOICE3"
        # どれにも属さなければ INVOICE1 に寄せる（必要なら変えてOK）
        return "INVOICE1"

    # 出力先ディレクトリ
    base_dir = Path(getattr(settings, "BASE_DIR", Path(__file__).resolve().parent.parent))
    out_dir = TEMP_PATH
    out_dir.mkdir(parents=True, exist_ok=True)

    # テンプレ探索（プロジェクト内に無ければ /mnt/data の配布物を使う）
    def _template_path(filename: str) -> Path:
        from pathlib import Path
        cand = [
            settings.BASE_DIR / "data" / filename,   # 例: BASE_DIR/data/InvoicePacking_Template.xlsx
            Path("/mnt/data") / filename,            # アップロード済みのフォールバック
        ]
        for p in cand:
            if p.exists():
                return p
        raise FileNotFoundError(f"テンプレートが見つかりません: {filename}")
    
    created_files = []

    thin_border = Border(
        left=Side(style="thin"),   # 左
        right=Side(style="thin"),  # 右
        top=Side(style="thin"),    # 上
        bottom=Side(style="thin")  # 下
    )
    invoice_font = Font(
        name="Times New Roman",   # フォント名（例: メイリオ）
        size=14,         # 文字サイズ
        bold=False,       # 太字
        italic=False,    # 斜体
        color="000000"   # 赤色（RGB指定）
    )
    top_font = Font(
        name="Arial",   # フォント名（例: メイリオ）
        size=16,         # 文字サイズ
        bold=True,       # 太字
        italic=False,    # 斜体
        color="000000"   # 赤色（RGB指定）
    )

    # --- ROYAL: InvoicePacking_Template.xlsx ---
    if royal_objects:
        sub_total_dict = {
            "INVOICE1": 14,
            "INVOICE2": 14,
            "INVOICE3": 14,
            "INVOICE4": 14,
            "INVOICE5": 14,
            "PL1": 13,
            "PL2": 13,
            "PL3": 13,
        }

        inv_sums = defaultdict(lambda: defaultdict(float))   # 例: inv_sums["INVOICE1"]["Q'ty"] += ...
        inv_header_pos = {}                                  # 例: inv_header_pos["INVOICE1"]["Q'ty"] == 列番号
        pack_sums = defaultdict(lambda: defaultdict(float))  # 例: pack_sums["PL1"][10] += 値  (J列など列番号)

        tpl = _template_path("InvoicePacking_Template.xlsx")
        wb = load_workbook(tpl)
        # 各明細を適切なシートに出力
        for obj in royal_objects:
            brand_name = getattr(obj, "Brand_name", "") or ""
            product_name = getattr(obj, "Product_name", "") or ""
            if brand_name in BRAND_TO_SHEET_ALIAS:
                sheet_name = BRAND_TO_SHEET_ALIAS[brand_name]
            else:
                logger.warning(f"ブランド名の割当が未定義: {brand_name}")
                continue
            
            tester_flg = False
            cont_flg = False
            for kw in TESTER_KEYWORDS:
                if kw in product_name:
                    if sheet_name in INVOICE4:
                        ws = wb["Invoice-4 (TESTER)"]
                        tester_flg = True
                        invoice = "INVOICE4"
                        sub_total_dict["INVOICE4"] += 1
                    elif sheet_name in INVOICE5:
                        ws = wb["Invoice-5 (TESTER) "]
                        tester_flg = True
                        invoice = "INVOICE5"
                        sub_total_dict["INVOICE5"] += 1
                    else:
                        logger.warning(f"ブランド名の割当が不明 TESTER配列定義に無し: {brand_name} / {product_name}")
                        cont_flg = True
                    break
            if cont_flg:
                continue

            if not tester_flg:
                if sheet_name in INVOICE1:
                    ws = wb["Invoice-1 "]
                    ws2 = wb["PL1"]
                    invoice = "INVOICE1"
                    packing = "PL1"
                    sub_total_dict["INVOICE1"] += 1
                    sub_total_dict["PL1"] += 1
                elif sheet_name in INVOICE2:
                    ws = wb["Invoice-2"]
                    ws2 = wb["PL2"]
                    invoice = "INVOICE2"
                    packing = "PL2"
                    sub_total_dict["INVOICE2"] += 1
                    sub_total_dict["PL2"] += 1
                elif sheet_name in INVOICE3:
                    ws = wb["Invoice-3"]
                    ws2 = wb["PL3"]
                    invoice = "INVOICE3"
                    packing = "PL3"
                    sub_total_dict["INVOICE3"] += 1
                    sub_total_dict["PL3"] += 1
                else:
                    logger.warning(f"ブランド名の割当が不明 INVOICE配列定義に無し: {brand_name}")
                    continue
            # ヘッダーが無い想定：必要なら追加（1行目）
            safe_insert_blank_rows(ws, 14, 1)
            # 初回だけ、このシートのヘッダー名 → 列番号をキャッシュ
            if invoice not in inv_header_pos:
                inv_header_pos[invoice] = {}
                for i in range(1, ws.max_column + 1):
                    hdr = ws.cell(13, i).value
                    if hdr is not None:
                        inv_header_pos[invoice][str(hdr).strip()] = i

            for i in range(1, 13):
                hdr_raw = ws.cell(13, i).value
                hdr = str(hdr_raw).strip() if hdr_raw is not None else ""
                col_name = RC_INVOICE_MAP.get(hdr, "")
                val = getattr(obj, col_name, "") if col_name else ""
                ws.cell(14, i).value = val
                ws.cell(14, i).border = thin_border
                ws.cell(14, i).font = invoice_font

            # ★ インボイス側：必要なヘッダーだけ加算（テンプレ固定想定）
            for key in ("Q'ty", "Unit Price", "TTL", "N/W", "TTL N/W"):
                col = inv_header_pos[invoice].get(key)
                if col:
                    inv_sums[invoice][key] += _to_num(ws.cell(14, col).value)

            if not tester_flg:
                # もともとの集計（ブランド別の数量/金額）も継続
                total_dict[sheet_name] = int(total_dict.get(sheet_name, 0)) + int(_to_num(obj.Amount))
                qty_dict[sheet_name]   = int(qty_dict.get(sheet_name, 0))   + int(_to_num(obj.Order))

                # Packing 明細
                safe_insert_blank_rows(ws2, 13, 1)
                for i in range(1, 21):
                    hdr_raw = ws2.cell(12, i).value
                    hdr = str(hdr_raw).strip() if hdr_raw is not None else ""
                    col_name = RC_PACKING_MAP.get(hdr, "")
                    val = getattr(obj, col_name, "") if col_name else ""
                    ws2.cell(13, i).value = val
                    ws2.cell(13, i).border = thin_border
                    ws2.cell(13, i).font = invoice_font

                    # ★ パッキング側：I～O 列 (9～15列) を加算
                    if 9 <= i <= 15:
                        pack_sums[packing][i] += _to_num(val)

                # 追加で直書きしている体積/重量も加算（この2つは上の for ではまだ入っていない）
                ws2.cell(13, 10).value = obj.TTL_volume   # J列
                ws2.cell(13, 15).value = obj.TTL_weight   # O列
                ws2.cell(13, 10).border = thin_border
                ws2.cell(13, 10).font = invoice_font
                ws2.cell(13, 15).border = thin_border
                ws2.cell(13, 15).font = invoice_font

                pack_sums[packing][10] += _to_num(obj.TTL_volume)
                pack_sums[packing][15] += _to_num(obj.TTL_weight)

        for inv_key in ("INVOICE1", "INVOICE2", "INVOICE3", "INVOICE4", "INVOICE5"):
            sum_row = sub_total_dict[inv_key]
            # シート取得
            if inv_key == "INVOICE1":
                ws_inv = wb["Invoice-1 "]
            elif inv_key == "INVOICE2":
                ws_inv = wb["Invoice-2"]
            elif inv_key == "INVOICE3":
                ws_inv = wb["Invoice-3"]
            elif inv_key == "INVOICE4":
                ws_inv = wb["Invoice-4 (TESTER)"]
            elif inv_key == "INVOICE5":
                ws_inv = wb["Invoice-5 (TESTER) "]
            else:
                continue

            # ヘッダー名→列番号
            pos = inv_header_pos.get(inv_key, {})
            # 数量だけは int、他は float のまま（必要なら丸めを加える）
            def _w(hname, cast_int=False):
                col = pos.get(hname)
                if not col:
                    return
                val = inv_sums[inv_key].get(hname, 0.0)
                if cast_int:
                    val = int(val)
                ws_inv.cell(sum_row, col).value = val
                ws_inv.cell(sum_row, col).border = thin_border
                ws_inv.cell(sum_row, col).font = invoice_font

            _w("Q'ty", cast_int=True)
            _w("Unit Price", cast_int=False)
            _w("TTL", cast_int=False)
            _w("N/W", cast_int=False)
            _w("TTL N/W", cast_int=False)

        # --- パッキング合計の固定値書き込み ---
        for pk_key, sheet_name_ in (("PL1","PL1"), ("PL2","PL2"), ("PL3","PL3")):
            sum_row = sub_total_dict[pk_key]
            ws_pk = wb[sheet_name_]
            for col in range(9, 16):  # I～O
                val = pack_sums[pk_key].get(col, 0.0)
                ws_pk.cell(sum_row, col).value = val
                ws_pk.cell(sum_row, col).border = thin_border
                ws_pk.cell(sum_row, col).font = invoice_font

        header_ws = wb["INVOICE(無償サンプル抜き価格)"]
        for b,t in total_dict.items():
            header_ws.cell(INVOICE_ALIASES[b],8).value = qty_dict[b]
            header_ws.cell(INVOICE_ALIASES[b],9).value = t

        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        out_path = out_dir / f"InvoicePacking_ROYAL_{ts}.xlsx"
        wb.save(out_path)
        created_files.append(out_path.name)
        # バッチに記録（複数あればカンマで保持）
        batch.InvoicePacking_file = ", ".join([p for p in created_files if p])
        batch.save(update_fields=["InvoicePacking_file"])

        #発注書作成
        Purchase_row_dict = {}
        Purchase_qty_dict = {}
        Purchase_buy_dict = {}
        Purchase_amount_dict = {}
        Tester_Purchase_row_dict = {}
        Tester_Purchase_qty_dict = {}
        Tester_Purchase_buy_dict = {}
        Tester_Purchase_amount_dict = {}
        Purchase_wb = load_workbook(PURCHASE_TEMPLATE_PATH)

        for obj in royal_objects:
            brand_name = getattr(obj, "Brand_name", "") or ""
            product_name = getattr(obj, "Product_name", "") or ""
            if brand_name not in BRAND_TO_SHEET_ALIAS:
                logger.warning(f"ブランド名の割当が未定義: {brand_name}")
                continue
            sheet_name = BRAND_TO_SHEET_ALIAS[brand_name]

            if sheet_name not in Purchase_wb.sheetnames:
                logger.warning(f"発注書作成Warning 対象ブランド:{brand_name} のシートがテンプレートに存在しません: {sheet_name}")
                continue
            else:
                ws = Purchase_wb[sheet_name]

            # TESTER 判定
            tester_flg = False
            for kw in TESTER_KEYWORDS:
                if kw in product_name:
                    tester_flg = True
                    break
                elif kw in brand_name:
                    tester_flg = True
                    break

            # ヘッダー自動検出
            b = "rc"
            hdr = detect_header_rows(ws, b ,min_hits=4, max_scan_rows=50)
            logger.info(f"[{sheet_name}] ヘッダー自動検出結果: {hdr}")
            header_row = None
            if tester_flg and hdr["tester"]:
                header_row = hdr["tester"]
            elif hdr["normal"]:
                header_row = hdr["normal"]
            elif hdr["tester"]:
                # 念のため：テスターのみ見つかった場合はそれを採用
                header_row = hdr["tester"]

            if not header_row:
                # フォールバック（従来固定行などがある場合はここを変更）
                header_row = 5
                logger.warning(f"[{sheet_name}] ヘッダー自動検出に失敗。fallback row={header_row}")

            # 列マップ作成
            b = "rc"
            colmap = build_po_colmap(ws,b, header_row)
            if not colmap:
                logger.warning(f"[{sheet_name}] マップに一致する列見出しが見つかりません（row={header_row}）")
                continue

            # ヘッダー直下へ1行挿入して書き込み
            insert_at = header_row + 1
            safe_insert_blank_rows(ws, insert_at, 1)

            if not tester_flg:
                if sheet_name in Purchase_row_dict:
                    Purchase_row_dict[sheet_name] += 1
                    Purchase_qty_dict[sheet_name] += _to_int(obj.Order) or 0
                    Purchase_buy_dict[sheet_name] += _to_int(obj.Purchase_price) or 0
                    Purchase_amount_dict[sheet_name] += _to_int(obj.Purchase_amount) or 0
                else:
                    Purchase_row_dict[sheet_name] = 1
                    Purchase_qty_dict[sheet_name] = _to_int(obj.Order) or 0
                    Purchase_buy_dict[sheet_name] = _to_int(obj.Purchase_price) or 0
                    Purchase_amount_dict[sheet_name] = _to_int(obj.Purchase_amount) or 0
            elif tester_flg:
                if sheet_name in Tester_Purchase_row_dict:
                    Tester_Purchase_row_dict[sheet_name] += 1
                    Tester_Purchase_qty_dict[sheet_name] += _to_int(obj.Order) or 0
                    Tester_Purchase_buy_dict[sheet_name] += _to_int(obj.Purchase_price) or 0
                    Tester_Purchase_amount_dict[sheet_name] += _to_int(obj.Purchase_amount) or 0
                else:
                    Tester_Purchase_row_dict[sheet_name] = 1
                    Tester_Purchase_qty_dict[sheet_name] = _to_int(obj.Order) or 0
                    Tester_Purchase_buy_dict[sheet_name] = _to_int(obj.Purchase_price) or 0
                    Tester_Purchase_amount_dict[sheet_name] = _to_int(obj.Purchase_amount) or 0

            for col, field in colmap.items():
                val = getattr(obj, field, "")
                # 数量/金額/数値っぽい列は数値として固定値投入
                norm_key = _norm_po_header(ws.cell(header_row, col).value)
                is_qty_col = norm_key in {"q'ty", "qty", "order", "数量"}
                is_amount_col = norm_key in {"amount", "合計", "小計","仕入値合計","仕入合計"}
                is_unit_price_col = norm_key in {"unit price", "金額", "単価","仕入値","定価","上代\n（税抜）","納価\n（税抜）"}

                if is_qty_col:
                    v_int = _to_int(val)
                    ws.cell(insert_at, col).value = v_int if v_int is not None else val
                elif is_amount_col or is_unit_price_col:
                    # 金額は float 化（整数で良ければ _to_int に切り替え）
                    try:
                        ws.cell(insert_at, col).number_format = r'¥#,##0'
                        ws.cell(insert_at, col).value = int(str(val).replace(",", ""))
                    except Exception:
                        ws.cell(insert_at, col).value = val
                else:
                    ws.cell(insert_at, col).value = val
            end_col = _last_header_col(ws, header_row)
            for c in range(1, end_col + 1):
                ws.cell(insert_at, c).border = thin_border

            top_time_str = datetime.datetime.now().strftime("%m.%Y")
            if sheet_name == "COCOCHI　発注書":
                pass
            elif sheet_name == "リレント通常注文":
                pass
            elif sheet_name == "Q'1st-1":
                pass
            else:
                ws.cell(1,1).value = f"ROYAL COSMETICS {top_time_str}輸出"
                ws.cell(1,1).font = top_font
            
        # 各シートの集計行に合計値を書き込み
        for ws_name, row in Purchase_row_dict.items():
            ws = Purchase_wb[ws_name]
            if ws_name == "COCOCHI　発注書":
                total_price = 0
                for row in range(12,Purchase_row_dict[ws_name]+12):
                    total_price += parse_money_to_int(Purchase_wb[ws_name].cell(row,6).value) if parse_money_to_int(Purchase_wb[ws_name].cell(row,6).value) else 0
                rate = Decimal("0.10")  # 10%
                t = Decimal(str(total_price))
                # 小計（税抜）—切り捨て
                subtotal = (t / (Decimal("1") + rate)).quantize(Decimal("1"), rounding=ROUND_DOWN)
                # 税額
                tax = t - subtotal
                ws.cell(Purchase_row_dict[ws_name]+12+1,6).number_format = r'¥#,##0'
                ws.cell(Purchase_row_dict[ws_name]+12+2,6).number_format = r'¥#,##0'
                ws.cell(Purchase_row_dict[ws_name]+12+3,6).number_format = r'¥#,##0'
                ws.cell(Purchase_row_dict[ws_name]+12+1,6).value = subtotal
                ws.cell(Purchase_row_dict[ws_name]+12+2,6).value = tax
                ws.cell(Purchase_row_dict[ws_name]+12+3,6).value = total_price
                continue
            elif ws_name == "リレント通常注文":
                total_price = 0
                total_buy = 0
                total_order = 0
                total_buy_amount = 0
                for row in range(3,Purchase_row_dict[ws_name]+3):
                    total_price += parse_money_to_int(ws.cell(row,5).value) if parse_money_to_int(ws.cell(row,5).value) else 0
                    total_buy += parse_money_to_int(ws.cell(row,6).value) if parse_money_to_int(ws.cell(row,6).value) else 0
                    total_order += parse_money_to_int(ws.cell(row,7).value) if parse_money_to_int(ws.cell(row,7).value) else 0
                    total_buy_amount += parse_money_to_int(ws.cell(row,8).value) if parse_money_to_int(ws.cell(row,8).value) else 0
                ws.cell(Purchase_row_dict[ws_name]+3,5).number_format = r'¥#,##0'
                ws.cell(Purchase_row_dict[ws_name]+3,6).number_format = r'¥#,##0'
                ws.cell(Purchase_row_dict[ws_name]+3,8).number_format = r'¥#,##0'
                ws.cell(Purchase_row_dict[ws_name]+3,5).value = total_price
                ws.cell(Purchase_row_dict[ws_name]+3,6).value = total_buy
                ws.cell(Purchase_row_dict[ws_name]+3,7).value = total_order
                ws.cell(Purchase_row_dict[ws_name]+3,8).value = total_buy_amount
                continue
            elif ws_name == "Q'1st-1":
                continue
            elif ws_name == "LAPIDEM" or ws_name == "ROSY DROP":
                ws.cell(row+6,9).number_format = r'¥#,##0'
                ws.cell(row+6,10).number_format = r'¥#,##0'
                ws.cell(row+6,8).value = Purchase_qty_dict.get(ws_name, 0)
                ws.cell(row+6,9).value = Purchase_buy_dict.get(ws_name, 0)
                ws.cell(row+6,10).value = Purchase_amount_dict.get(ws_name, 0)
                continue
            elif ws_name == "Diaas":
                ws.cell(row+6,8).number_format = r'¥#,##0'
                ws.cell(row+6,9).number_format = r'¥#,##0'
                ws.cell(row+6,7).value = Purchase_qty_dict.get(ws_name, 0)
                ws.cell(row+6,8).value = Purchase_buy_dict.get(ws_name, 0)
                ws.cell(row+6,9).value = Purchase_amount_dict.get(ws_name, 0)
                continue

            ws.cell(row+6,8).number_format = r'¥#,##0'
            ws.cell(row+6,9).number_format = r'¥#,##0'
            ws.cell(row+6,7).value = Purchase_qty_dict.get(ws_name, 0)
            ws.cell(row+6,8).value = Purchase_buy_dict.get(ws_name, 0)
            ws.cell(row+6,9).value = Purchase_amount_dict.get(ws_name, 0)
        for ws_name, row in Tester_Purchase_row_dict.items():
            add_row = Purchase_row_dict[ws_name] if ws_name in Purchase_row_dict else 0
            ws = Purchase_wb[ws_name]
            if ws_name == "COCOCHI　発注書":
                continue
            elif ws_name == "リレント通常注文":
                continue
            elif ws_name == "Q'1st-1":
                continue
            elif ws_name == "LAPIDEM":
                ws.cell(row+12,9).number_format = r'¥#,##0'
                ws.cell(row+12,10).number_format = r'¥#,##0'
                ws.cell(row+12,8).value = Purchase_qty_dict.get(ws_name, 0)
                ws.cell(row+12,9).value = Tester_Purchase_buy_dict.get(ws_name, 0)
                ws.cell(row+12,10).value = Tester_Purchase_amount_dict.get(ws_name, 0)
                continue
            elif ws_name == "Diaas":
                ws.cell(row+11,8).number_format = r'¥#,##0'
                ws.cell(row+11,9).number_format = r'¥#,##0'
                ws.cell(row+11,7).value = Purchase_qty_dict.get(ws_name, 0)
                ws.cell(row+11,8).value = Tester_Purchase_buy_dict.get(ws_name, 0)
                ws.cell(row+11,9).value = Tester_Purchase_amount_dict.get(ws_name, 0)
                continue
            elif ws_name == "ROSY DROP":
                ws.cell(row+10+add_row,9).number_format = r'¥#,##0'
                ws.cell(row+10+add_row,10).number_format = r'¥#,##0'
                ws.cell(row+10+add_row,8).value = Tester_Purchase_qty_dict.get(ws_name, 0)
                ws.cell(row+10+add_row,9).value = Tester_Purchase_buy_dict.get(ws_name, 0)
                ws.cell(row+10+add_row,10).value = Tester_Purchase_amount_dict.get(ws_name, 0)
                continue
            
            ws.cell(row+10+add_row,8).number_format = r'¥#,##0'
            ws.cell(row+10+add_row,9).number_format = r'¥#,##0'
            ws.cell(row+10+add_row,7).value = Tester_Purchase_qty_dict.get(ws_name, 0)
            ws.cell(row+10+add_row,8).value = Tester_Purchase_buy_dict.get(ws_name, 0)
            ws.cell(row+10+add_row,9).value = Tester_Purchase_amount_dict.get(ws_name, 0)

        Purchase_wb.remove(Purchase_wb["ORDER SHEET"])
        Purchase_wb.remove(Purchase_wb["TOTAL"])
        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        out_path = out_dir / f"PURCHASE_ROYAL_{ts}.xlsx"
        Purchase_wb.save(out_path)
        created_files.append(out_path.name)
        # バッチに記録（複数あればカンマで保持）
        batch.PurchaseOrder_file = ", ".join([p for p in created_files if p])
        batch.save(update_fields=["PurchaseOrder_file"])

    # --- NIPPONIKA: InvoicePacking_Nipponika.xlsx ---
    elif nippon_objects:

        tpl = _template_path("InvoicePacking_Nipponika.xlsx")
        wb = load_workbook(tpl)

        for obj in nippon_objects:
            brand_name = getattr(obj, "Brand_name", "") or ""
            product_name = getattr(obj, "Description_of_goods", "") or ""
            if brand_name in BRAND_TO_SHEET_ALIAS:
                sheet_name = BRAND_TO_SHEET_ALIAS[brand_name]
            else:
                logger.warning(f"ブランド名の割当が未定義: {brand_name}")
                continue

            if sheet_name in NIPPONIKA_INVOICE:
                ws = wb["Invoice-1 "]
                ws2 = wb["PL1"]
            else:
                logger.warning(f"ブランド名の割当が不明 NIPPONIKA配列定義に無し: {brand_name}")
                continue

            safe_insert_blank_rows(ws,14,1)
            ws.cell(14,2).value = obj.Description_of_goods
            ws.cell(14,3).value = obj.ORDER
            ws.cell(14,4).value = obj.Unit_price
            ws.cell(14,5).value = obj.Amount
            ws.cell(14,6).value = obj.Unit_NW
            ws.cell(14,7).value = obj.Total_NW
            ws.cell(14,8).value = obj.HS_CODE
            ws.cell(14,9).value = obj.Jan_code
            ws.cell(14,10).value = obj.Артикул
            for i in range(1,11):
                ws.cell(14,i).border = thin_border
                ws.cell(14,i).font = invoice_font

            safe_insert_blank_rows(ws2,14,1)
            ws2.cell(14,3).value = obj.ケース重量
            ws2.cell(14,4).value = obj.ORDER
            ws2.cell(14,5).value = obj.Description_of_goods
            ws2.cell(14,6).value = obj.Contents
            ws2.cell(14,7).value = obj.ケース容積
            ws2.cell(14,9).value = obj.ケース重量
            ws2.cell(14,12).value = obj.ケース数量
            ws2.cell(14,14).value = obj.Unit_price
            ws2.cell(14,15).value = obj.Amount
            for i in range(1,21):
                ws2.cell(14,i).border = thin_border
                ws2.cell(14,i).font = invoice_font


        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        out_path = out_dir / f"InvoicePacking_Nipponika_{ts}.xlsx"
        wb.save(out_path)
        created_files.append(out_path.name)
        batch.InvoicePacking_file = ", ".join([p for p in created_files if p])
        batch.save(update_fields=["InvoicePacking_file"])

        #発注書作成
        created_files = []
        Purchase_row_dict = {}
        Purchase_qty_dict = {}
        Purchase_buy_dict = {}
        Purchase_amount_dict = {}
        Tester_Purchase_row_dict = {}
        Tester_Purchase_qty_dict = {}
        Tester_Purchase_buy_dict = {}
        Tester_Purchase_amount_dict = {}
        Purchase_wb = load_workbook(PURCHASE_TEMPLATE_PATH)

        for obj in nippon_objects:
            brand_name = getattr(obj, "Brand_name", "") or ""
            product_name = getattr(obj, "Description_of_goods", "") or ""
            if brand_name not in BRAND_TO_SHEET_ALIAS:
                logger.warning(f"ブランド名の割当が未定義: {brand_name}")
                continue
            sheet_name = BRAND_TO_SHEET_ALIAS[brand_name]

            if sheet_name not in Purchase_wb.sheetnames:
                logger.warning(f"発注書作成Warning 対象ブランド:{brand_name} のシートがテンプレートに存在しません: {sheet_name}")
                continue
            else:
                ws = Purchase_wb[sheet_name]

            # TESTER 判定
            tester_flg = False
            for kw in TESTER_KEYWORDS:
                if kw in product_name:
                    tester_flg = True
                    break
                elif kw in brand_name:
                    tester_flg = True
                    break

            # ヘッダー自動検出
            b = "nipponika"
            hdr = detect_header_rows(ws,b, min_hits=4, max_scan_rows=50)
            logger.info(f"[{sheet_name}] ヘッダー自動検出結果: {hdr}")
            header_row = None
            if tester_flg and hdr["tester"]:
                header_row = hdr["tester"]
            elif hdr["normal"]:
                header_row = hdr["normal"]
            elif hdr["tester"]:
                # 念のため：テスターのみ見つかった場合はそれを採用
                header_row = hdr["tester"]

            if not header_row:
                # フォールバック（従来固定行などがある場合はここを変更）
                header_row = 5
                logger.warning(f"[{sheet_name}] ヘッダー自動検出に失敗。fallback row={header_row}")

            # 列マップ作成
            b = "nipponika"
            colmap = build_po_colmap(ws,b, header_row)
            if not colmap:
                logger.warning(f"[{sheet_name}] マップに一致する列見出しが見つかりません（row={header_row}）")
                continue

            # ヘッダー直下へ1行挿入して書き込み
            insert_at = header_row + 1
            safe_insert_blank_rows(ws, insert_at, 1)

            if not tester_flg:
                if sheet_name in Purchase_row_dict:
                    Purchase_row_dict[sheet_name] += 1
                    Purchase_qty_dict[sheet_name] += _to_int(obj.ORDER) or 0
                    Purchase_buy_dict[sheet_name] += _to_int(obj.仕入値) or 0
                    Purchase_amount_dict[sheet_name] += _to_int(obj.仕入値合計) or 0
                else:
                    Purchase_row_dict[sheet_name] = 1
                    Purchase_qty_dict[sheet_name] = _to_int(obj.ORDER) or 0
                    Purchase_buy_dict[sheet_name] = _to_int(obj.仕入値) or 0
                    Purchase_amount_dict[sheet_name] = _to_int(obj.仕入値合計) or 0
            elif tester_flg:
                if sheet_name in Tester_Purchase_row_dict:
                    Tester_Purchase_row_dict[sheet_name] += 1
                    Tester_Purchase_qty_dict[sheet_name] += _to_int(obj.ORDER) or 0
                    Tester_Purchase_buy_dict[sheet_name] += _to_int(obj.仕入値) or 0
                    Tester_Purchase_amount_dict[sheet_name] += _to_int(obj.仕入値合計) or 0
                else:
                    Tester_Purchase_row_dict[sheet_name] = 1
                    Tester_Purchase_qty_dict[sheet_name] = _to_int(obj.ORDER) or 0
                    Tester_Purchase_buy_dict[sheet_name] = _to_int(obj.仕入値) or 0
                    Tester_Purchase_amount_dict[sheet_name] = _to_int(obj.仕入値合計) or 0

            for col, field in colmap.items():
                val = getattr(obj, field, "")
                # 数量/金額/数値っぽい列は数値として固定値投入
                norm_key = _norm_po_header(ws.cell(header_row, col).value)
                is_qty_col = norm_key in {"q'ty", "qty", "order", "数量"}
                is_amount_col = norm_key in {"amount", "合計", "小計","仕入値合計","仕入合計"}
                is_unit_price_col = norm_key in {"unit price", "金額", "単価","仕入値","定価","上代\n（税抜）","納価\n（税抜）"}

                if is_qty_col:
                    v_int = _to_int(val)
                    ws.cell(insert_at, col).value = v_int if v_int is not None else val
                elif is_amount_col or is_unit_price_col:
                    # 金額は float 化（整数で良ければ _to_int に切り替え）
                    try:
                        ws.cell(insert_at, col).number_format = r'¥#,##0'
                        ws.cell(insert_at, col).value = int(str(val).replace(",", ""))
                    except Exception as e:
                        logger.warning(f"金額列の変換エラー: {e} / val={val}")
                        ws.cell(insert_at, col).value = val
                else:
                    ws.cell(insert_at, col).value = val
            end_col = _last_header_col(ws, header_row)
            for c in range(1, end_col + 1):
                ws.cell(insert_at, c).border = thin_border

            top_time_str = datetime.datetime.now().strftime("%m.%Y")
            if sheet_name == "COCOCHI　発注書":
                pass
            elif sheet_name == "リレント通常注文":
                pass
            elif sheet_name == "Q'1st-1":
                pass
            else:
                ws.cell(1,1).value = f"NIPPONIKA TRADING {top_time_str}輸出"
                ws.cell(1,1).font = top_font

        # 各シートの集計行に合計値を書き込み
        for ws_name, row in Purchase_row_dict.items():
            ws = Purchase_wb[ws_name]
            if ws_name == "COCOCHI　発注書":
                total_price = 0
                for row in range(12,Purchase_row_dict[ws_name]+12):
                    total_price += parse_money_to_int(Purchase_wb[ws_name].cell(row,6).value) if parse_money_to_int(Purchase_wb[ws_name].cell(row,6).value) else 0
                rate = Decimal("0.10")  # 10%
                t = Decimal(str(total_price))
                # 小計（税抜）—切り捨て
                subtotal = (t / (Decimal("1") + rate)).quantize(Decimal("1"), rounding=ROUND_DOWN)
                # 税額
                tax = t - subtotal
                ws.cell(Purchase_row_dict[ws_name]+12+1,6).number_format = r'¥#,##0'
                ws.cell(Purchase_row_dict[ws_name]+12+2,6).number_format = r'¥#,##0'
                ws.cell(Purchase_row_dict[ws_name]+12+3,6).number_format = r'¥#,##0'
                ws.cell(Purchase_row_dict[ws_name]+12+1,6).value = subtotal
                ws.cell(Purchase_row_dict[ws_name]+12+2,6).value = tax
                ws.cell(Purchase_row_dict[ws_name]+12+3,6).value = total_price
                continue
            elif ws_name == "リレント通常注文":
                total_price = 0
                total_buy = 0
                total_order = 0
                total_buy_amount = 0
                for row in range(3,Purchase_row_dict[ws_name]+3):
                    total_price += parse_money_to_int(ws.cell(row,5).value) if parse_money_to_int(ws.cell(row,5).value) else 0
                    total_buy += parse_money_to_int(ws.cell(row,6).value) if parse_money_to_int(ws.cell(row,6).value) else 0
                    total_order += parse_money_to_int(ws.cell(row,7).value) if parse_money_to_int(ws.cell(row,7).value) else 0
                    total_buy_amount += parse_money_to_int(ws.cell(row,8).value) if parse_money_to_int(ws.cell(row,8).value) else 0
                ws.cell(Purchase_row_dict[ws_name]+3,5).number_format = r'¥#,##0'
                ws.cell(Purchase_row_dict[ws_name]+3,6).number_format = r'¥#,##0'
                ws.cell(Purchase_row_dict[ws_name]+3,8).number_format = r'¥#,##0'
                ws.cell(Purchase_row_dict[ws_name]+3,5).value = total_price
                ws.cell(Purchase_row_dict[ws_name]+3,6).value = total_buy
                ws.cell(Purchase_row_dict[ws_name]+3,7).value = total_order
                ws.cell(Purchase_row_dict[ws_name]+3,8).value = total_buy_amount
                continue
            elif ws_name == "Q'1st-1":
                continue
            elif ws_name == "LAPIDEM" or ws_name == "ROSY DROP":
                ws.cell(row+6,9).number_format = r'¥#,##0'
                ws.cell(row+6,10).number_format = r'¥#,##0'
                ws.cell(row+6,8).value = Purchase_qty_dict.get(ws_name, 0)
                ws.cell(row+6,9).value = Purchase_buy_dict.get(ws_name, 0)
                ws.cell(row+6,10).value = Purchase_amount_dict.get(ws_name, 0)
                continue
            elif ws_name == "Diaas":
                ws.cell(row+6,8).number_format = r'¥#,##0'
                ws.cell(row+6,9).number_format = r'¥#,##0'
                ws.cell(row+6,7).value = Purchase_qty_dict.get(ws_name, 0)
                ws.cell(row+6,8).value = Purchase_buy_dict.get(ws_name, 0)
                ws.cell(row+6,9).value = Purchase_amount_dict.get(ws_name, 0)
                continue

            ws.cell(row+6,8).number_format = r'¥#,##0'
            ws.cell(row+6,9).number_format = r'¥#,##0'
            ws.cell(row+6,7).value = Purchase_qty_dict.get(ws_name, 0)
            ws.cell(row+6,8).value = Purchase_buy_dict.get(ws_name, 0)
            ws.cell(row+6,9).value = Purchase_amount_dict.get(ws_name, 0)
        for ws_name, row in Tester_Purchase_row_dict.items():
            add_row = Purchase_row_dict[ws_name] if ws_name in Purchase_row_dict else 0
            ws = Purchase_wb[ws_name]
            if ws_name == "COCOCHI　発注書":
                continue
            elif ws_name == "リレント通常注文":
                continue
            elif ws_name == "Q'1st-1":
                continue
            elif ws_name == "LAPIDEM":
                ws.cell(row+12,9).number_format = r'¥#,##0'
                ws.cell(row+12,10).number_format = r'¥#,##0'
                ws.cell(row+12,8).value = Purchase_qty_dict.get(ws_name, 0)
                ws.cell(row+12,9).value = Tester_Purchase_buy_dict.get(ws_name, 0)
                ws.cell(row+12,10).value = Tester_Purchase_amount_dict.get(ws_name, 0)
                continue
            elif ws_name == "Diaas":
                ws.cell(row+11,8).number_format = r'¥#,##0'
                ws.cell(row+11,9).number_format = r'¥#,##0'
                ws.cell(row+11,7).value = Purchase_qty_dict.get(ws_name, 0)
                ws.cell(row+11,8).value = Tester_Purchase_buy_dict.get(ws_name, 0)
                ws.cell(row+11,9).value = Tester_Purchase_amount_dict.get(ws_name, 0)
                continue
            elif ws_name == "ROSY DROP":
                ws.cell(row+10+add_row,9).number_format = r'¥#,##0'
                ws.cell(row+10+add_row,10).number_format = r'¥#,##0'
                ws.cell(row+10+add_row,8).value = Tester_Purchase_qty_dict.get(ws_name, 0)
                ws.cell(row+10+add_row,9).value = Tester_Purchase_buy_dict.get(ws_name, 0)
                ws.cell(row+10+add_row,10).value = Tester_Purchase_amount_dict.get(ws_name, 0)
                continue
            
            ws.cell(row+10+add_row,8).number_format = r'¥#,##0'
            ws.cell(row+10+add_row,9).number_format = r'¥#,##0'
            ws.cell(row+10+add_row,7).value = Tester_Purchase_qty_dict.get(ws_name, 0)
            ws.cell(row+10+add_row,8).value = Tester_Purchase_buy_dict.get(ws_name, 0)
            ws.cell(row+10+add_row,9).value = Tester_Purchase_amount_dict.get(ws_name, 0)

        Purchase_wb.remove(Purchase_wb["ORDER SHEET"])
        Purchase_wb.remove(Purchase_wb["TOTAL"])
        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        out_path = out_dir / f"PURCHASE_NIPPONIKA_{ts}.xlsx"
        Purchase_wb.save(out_path)
        created_files.append(out_path.name)
        # バッチに記録（複数あればカンマで保持）
        batch.PurchaseOrder_file = ", ".join([p for p in created_files if p])
        batch.save(update_fields=["PurchaseOrder_file"])

    # --- YAMATO: （必要ならテンプレを用意して同様に出力） ---
    if yamato_objects:
        input_wb = openpyxl.load_workbook(io.BytesIO(raw.getvalue()), data_only=False)
        input_ws = input_wb[sheet_name] if (isinstance(sheet_name, str) and sheet_name in input_wb.sheetnames) else input_wb.active

        product_info_count = YAMATO_TOYO_ProductInfo.objects.count()
        for row in range(21,product_info_count+21):
            if input_ws.cell(row,5).value is None:
                continue
            product_name = input_ws.cell(row,4).value if input_ws.cell(row,4).value else ""
            order_quantity = int(input_ws.cell(row,5).value) if input_ws.cell(row,5).value else 0
            if order_quantity <= 0:
                continue
            obj = get_object_or_404(YAMATO_TOYO_ProductInfo, Item_Name__iexact=product_name)
            if not obj:
                logger.warning(f"商品マスタに存在しない商品: {product_name}")
                continue
            sell_price = obj.販売価格 if obj.販売価格 else 0
            purchase_price = obj.Unit_price_JPY if obj.Unit_price_JPY else 0
            total_sell_price = order_quantity * sell_price
            total_purchase_price = order_quantity * purchase_price
            input_ws.cell(row,7).value = total_purchase_price
            input_ws.cell(row,9).value = total_sell_price
            input_ws.cell(row,10).value = total_sell_price - total_purchase_price
            input_ws.cell(row,11).value = (total_sell_price - total_purchase_price) / total_sell_price
            input_ws.cell(row,7).number_format = r'¥#,##0'
            input_ws.cell(row,9).number_format = r'¥#,##0'
            input_ws.cell(row,10).number_format = r'¥#,##0'
            input_ws.cell(row,11).number_format = '0.00%'

        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        out_path = out_dir / f"PURCHASE_YAMATO_{ts}.xlsx"
        input_wb.save(out_path)
        created_files.append(out_path.name)
        # バッチに記録（複数あればカンマで保持）
        batch.PurchaseOrder_file = ", ".join([p for p in created_files if p])
        batch.save(update_fields=["PurchaseOrder_file"])


    messages.success(request, "取り込みと帳票の下書き作成が完了しました。")
    return redirect("import_batch_list")

def order_list(request):
    """
    注文の一覧（初期表示で全件：新しい順）
    - ページネーション
    - 任意でバッチ/キーワード検索も可能（未指定なら全て）
    """
    batch_id = request.GET.get("batch")
    q = request.GET.get("q", "")
    page = request.GET.get("page", 1)

    # ベースクエリ：新しい順
    qs = OrderContent.objects.select_related("batch").order_by("-id")

    # 任意：バッチ絞り込み
    if batch_id:
        qs = qs.filter(batch_id=batch_id)

    # 任意：簡易検索（必要な項目は適宜追加）
    if q:
        qs = qs.filter(
            Q(Jan_code__icontains=q) |
            Q(Product_name__icontains=q) |
            Q(Brand_name__icontains=q) |
            Q(SKU_number__icontains=q)
        )

    # ページネーション
    paginator = Paginator(qs, 50)  # 1ページ50件
    page_obj = paginator.get_page(page)

    # バッチ選択用（ドロップダウン）
    batches = ImportBatch.objects.all().only("id", "created_at", "source_filename")

    return render(request, "kseurasia_manage_app/order_list.html", {
        "page_obj": page_obj,
        "total_count": qs.count(),
        "batches": batches,
        "selected_batch_id": batch_id,
        "q": q,
    })


def order_detail(request, pk: int):
    """
    単体の注文詳細（会社に応じて自動でモデルを判断）。
    GET: 表示、POST: 保存/削除
    vendor は ?vendor=royal|nipponika|yamato でヒント指定可（無指定でも自動判定）
    """

    vendor_hint = (request.GET.get("vendor") or request.POST.get("vendor") or "").lower() or None
    obj = _get_order_obj_by_pk(pk, vendor_hint)
    vendor = _infer_vendor_from_obj(obj)
    vendor_label = ORDER_VENDOR_LABELS[vendor]

    if request.method == "POST":
        action = request.POST.get("action", "save")
        if action == "delete":
            obj.delete()
            messages.success(request, f"注文 #{pk} を削除しました。")
            # バッチから来た場合はその一覧に戻す
            from_batch_id = request.GET.get("from_batch_id") or request.POST.get("from_batch_id")
            if from_batch_id:
                return redirect(reverse("batch_order_list", kwargs={"batch_id": from_batch_id}))
            return redirect("order_list")

        # 保存
        editable = _editable_field_names(obj.__class__)
        for name in editable:
            if name in request.POST:
                val = request.POST.get(name, "")

                if name == "RequiredProduct_flg":
                    setattr(obj, name, _to_bool_required(val))
                    continue

                setattr(obj, name, (val if val != "" else None))
        obj.save()
        messages.success(request, "変更を保存しました。")
        # 同じページに戻る（クエリ維持）
        return redirect(request.path + f"?vendor={vendor}")

    fields = _fields_for_display(obj, vendor)
    from_batch_id = request.GET.get("from_batch_id")
    return render(request, "kseurasia_manage_app/order_detail.html", {
        "obj": obj,
        "vendor": vendor,
        "vendor_label": vendor_label,
        "fields": fields,
        "from_batch_id": from_batch_id,
    })

def batch_order_list(request, batch_id: int):
    batch = get_object_or_404(ImportBatch, pk=batch_id)
    page = request.GET.get("page", 1)
    q = request.GET.get("q", "").strip()

    # buyers でモデル切替
    if batch.buyers == "NIPPONIKATRADING":
        Model = NIPPONIKATRADING_OrderContent
        vendor = "NIPPONIKATRADING"
    elif batch.buyers == "YAMATO_TOYO":
        Model = YAMATO_TOYO_OrderContent
        vendor = "YAMATO_TOYO"
    else:  # ROYAL
        Model = OrderContent
        vendor = "ROYAL"

    # ベンダー別：表示カラム(フィールド名, 見出し) と、数値カラム（右寄せ用）
    LIST_COLUMNS = {
        "ROYAL": [
            ("Product_name", "商品名"),
            ("Brand_name",   "ブランド"),
            ("Jan_code",     "JAN"),
            ("SKU_number",   "SKU"),
            ("Order",        "数量"),
            ("Amount",       "金額"),
        ],
        "NIPPONIKATRADING": [
            ("Description_of_goods", "商品名"),
            ("Brand_name",           "ブランド"),
            ("Артикул",              "SKU"),
            ("ORDER",                "数量"),
            ("Amount",               "金額"),
        ],
        "YAMATO_TOYO": [
            ("Item_Name",      "商品名"),
            ("Brand",          "ブランド"),
            ("Order_Code",     "JAN/コード"),
            ("Quantity",       "数量"),
            ("総販売価格",     "金額"),   # ← ここを売上合計に
        ],
    }
    NUMERIC_FIELDS = {
        "ROYAL": {"Order", "Amount"},
        "NIPPONIKATRADING": {"ORDER", "Amount"},
        "YAMATO_TOYO": {"Quantity", "総販売価格"},
    }

    # 実在するフィールドだけ残す（古いデータ差異への保険）
    model_fields = {f.name for f in Model._meta.get_fields()}
    list_columns = [(f, label) for (f, label) in LIST_COLUMNS[vendor] if f in model_fields]
    numeric_fields = NUMERIC_FIELDS[vendor] & model_fields

    qs = Model.objects.select_related("batch").filter(batch_id=batch.id).order_by("-id")

    # 検索：ベンダーごとの代表列で
    SEARCH_FIELDS = {
        "ROYAL": ["Jan_code", "Product_name", "Brand_name", "SKU_number"],
        "NIPPONIKATRADING": ["Jan_code", "Description_of_goods", "Brand_name", "Артикул"],
        "YAMATO_TOYO": ["Order_Code", "Item_Name", "Brand"],
    }
    if q:
        cond = Q()
        for f in SEARCH_FIELDS[vendor]:
            if f in model_fields:
                cond |= Q(**{f + "__icontains": q})
        qs = qs.filter(cond)

    paginator = Paginator(qs, 50)
    page_obj = paginator.get_page(page)

    # テンプレが簡単に描画できるよう、セル値＋右寄せフラグを作る
    rows = []
    for obj in page_obj:
        cells = []
        for field, _label in list_columns:
            val = getattr(obj, field, "")
            cells.append({
                "value": val,
                "is_num": field in numeric_fields,
            })
        rows.append({"obj": obj, "cells": cells})

    return render(request, "kseurasia_manage_app/batch_order_list.html", {
        "batch": batch,
        "page_obj": page_obj,
        "total_count": qs.count(),
        "q": q,
        "list_columns": list_columns,   # ← 追加
        "rows": rows,                   # ← 追加
    })

def batch_order_detail(request, batch_id: int, pk: int):
    """
    バッチ配下の注文詳細。基本は order_detail と同じ振る舞い。
    """
    vendor_hint = (request.GET.get("vendor") or request.POST.get("vendor") or "").lower() or None
    obj = _get_order_obj_by_pk(pk, vendor_hint)
    vendor = _infer_vendor_from_obj(obj)
    vendor_label = ORDER_VENDOR_LABELS[vendor]

    if request.method == "POST":
        action = request.POST.get("action", "save")
        if action == "delete":
            obj.delete()
            messages.success(request, f"注文 #{pk} を削除しました。")
            return redirect(reverse("batch_order_list", kwargs={"batch_id": batch_id}))

        editable = _editable_field_names(obj.__class__)
        for name in editable:
            if name in request.POST:
                val = request.POST.get(name, "")

                if name == "RequiredProduct_flg":
                    setattr(obj, name, _to_bool_required(val))
                    continue

                setattr(obj, name, (val if val != "" else None))
        obj.save()
        messages.success(request, "変更を保存しました。")
        return redirect(request.path + f"?vendor={vendor}")

    fields = _fields_for_display(obj, vendor)
    return render(request, "kseurasia_manage_app/order_detail.html", {
        "obj": obj,
        "vendor": vendor,
        "vendor_label": vendor_label,
        "fields": fields,
        "from_batch_id": batch_id,   # 戻り先に使う
    })

@require_POST
def delete_import_batch(request, batch_id: int):
    """バッチと関連OrderContentを削除。関連ファイルも削除。"""
    batch = get_object_or_404(ImportBatch, pk=batch_id)

    # 関連ファイル（保存していれば）を削除
    for attr in ("PurchaseOrder_file", "InvoicePacking_file"):
        val = getattr(batch, attr, "") or ""
        if val:
            p = Path(val)
            if not p.is_absolute():
                p = Path(TEMP_PATH) / val
            try:
                if p.exists() and p.is_file():
                    p.unlink()
            except Exception as e:
                logger.warning("Failed to delete file %s: %s", p, e)

    # OrderContent は on_delete=SET_NULL のため手動削除
    OrderContent.objects.filter(batch_id=batch.id).delete()

    # バッチ削除
    batch.delete()

    messages.success(request, "バッチと関連オーダーを削除しました。")
    return redirect("import_batch_list")

#以下はインポートバッチ関連
def import_batch_list(request):
    q = request.GET.get("q", "").strip()
    date_from = request.GET.get("from", "").strip()
    date_to   = request.GET.get("to", "").strip()
    page = int(request.GET.get("page", 1))

    # ここを修正：ordercontent → items（モデルの related_name に合わせる）
    batches = (
        ImportBatch.objects.all()
        .annotate(cnt_rc=Count("items", distinct=True))
        .annotate(cnt_nk=Count("nipponika_items", distinct=True))
        .annotate(cnt_yt=Count("yamato_toyo_items", distinct=True))
        .annotate(
            item_count=ExpressionWrapper(
                Coalesce(F("cnt_rc"), 0) + Coalesce(F("cnt_nk"), 0) + Coalesce(F("cnt_yt"), 0),
                output_field=IntegerField(),
            )
        )
        .order_by("-id")
    )

    if q:
        batches = (
            ImportBatch.objects.all()
            .annotate(cnt_rc=Count("items", distinct=True))
            .annotate(cnt_nk=Count("nipponika_items", distinct=True))
            .annotate(cnt_yt=Count("yamato_toyo_items", distinct=True))
            .annotate(
                item_count=ExpressionWrapper(
                    Coalesce(F("cnt_rc"), 0) + Coalesce(F("cnt_nk"), 0) + Coalesce(F("cnt_yt"), 0),
                    output_field=IntegerField(),
                )
            )
            .order_by("-id")
        )

    if date_from:
        batches = batches.filter(created_at__date__gte=date_from)
    if date_to:
        batches = batches.filter(created_at__date__lte=date_to)

    paginator = Paginator(batches, 20)
    page_obj = paginator.get_page(page)

    ctx = {
        "page_obj": page_obj,
        "q": q,
        "date_from": date_from,
        "date_to": date_to,
    }
    return render(request, "kseurasia_manage_app/import_batch_list.html", ctx)


# --- 以降はダミー（あなたが実装） -----------------
from django.http import HttpResponseNotAllowed
from django.http import FileResponse, Http404

def download_purchase_order(request, batch_id: int):
    batch = get_object_or_404(ImportBatch, pk=batch_id)

    filename_or_path = getattr(batch, "PurchaseOrder_file", "") or ""
    if not filename_or_path:
        raise Http404("このバッチには発注書ファイルが登録されていません。")

    p = Path(filename_or_path)
    # 相対（ファイル名のみ）で保存している場合は TEMP_PATH を前置
    if not p.is_absolute():
        p = Path(TEMP_PATH) / filename_or_path

    if not p.exists() or not p.is_file():
        raise Http404("発注書ファイルが見つかりません。")

    f = open(p, "rb")
    resp = FileResponse(
        f,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    quoted = quote(p.name)
    # ダウンロード時のファイル名（UTF-8対応）
    resp["Content-Disposition"] = f'attachment; filename="purchase_order.xlsx"; filename*=UTF-8\'\'{quoted}'
    return resp

def export_invoice_packing(request, batch_id: int):
    batch = get_object_or_404(ImportBatch, pk=batch_id)

    filename_or_path = getattr(batch, "InvoicePacking_file", "") or ""
    if not filename_or_path:
        raise Http404("このバッチにはIVPLファイルが登録されていません。")

    p = Path(filename_or_path)
    # 相対（ファイル名のみ）で保存している場合は TEMP_PATH を前置
    if not p.is_absolute():
        p = Path(TEMP_PATH) / filename_or_path

    if not p.exists() or not p.is_file():
        raise Http404("IVPLファイルファイルが見つかりません。")

    f = open(p, "rb")
    resp = FileResponse(
        f,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    quoted = quote(p.name)
    # ダウンロード時のファイル名（UTF-8対応）
    resp["Content-Disposition"] = f'attachment; filename="InvoicePackingList.xlsx"; filename*=UTF-8\'\'{quoted}'
    return resp

def report_console(request):
    return render(request, "kseurasia_manage_app/reports.html")

#以下売り上げ票関連
def _parse_month_yyyy_mm(s: str) -> date:
    """'YYYY-MM' → その月の1日(naive date)"""
    try:
        dt = datetime.strptime(s.strip(), "%Y-%m")
        return date(dt.year, dt.month, 1)
    except Exception:
        raise Http404("start_month / end_month は 'YYYY-MM' 形式で指定してください。")

def _month_end(d: date) -> date:
    """その月の末日(naive date)を返す"""
    y, m = d.year, d.month
    if m == 12:
        return date(y, 12, 31)
    return date(y, m + 1, 1) - timedelta(days=1)

def _get_report_params_range(request):
    """
    レポート用：クエリから
      - clients: 複数 (client=A&client=B…)
      - start_month: 'YYYY-MM'
      - end_month  : 'YYYY-MM'
      - format     : 'csv' or 'pdf'（デフォルト csv）
    を取り出し、期間 [start,end) と表示用ラベルを返す。
    """
    start_ym = (request.GET.get("start_month") or "").strip()
    end_ym   = (request.GET.get("end_month") or "").strip()
    fmt = (request.GET.get("format") or "csv").lower()

    if not start_ym or not end_ym:
        raise ValueError("client（1つ以上）, start_month, end_month は必須です")

    start_dt, end_dt, label = _month_range_aware(start_ym, end_ym)
    return start_dt, end_dt, fmt, label

def ym_to_year(ym: str) -> int:
    """'YYYY-MM' から年(YYYY)だけ取り出す。'YYYY' 単体も許容。"""
    ym = (ym or "").strip()
    try:
        return datetime.strptime(ym, "%Y-%m").year
    except ValueError:
        if re.fullmatch(r"\d{4}", ym):
            return int(ym)
        raise ValueError("ym は 'YYYY' または 'YYYY-MM' で指定してください")
    
def orders_imported_in_year(year: int):
    """指定年(YYYY)に ImportBatch.created_at が属する OrderContent を全件返す。"""
    start = timezone.make_aware(datetime(year,     1, 1, 0, 0, 0))
    end   = timezone.make_aware(datetime(year + 1, 1, 1, 0, 0, 0))
    return (
        OrderContent.objects
        .select_related("batch")
        .filter(batch__created_at__gte=start, batch__created_at__lt=end)
        .order_by("id")
    )

def _parse_ym(ym: str) -> Tuple[int, int]:
    """'YYYY-MM' → (year, month)。不正は ValueError。"""
    dt = datetime.strptime(ym, "%Y-%m")
    return dt.year, dt.month

def _month_range_aware(start_ym: str, end_ym: str) -> Tuple[datetime, datetime, str]:
    """
    [start, end) の aware datetime（終了は『終了月の翌月1日 00:00』未満）。
    併せてファイル名用のラベルも返す（例: '2024-04_2024-06'）。
    """
    sy, sm = _parse_ym(start_ym)
    ey, em = _parse_ym(end_ym)

    start = timezone.make_aware(datetime(sy, sm, 1, 0, 0, 0))
    # end = 終了月の翌月1日
    if em == 12:
        end = timezone.make_aware(datetime(ey + 1, 1, 1, 0, 0, 0))
    else:
        end = timezone.make_aware(datetime(ey, em + 1, 1, 0, 0, 0))

    if start >= end:
        raise ValueError("開始年月は終了年月以前にしてください")

    label = f"{start_ym}_{end_ym}"
    return start, end, label

def import_batches_between(start_dt: datetime, end_dt: datetime):
    """期間に作成された ImportBatch を返す queryset。"""
    return (
        ImportBatch.objects
        .filter(created_at__gte=start_dt, created_at__lt=end_dt)
        .order_by("id")
    )

def orders_between_by_batch_created(b,start_dt: datetime, end_dt: datetime):
    """
    期間に作成されたバッチに紐づく OrderContent を返す queryset。
    （売上表など“取り込み期間ベース”で集計したい場合）
    """
    if b=="ROYAL COSMETICS":
        return (
            OrderContent.objects
            .select_related("batch")
            .filter(batch__created_at__gte=start_dt, batch__created_at__lt=end_dt)
            .order_by("id")
        )
    elif b=="NIPPONIKA":
        return (
            NIPPONIKATRADING_OrderContent.objects
            .select_related("batch")
            .filter(batch__created_at__gte=start_dt, batch__created_at__lt=end_dt)
            .order_by("id")
        )
    elif b=="YAMATO_TOYO":
        return (
            YAMATO_TOYO_OrderContent.objects
            .select_related("batch")
            .filter(batch__created_at__gte=start_dt, batch__created_at__lt=end_dt)
            .order_by("id")
        )

def _download_response(binary: bytes, filename: str, content_type: str):
    resp = HttpResponse(binary, content_type=content_type)
    # 日本語ファイル名対応
    quoted = escape_uri_path(filename)
    resp["Content-Disposition"] = f'attachment; filename="{quoted}"'
    return resp

# ---- 以下4つを実装すればフロントのDLボタンが動きます ----

def reports_sales_export(request):

    thin_border = Border(
        left=Side(style="thin"),   # 左
        right=Side(style="thin"),  # 右
        top=Side(style="thin"),    # 上
        bottom=Side(style="thin")  # 下
    )
    yamato_font = Font(
        name="ＭＳ Ｐゴシック",   # フォント名（例: メイリオ）
        size=11,         # 文字サイズ
        bold=False,       # 太字
        italic=False,    # 斜体
        color="000000"   # 赤色（RGB指定）
    )

    # 期間・フォーマット等の取得（既存のヘルパ想定）
    start_dt, end_dt, fmt, label = _get_report_params_range(request)

    #R&C
    # --- 売上データの取得（辞書化） ---
    # ※ select_related は使わず、必要な列を .values(...) で取得
    rows = list(
        orders_between_by_batch_created("ROYAL COSMETICS",start_dt, end_dt)
        .filter(batch__buyers="ROYAL COSMETICS")
        .values("Brand_name", "Amount", "Purchase_amount", "batch__created_at")
        .order_by("id")
    )

    # --- 日付×ブランド(=テンプレ内のシート別名) で集計 ---
    # 例: sales_purchase_dict[date][sheet_alias] = 仕入合計
    sales_purchase_dict: Dict[datetime, Dict[str, int]] = defaultdict(lambda: defaultdict(int))
    sales_sell_dict: dict[datetime.datetime, dict[str, int]] = defaultdict(lambda: defaultdict(int))

    for r in rows:
        brand = (r["Brand_name"] or "").strip()
        created_at = r["batch__created_at"]  # datetime
        created_at = timezone.localtime(created_at)
        created_at = created_at.strftime("%Y/%m/%d")
        if not brand:
            continue

        sheet_alias = BRAND_TO_SHEET_ALIAS.get(brand)
        if not sheet_alias:
            logger.warning(f"ブランド名の割当が未定義: {brand}")
            continue

        amount = int(r["Amount"] or 0)
        purchase = int(r["Purchase_amount"] or 0)

        sales_purchase_dict[created_at][sheet_alias] += purchase
        sales_sell_dict[created_at][sheet_alias] += amount

    # データが無ければエラーメッセージ or 空返し
    if not sales_sell_dict:
        logger.info("指定期間に売上データがありません。")
        messages.warning(request, "指定期間に売上データがありません。")
        return redirect("report_console")

    # --- テンプレート読込 ---
    wb = openpyxl.load_workbook(SALES_TABLE_TEMPLATE_PATH, data_only=False)
    ws = wb["R&C"]

    def safe_insert_blank_cols(ws_, col_idx: int, how_many: int = 1):
        """結合の影響で insert_cols が失敗しないようにしたい最低限のラッパ（テンプレに合わせてそのまま挿入）。"""
        ws_.insert_cols(col_idx, how_many)

    for created_at in sorted(sales_purchase_dict.keys(), reverse=True):
        safe_insert_blank_cols(ws, 4, 1)  # D列に1列挿入
        ws.cell(2, 4).value = created_at
        ws.cell(2, 4).border = thin_border

        purchase_row = 3
        sell_row = 97
        profit_row = 147

        # 同日のブランド列挙（テンプレの行順に合わせたい場合は、ここを SALES_TABLE_RC_MAP の順で回してください）
        day_purchase = sales_purchase_dict[created_at]
        day_sell = sales_sell_dict.get(created_at, {})

        # ブランドごとに記入
        for sheet_alias, purchase_total in day_purchase.items():
            sell_total = int(day_sell.get(sheet_alias, 0))

            # B列にブランド（シート別名）を書きたい場合
            ws.cell(purchase_row, 2).value = sheet_alias
            ws.cell(sell_row, 2).value = sheet_alias
            ws.cell(profit_row, 2).value = sheet_alias
            ws.cell(purchase_row, 2).border = thin_border
            ws.cell(sell_row, 2).border = thin_border
            ws.cell(profit_row, 2).border = thin_border


            # 仕入（上段:Total / 下段:税込）
            ws.cell(purchase_row, 4).value = int(purchase_total)
            ws.cell(purchase_row + 1, 4).value = int(round(purchase_total * 1.1))
            ws.cell(purchase_row, 4).number_format = r'¥#,##0'
            ws.cell(purchase_row + 1, 4).number_format = r'¥#,##0'
            ws.cell(purchase_row, 4).border = thin_border
            ws.cell(purchase_row + 1, 4).border = thin_border

            # 売上（Totalのみ）
            ws.cell(sell_row, 4).value = int(sell_total)
            ws.cell(sell_row, 4).number_format = r'¥#,##0'
            ws.cell(sell_row, 4).border = thin_border

            # 利益（上段:金額 / 下段:利益率%）
            profit_amt = sell_total - purchase_total
            profit_rate = (profit_amt / sell_total) if sell_total else 0.0
            ws.cell(profit_row, 4).value = int(profit_amt)
            ws.cell(profit_row + 1, 4).value = float(profit_rate)
            ws.cell(profit_row, 4).number_format = r'¥#,##0'
            ws.cell(profit_row + 1, 4).number_format = '0.00%'
            ws.cell(profit_row, 4).border = thin_border
            ws.cell(profit_row + 1, 4).border = thin_border

            # 次の行へ
            purchase_row += 2
            sell_row += 1
            profit_row += 2

    col_count = len(sales_purchase_dict.keys())
    for col in range(4, 4 + col_count):
        for row in range(3, purchase_row+1):
            total_purchase = parse_money_to_int(ws.cell(row, col_count+4).value) or 0
            total_purchase += (ws.cell(row, col).value) or 0
            ws.cell(row, col_count+4).value = total_purchase
            ws.cell(row, col_count+4).number_format = r'¥#,##0'
            ws.cell(row, col_count+4).border = thin_border
        for row2 in range(97, sell_row+1):
            total_sell = parse_money_to_int(ws.cell(row2, col_count+4).value) or 0
            total_sell += parse_money_to_int(ws.cell(row2, col).value) or 0
            ws.cell(row2, col_count+4).value = total_sell
            ws.cell(row2, col_count+4).number_format = r'¥#,##0'
            ws.cell(row2, col_count+4).border = thin_border
        for row3 in range(147, profit_row+1,2):
            total_profit = parse_money_to_int(ws.cell(row3, col_count+4).value) or 0
            total_profit += parse_money_to_int(ws.cell(row3, col).value) or 0
            ws.cell(row3, col_count+4).value = total_profit
            ws.cell(row3, col_count+4).number_format = r'¥#,##0'
            ws.cell(row3, col_count+4).border = thin_border

            sell = parse_money_to_int(ws.cell(row3-50, col_count+4).value) or 0
            profit = parse_money_to_int(ws.cell(row3, col_count+4).value) or 0
            profit_rate = (profit / sell) if sell else 0.0
            ws.cell(row3+1, col_count+4).value = float(profit_rate)
            ws.cell(row3+1, col_count+4).number_format = '0.00%'
            ws.cell(row3+1, col_count+4).border = thin_border

    #NIPPONIKA
    rows = list(
        orders_between_by_batch_created("NIPPONIKA",start_dt, end_dt)
        .filter(batch__buyers="NIPPONIKATRADING")
        .values("Brand_name", "Amount", "仕入値合計", "batch__created_at")
        .order_by("id")
    )

    # --- 日付×ブランド(=テンプレ内のシート別名) で集計 ---
    # 例: sales_purchase_dict[date][sheet_alias] = 仕入合計
    sales_purchase_dict: Dict[datetime, Dict[str, int]] = defaultdict(lambda: defaultdict(int))
    sales_sell_dict: dict[datetime.datetime, dict[str, int]] = defaultdict(lambda: defaultdict(int))

    for r in rows:
        brand = (r["Brand_name"] or "").strip()
        created_at = r["batch__created_at"]  # datetime
        created_at = timezone.localtime(created_at)
        created_at = created_at.strftime("%Y/%m/%d")
        if not brand:
            continue

        sheet_alias = BRAND_TO_SHEET_ALIAS.get(brand)
        if not sheet_alias:
            logger.warning(f"ブランド名の割当が未定義: {brand}")
            continue

        amount = int(r["Amount"] or 0)
        purchase = int(r["仕入値合計"] or 0)

        sales_purchase_dict[created_at][sheet_alias] += purchase
        sales_sell_dict[created_at][sheet_alias] += amount

    # データが無ければエラーメッセージ or 空返し
    if not sales_sell_dict:
        logger.info("指定期間に売上データがありません。")
        messages.warning(request, "指定期間に売上データがありません。")
        return redirect("report_console")

    # --- テンプレート読込 ---
    #wb = openpyxl.load_workbook(SALES_TABLE_TEMPLATE_PATH, data_only=False)
    ws = wb["NIPPONIKA"]

    def safe_insert_blank_cols(ws_, col_idx: int, how_many: int = 1):
        """結合の影響で insert_cols が失敗しないようにしたい最低限のラッパ（テンプレに合わせてそのまま挿入）。"""
        ws_.insert_cols(col_idx, how_many)

    for created_at in sorted(sales_purchase_dict.keys(), reverse=True):
        safe_insert_blank_cols(ws, 4, 1)  # D列に1列挿入
        ws.cell(2, 4).value = created_at
        ws.cell(2, 4).border = thin_border

        purchase_row = 3
        sell_row = 97
        profit_row = 147

        # 同日のブランド列挙（テンプレの行順に合わせたい場合は、ここを SALES_TABLE_RC_MAP の順で回してください）
        day_purchase = sales_purchase_dict[created_at]
        day_sell = sales_sell_dict.get(created_at, {})

        # ブランドごとに記入
        for sheet_alias, purchase_total in day_purchase.items():
            sell_total = int(day_sell.get(sheet_alias, 0))

            # B列にブランド（シート別名）を書きたい場合
            ws.cell(purchase_row, 2).value = sheet_alias
            ws.cell(sell_row, 2).value = sheet_alias
            ws.cell(profit_row, 2).value = sheet_alias
            ws.cell(purchase_row, 2).border = thin_border
            ws.cell(sell_row, 2).border = thin_border
            ws.cell(profit_row, 2).border = thin_border


            # 仕入（上段:Total / 下段:税込）
            ws.cell(purchase_row, 4).value = int(purchase_total)
            ws.cell(purchase_row + 1, 4).value = int(round(purchase_total * 1.1))
            ws.cell(purchase_row, 4).number_format = r'¥#,##0'
            ws.cell(purchase_row + 1, 4).number_format = r'¥#,##0'
            ws.cell(purchase_row, 4).border = thin_border
            ws.cell(purchase_row + 1, 4).border = thin_border

            # 売上（Totalのみ）
            ws.cell(sell_row, 4).value = int(sell_total)
            ws.cell(sell_row, 4).number_format = r'¥#,##0'
            ws.cell(sell_row, 4).border = thin_border

            # 利益（上段:金額 / 下段:利益率%）
            profit_amt = sell_total - purchase_total
            profit_rate = (profit_amt / sell_total) if sell_total else 0.0
            ws.cell(profit_row, 4).value = int(profit_amt)
            ws.cell(profit_row + 1, 4).value = float(profit_rate)
            ws.cell(profit_row, 4).number_format = r'¥#,##0'
            ws.cell(profit_row + 1, 4).number_format = '0.00%'
            ws.cell(profit_row, 4).border = thin_border
            ws.cell(profit_row + 1, 4).border = thin_border

            # 次の行へ
            purchase_row += 2
            sell_row += 1
            profit_row += 2

    col_count = len(sales_purchase_dict.keys())
    for col in range(4, 4 + col_count):
        for row in range(3, purchase_row+1):
            total_purchase = parse_money_to_int(ws.cell(row, col_count+4).value) or 0
            total_purchase += parse_money_to_int(ws.cell(row, col).value) or 0
            ws.cell(row, col_count+4).value = total_purchase
            ws.cell(row, col_count+4).number_format = r'¥#,##0'
            ws.cell(row, col_count+4).border = thin_border
        for row2 in range(97, sell_row+1):
            total_sell = parse_money_to_int(ws.cell(row2, col_count+4).value) or 0
            total_sell += parse_money_to_int(ws.cell(row2, col).value) or 0
            ws.cell(row2, col_count+4).value = total_sell
            ws.cell(row2, col_count+4).number_format = r'¥#,##0'
            ws.cell(row2, col_count+4).border = thin_border
        for row3 in range(147, profit_row+1,2):
            total_profit = parse_money_to_int(ws.cell(row3, col_count+4).value) or 0
            total_profit += parse_money_to_int(ws.cell(row3, col).value) or 0
            ws.cell(row3, col_count+4).value = total_profit
            ws.cell(row3, col_count+4).number_format = r'¥#,##0'
            ws.cell(row3, col_count+4).border = thin_border

            sell = parse_money_to_int(ws.cell(row3-50, col_count+4).value) or 0
            profit = parse_money_to_int(ws.cell(row3, col_count+4).value) or 0
            profit_rate = (profit / sell) if sell else 0.0
            ws.cell(row3+1, col_count+4).value = float(profit_rate)
            ws.cell(row3+1, col_count+4).number_format = '0.00%'
            ws.cell(row3+1, col_count+4).border = thin_border

    #YAMATO/TOYO
    rows = list(
        orders_between_by_batch_created("YAMATO_TOYO",start_dt, end_dt)
        .filter(batch__buyers="YAMATO_TOYO")
        .values("Brand", "Amount_JPY", "総販売価格", "Quantity" , "batch__created_at")
        .order_by("id")
    )

    # --- 日付×ブランド(=テンプレ内のシート別名) で集計 ---
    # 例: sales_purchase_dict[date][sheet_alias] = 仕入合計
    sales_purchase_dict: Dict[datetime, Dict[str, int]] = defaultdict(lambda: defaultdict(int))
    sales_sell_dict: dict[datetime.datetime, dict[str, int]] = defaultdict(lambda: defaultdict(int))

    for r in rows:
        brand = (r["Brand"] or "").strip()
        created_at = r["batch__created_at"]  # datetime
        created_at = timezone.localtime(created_at)
        created_at = created_at.strftime("%Y/%m/%d")
        if not brand:
            continue

        amount = int(r["総販売価格"] or 0)
        purchase = int(r["Amount_JPY"] or 0)

        sales_purchase_dict[created_at][brand] += purchase
        sales_sell_dict[created_at][brand] += amount

    # データが無ければエラーメッセージ or 空返し
    if not sales_sell_dict:
        logger.info("指定期間に売上データがありません。")
        messages.warning(request, "指定期間に売上データがありません。")
        return redirect("report_console")

    # --- テンプレート読込 ---
    #wb = openpyxl.load_workbook(SALES_TABLE_TEMPLATE_PATH, data_only=False)
    ws = wb["YAMATO"]
    for key,val in sorted(sales_purchase_dict.items(), reverse=True):
        safe_insert_blank_cols(ws,4,1)
        ws.cell(2,4).value = key
        ws.cell(2,4).border = thin_border
        ws.cell(2,4).font = yamato_font

        p_total = 0
        s_total = 0
        for YAMATO_row in range(3,43,2):
            brand_key = ws.cell(YAMATO_row,2).value
            purchase_total = int(val.get(brand_key,0))
            sell_total = int(sales_sell_dict.get(key,{}).get(brand_key,0))

            ws.cell(YAMATO_row,4).value = int(purchase_total)
            ws.cell(YAMATO_row+1,4).value = int(round(purchase_total * 1.1))
            ws.cell(YAMATO_row,4).number_format = r'¥#,##0'
            ws.cell(YAMATO_row+1,4).number_format = r'¥#,##0'
            ws.cell(YAMATO_row,4).border = thin_border
            ws.cell(YAMATO_row+1,4).border = thin_border
            ws.cell(YAMATO_row,4).font = yamato_font
            ws.cell(YAMATO_row+1,4).font = yamato_font

            p_total += purchase_total
        ws.cell(43,4).value = int(p_total)
        ws.cell(43,4).number_format = r'¥#,##0'
        ws.cell(43,4).border = thin_border
        ws.cell(43,4).font = yamato_font

        for YAMATO_row in range(44,63):
            brand_key = ws.cell(YAMATO_row,2).value
            purchase_total = int(val.get(brand_key,0))
            sell_total = int(sales_sell_dict.get(key,{}).get(brand_key,0))

            ws.cell(YAMATO_row,4).value = int(sell_total)
            ws.cell(YAMATO_row,4).number_format = r'¥#,##0'
            ws.cell(YAMATO_row,4).border = thin_border
            ws.cell(YAMATO_row,4).font = yamato_font

            s_total += sell_total
        ws.cell(63,4).value = int(s_total)
        ws.cell(63,4).number_format = r'¥#,##0'
        ws.cell(63,4).border = thin_border
        ws.cell(63,4).font = yamato_font

        ws.cell(64,4).value = int(s_total - p_total)
        ws.cell(64,4).number_format = r'¥#,##0'
        ws.cell(64,4).border = thin_border
        ws.cell(64,4).font = yamato_font
        ws.cell(65,4).value = float((s_total - p_total) / s_total if s_total else 0.0)
        ws.cell(65,4).number_format = '0.00%'
        ws.cell(65,4).border = thin_border
        ws.cell(64,4).font = yamato_font

        ws.cell(70,4).value = int(s_total)
        ws.cell(70,4).number_format = r'¥#,##0'
        ws.cell(70,4).border = thin_border
        ws.cell(70,4).font = yamato_font

    col_count = len(sales_purchase_dict.keys())
    for col in range(3, 4 + col_count):
        for row in range(3,42):
            total_purchase = parse_money_to_int(ws.cell(row, col_count+4).value) or 0
            total_purchase += parse_money_to_int(ws.cell(row, col).value) or 0
            ws.cell(row, col_count+4).value = total_purchase
            ws.cell(row, col_count+4).number_format = r'¥#,##0'
            ws.cell(row, col_count+4).border = thin_border
            ws.cell(row, col_count+4).font = yamato_font
        for row2 in range(44,62):
            total_sell = parse_money_to_int(ws.cell(row2, col_count+4).value) or 0
            total_sell += parse_money_to_int(ws.cell(row2, col).value) or 0
            ws.cell(row2, col_count+4).value = total_sell
            ws.cell(row2, col_count+4).number_format = r'¥#,##0'
            ws.cell(row2, col_count+4).border = thin_border
            ws.cell(row2, col_count+4).font = yamato_font

    # --- 保存 & 返却 ---
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = TEMP_PATH / f"SELL_{label}_{ts}.xlsx"
    wb.save(out_path)

    return FileResponse(open(out_path, "rb"), as_attachment=True, filename=out_path.name)

def reports_ar_export(request):
    thin_border = Border(
        left=Side(style="thin"),   # 左
        right=Side(style="thin"),  # 右
        top=Side(style="thin"),    # 上
        bottom=Side(style="thin")  # 下
    )
    ar_font = Font(
        name="ＭＳ Ｐゴシック",   # フォント名（例: メイリオ）
        size=12,         # 文字サイズ
        bold=False,       # 太字
        italic=False,    # 斜体
        color="000000"   # 赤色（RGB指定）
    )

    start_dt, end_dt, fmt, label = _get_report_params_range(request)
    tz = timezone.get_current_timezone()
    start_dt = timezone.make_aware(datetime(1970, 1, 1, 0, 0, 0), tz)
    end_dt   = timezone.make_aware(datetime(2100, 1, 1, 0, 0, 0), tz)

    wb = openpyxl.load_workbook(AR_TEMPLATE_PATH, data_only=False)
    ws = wb["template"]

    #R&C
    RC_rows = list(
        orders_between_by_batch_created("ROYAL COSMETICS",start_dt, end_dt)
        .filter(batch__buyers="ROYAL COSMETICS")
        .values("Brand_name", "Amount", "Purchase_amount", "batch__created_at")
        .order_by("id")
    )
    #NIPPONIKA
    nipponika_rows = list(
        orders_between_by_batch_created("NIPPONIKA",start_dt, end_dt)
        .filter(batch__buyers="NIPPONIKATRADING")
        .values("Brand_name", "Amount", "仕入値合計", "batch__created_at")
        .order_by("id")
    )
    #YAMATO
    yamato_rows = list(
        orders_between_by_batch_created("YAMATO_TOYO",start_dt, end_dt)
        .filter(batch__buyers="YAMATO_TOYO")
        .values("Brand", "Amount_JPY", "総販売価格", "Quantity" , "batch__created_at")
        .order_by("id")
    )

    # --- 日付×ブランド で集計（キーは "YYYY/MM" へ統一） ---
    RC_sales_purchase_dict = defaultdict(lambda: defaultdict(int))
    RC_sales_sell_dict     = defaultdict(lambda: defaultdict(int))
    NIPPONIKA_sales_purchase_dict = defaultdict(lambda: defaultdict(int))
    NIPPONIKA_sales_sell_dict     = defaultdict(lambda: defaultdict(int))
    YAMATO_sales_purchase_dict = defaultdict(lambda: defaultdict(int))
    YAMATO_sales_sell_dict     = defaultdict(lambda: defaultdict(int))
    RC_pay_dict = defaultdict(int)          # { "YYYY/MM": number }
    NIPPONIKA_pay_dict = defaultdict(int)   # { "YYYY/MM": number }
    YAMATO_pay_dict = defaultdict(int)      # { "YYYY/MM": number }
    month_list = []  # ★ "YYYY/MM" の文字列だけ入れる

    for r in RC_rows:
        brand = (r["Brand_name"] or "").strip()
        created_at = timezone.localtime(r["batch__created_at"])  # aware dt → ローカル
        created_ym = created_at.strftime("%Y/%m")                # ★ 文字列キー

        month_list.append(created_ym)                            # ★ dt ではなく文字列
        if not brand:
            continue
        sheet_alias = BRAND_TO_SHEET_ALIAS.get(brand)
        if not sheet_alias:
            logger.warning(f"ブランド名の割当が未定義: {brand}")
            continue

        amount = int(r["Amount"] or 0)
        purchase = int(r["Purchase_amount"] or 0)

        RC_sales_purchase_dict[created_ym]["RC"] += purchase
        RC_sales_sell_dict[created_ym]["RC"] += amount

    for r in nipponika_rows:
        brand = (r["Brand_name"] or "").strip()
        created_at = timezone.localtime(r["batch__created_at"])
        created_ym = created_at.strftime("%Y/%m")                # ★

        month_list.append(created_ym)                            # ★
        if not brand:
            continue
        sheet_alias = BRAND_TO_SHEET_ALIAS.get(brand)
        if not sheet_alias:
            logger.warning(f"ブランド名の割当が未定義: {brand}")
            continue

        amount = int(r["Amount"] or 0)
        purchase = int(r["仕入値合計"] or 0)

        NIPPONIKA_sales_purchase_dict[created_ym]["NIPPONIKA"] += purchase
        NIPPONIKA_sales_sell_dict[created_ym]["NIPPONIKA"] += amount

    for r in yamato_rows:
        brand = (r["Brand"] or "").strip()
        created_at = timezone.localtime(r["batch__created_at"])
        created_ym = created_at.strftime("%Y/%m")                # ★

        month_list.append(created_ym)                            # ★
        if not brand:
            continue
        sheet_alias = brand
        if not sheet_alias:
            logger.warning(f"ブランド名の割当が未定義: {brand}")
            continue

        amount = int(r["総販売価格"] or 0)
        purchase = int(r["Amount_JPY"] or 0)

        YAMATO_sales_purchase_dict[created_ym]["YAMATO"] += purchase
        YAMATO_sales_sell_dict[created_ym]["YAMATO"] += amount

    month_list = sorted(set(month_list))                         # ★ 文字列で時系列ソート

    for month in month_list:  # ★ month は "YYYY/MM"（文字列）

        #和暦計算
        base_dt = datetime.strptime(month, "%Y/%m")
        d = date(base_dt.year, base_dt.month, 1)
        if d >= date(2019, 5, 1):
            era = "令和"
            y = d.year - 2018      # 2019→1年
        elif d >= date(1989, 1, 8):
            era = "平成"
            y = d.year - 1988
        elif d >= date(1926, 12, 25):
            era = "昭和"
            y = d.year - 1925
        elif d >= date(1912, 7, 30):
            era = "大正"
            y = d.year - 1911
        else:
            era = "明治"
            y = d.year - 1867
        year_str = "元" if y == 1 else str(y)
        wareki_month = f"{era}{year_str}年{d.month}月"

        safe_insert_blank_cols(ws, 5, 7)

        for col in range(5,12):
            for row in range(6,17):
                ws.cell(row,col).border = thin_border
                ws.cell(row,col).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row,col).font = ar_font

        ws.column_dimensions['E'].width = 14
        ws.column_dimensions['F'].width = 14
        ws.column_dimensions['G'].width = 14
        ws.column_dimensions['H'].width = 14
        ws.column_dimensions['I'].width = 14
        ws.column_dimensions['J'].width = 14
        ws.column_dimensions['K'].width = 14

        ws.cell(3,5).value = wareki_month
        ws.merge_cells(start_row=3, start_column=5, end_row=3, end_column=11)
        ws.cell(4,5).value = "当月売上高"
        ws.merge_cells(start_row=4, start_column=5, end_row=5, end_column=5)
        ws.cell(4,6).value = "入金"
        ws.merge_cells(start_row=4, start_column=6, end_row=4, end_column=10)
        ws.cell(5,6).value = "入金①"
        ws.cell(5,7).value = "入金②"
        ws.cell(5,8).value = "入金③"
        ws.cell(5,9).value = "相殺"
        ws.cell(5,10).value = "入金合計"
        ws.cell(4,11).value = "残高"
        ws.merge_cells(start_row=4, start_column=11, end_row=5, end_column=11)
        for r in [(3,5),(3,6),(3,7),(3,8),(3,9),(3,10),(3,11),(4,5),(4,6),(4,7),(4,8),(4,9),(4,10),(5,5),(4,6),(5,6),(5,7),(5,8),(5,9),(5,10),(4,11),(5,11)]:
            ws.cell(*r).border = thin_border
            ws.cell(*r).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(*r).font = ar_font

        # === R&C ===
        sell_total = RC_sales_sell_dict.get(month, {}).get("RC", 0)
        ws.cell(6,5).value = int(sell_total)
        ws.cell(6,5).number_format = r'¥#,##0'
        ws.cell(6,5).border = thin_border
        ws.cell(6,5).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(6,5).font = ar_font

        ws.cell(6,6).value = int(sell_total) / 2
        ws.cell(6,6).number_format = r'¥#,##0'
        ws.cell(6,6).border = thin_border
        ws.cell(6,6).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(6,6).font = ar_font

        if month in RC_pay_dict:
            ws.cell(6,7).value = int(RC_pay_dict[month])
            ws.cell(6,7).number_format = r'¥#,##0'
            ws.cell(6,7).border = thin_border
            ws.cell(6,7).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(6,7).font = ar_font

        # 入金合計（当月半金 + 過去からの入金②）
        ws.cell(6,10).value = int((sell_total / 2) + RC_pay_dict.get(month, 0))
        ws.cell(6,10).number_format = r'¥#,##0'
        ws.cell(6,10).border = thin_border
        ws.cell(6,10).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(6,10).font = ar_font

        # ★ ここで month("YYYY/MM") を dt に一時的に変換し、+3ヶ月して再び "YYYY/MM" へ
        base_dt = datetime.strptime(month, "%Y/%m")              # ★
        pay_month_plus3 = (base_dt.replace(day=1) + relativedelta(months=+3)).strftime("%Y/%m")  # ★
        RC_pay_dict[pay_month_plus3] += (sell_total / 2)         # ★ そのまま += でOK

        # 残高（その月までの入金②の累計）…文字列比較でOK
        balance = 0
        for key, val in RC_pay_dict.items():                     # key も "YYYY/MM"
            if key >= month:
                balance += val
        ws.cell(6,11).value = int(balance)
        ws.cell(6,11).number_format = r'¥#,##0'
        ws.cell(6,11).border = thin_border
        ws.cell(6,11).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(6,11).font = ar_font

        # === NIPPONIKA ===
        sell_total = NIPPONIKA_sales_sell_dict.get(month, {}).get("NIPPONIKA", 0)
        ws.cell(10,5).value = int(sell_total)
        ws.cell(10,5).number_format = r'¥#,##0'
        ws.cell(10,5).border = thin_border
        ws.cell(10,5).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(10,5).font = ar_font

        ws.cell(10,6).value = int(sell_total) / 2
        ws.cell(10,6).number_format = r'¥#,##0'
        ws.cell(10,6).border = thin_border
        ws.cell(10,6).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(10,6).font = ar_font

        if month in NIPPONIKA_pay_dict:
            ws.cell(10,7).value = int(NIPPONIKA_pay_dict[month])
            ws.cell(10,7).number_format = r'¥#,##0'
            ws.cell(10,7).border = thin_border
            ws.cell(10,7).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(10,7).font = ar_font

        # ★ 同様に3ヶ月後キーを関数外で直接計算
        base_dt = datetime.strptime(month, "%Y/%m")              # ★
        pay_month_plus3 = (base_dt.replace(day=1) + relativedelta(months=+3)).strftime("%Y/%m")  # ★
        NIPPONIKA_pay_dict[pay_month_plus3] += (sell_total / 2)  # ★

        balance = 0
        for key, val in NIPPONIKA_pay_dict.items():
            if key >= month:
                balance += val
        ws.cell(10,11).value = int(balance)
        ws.cell(10,11).number_format = r'¥#,##0'
        ws.cell(10,11).border = thin_border
        ws.cell(10,11).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(10,11).font = ar_font

        # === YAMATO_TOYO ===
        sell_total = YAMATO_sales_sell_dict.get(month, {}).get("YAMATO", 0)
        ws.cell(8,5).value = int(sell_total)
        ws.cell(8,5).number_format = r'¥#,##0'
        ws.cell(8,5).border = thin_border
        ws.cell(8,5).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(8,5).font = ar_font
        ws.cell(8,6).value = int(sell_total) / 2
        ws.cell(8,6).number_format = r'¥#,##0'
        ws.cell(8,6).border = thin_border
        ws.cell(8,6).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(8,6).font = ar_font
        if month in YAMATO_pay_dict:
            ws.cell(8,7).value = int(YAMATO_pay_dict[month])
            ws.cell(8,7).number_format = r'¥#,##0'
            ws.cell(8,7).border = thin_border
            ws.cell(8,7).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(8,7).font = ar_font
        # ★ 同様に3ヶ月後キーを関数外で直接計算
        base_dt = datetime.strptime(month, "%Y/%m")              # ★
        pay_month_plus4 = (base_dt.replace(day=1) + relativedelta(months=+4)).strftime("%Y/%m")  # ★
        YAMATO_pay_dict[pay_month_plus4] += (sell_total / 2)     # ★
        balance = 0
        for key, val in YAMATO_pay_dict.items():
            if key >= month:
                balance += val
        ws.cell(8,11).value = int(balance)
        ws.cell(8,11).number_format = r'¥#,##0'
        ws.cell(8,11).border = thin_border
        ws.cell(8,11).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(8,11).font = ar_font

        # 合計行（16行目）
        for col in range(5,13):
            total = 0
            for row in [6,8,10]:
                v = ws.cell(row, col).value
                if v:
                    total += int(v)
            ws.cell(16, col).value = int(total)
            ws.cell(16, col).number_format = r'¥#,##0'
            ws.cell(16, col).border = thin_border
            ws.cell(16, col).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(16, col).font = ar_font

    # --- 保存 & 返却 ---
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = TEMP_PATH / f"AR_{label}_{ts}.xlsx"
    wb.save(out_path)
    return FileResponse(open(out_path, "rb"), as_attachment=True, filename=out_path.name)

def reports_ap_export(request):
    thin_border = Border(
        left=Side(style="thin"),   # 左
        right=Side(style="thin"),  # 右
        top=Side(style="thin"),    # 上
        bottom=Side(style="thin")  # 下
    )

    start_dt, end_dt, fmt, label = _get_report_params_range(request)
    tz = timezone.get_current_timezone()
    start_dt = timezone.make_aware(datetime(1970, 1, 1, 0, 0, 0), tz)
    end_dt   = timezone.make_aware(datetime(2100, 1, 1, 0, 0, 0), tz)

    wb = openpyxl.load_workbook(AP_TEMPLATE_PATH, data_only=False)
    ws = wb["template"]

    #R&C
    RC_rows = list(
        orders_between_by_batch_created("ROYAL COSMETICS",start_dt, end_dt)
        .filter(batch__buyers="ROYAL COSMETICS")
        .values("Brand_name", "Amount", "Purchase_amount", "batch__created_at")
        .order_by("id")
    )
    #NIPPONIKA
    nipponika_rows = list(
        orders_between_by_batch_created("NIPPONIKA",start_dt, end_dt)
        .filter(batch__buyers="NIPPONIKATRADING")
        .values("Brand_name", "Amount", "仕入値合計", "batch__created_at")
        .order_by("id")
    )
    #YAMATO
    yamato_rows = list(
        orders_between_by_batch_created("YAMATO_TOYO",start_dt, end_dt)
        .filter(batch__buyers="YAMATO_TOYO")
        .values("Brand", "Amount_JPY", "総販売価格", "Quantity" , "batch__created_at")
        .order_by("id")
    )

    # --- 日付×ブランド(=テンプレ内のシート別名) で集計 ---
    # 例: sales_purchase_dict[date][sheet_alias] = 仕入合計
    sales_purchase_dict: Dict[datetime, Dict[str, int]] = defaultdict(lambda: defaultdict(int))
    sales_sell_dict: Dict[datetime, Dict[str, int]] = defaultdict(lambda: defaultdict(int))
    zandaka_dict: Dict[datetime, Dict[str, int]] = defaultdict(lambda: defaultdict(int))
    next_dict: Dict[datetime, Dict[str, int]] =  defaultdict(lambda: defaultdict(int))

    for r in RC_rows:
        brand = (r["Brand_name"] or "").strip()
        created_at = r["batch__created_at"]  # datetime
        created_at = timezone.localtime(created_at)
        created_at = created_at.strftime("%Y/%m")
        if not brand:
            continue

        sheet_alias = BRAND_TO_SHEET_ALIAS.get(brand)
        if not sheet_alias:
            logger.warning(f"ブランド名の割当が未定義: {brand}")
            continue

        amount = int(r["Amount"] or 0)
        purchase = int(r["Purchase_amount"] or 0)

        sales_purchase_dict[created_at][sheet_alias] += purchase
        sales_sell_dict[created_at][sheet_alias] += amount

    for r in nipponika_rows:
        brand = (r["Brand_name"] or "").strip()
        created_at = r["batch__created_at"]  # datetime
        created_at = timezone.localtime(created_at)
        created_at = created_at.strftime("%Y/%m")
        if not brand:
            continue

        sheet_alias = BRAND_TO_SHEET_ALIAS.get(brand)
        if not sheet_alias:
            logger.warning(f"ブランド名の割当が未定義: {brand}")
            continue

        amount = int(r["Amount"] or 0)
        purchase = int(r["仕入値合計"] or 0)

        sales_purchase_dict[created_at][sheet_alias] += purchase
        sales_sell_dict[created_at][sheet_alias] += amount

    for r in yamato_rows:
        brand = (r["Brand"] or "").strip()
        created_at = r["batch__created_at"]  # datetime
        created_at = timezone.localtime(created_at)
        created_at = created_at.strftime("%Y/%m")
        if not brand:
            continue

        sheet_alias = brand
        if not sheet_alias:
            logger.warning(f"ブランド名の割当が未定義: {brand}")
            continue

        amount = int(r["総販売価格"] or 0)
        purchase = int(r["Amount_JPY"] or 0)

        sales_purchase_dict[created_at][sheet_alias] += purchase
        sales_sell_dict[created_at][sheet_alias] += amount

    for ym_key,val in sorted(sales_purchase_dict.items(), reverse=False):
        safe_insert_blank_cols(ws, 6, 5)  # D列に1列挿入
        ws.column_dimensions['F'].width = 14
        ws.column_dimensions['G'].width = 14
        ws.column_dimensions['H'].width = 14
        ws.column_dimensions['I'].width = 14
        ws.column_dimensions['J'].width = 14
        dt = datetime.strptime(ym_key, "%Y/%m")  # ← datetime 化
        m = dt.strftime("%Y%m")

        base = dt.replace(day=1, hour=0, minute=0, second=0, microsecond=0)

        dt_plus1 = base + relativedelta(months=+1)  # 1か月後(翌月1日)
        dt_plus2 = base + relativedelta(months=+2)  # 2か月後
        dt_plus5 = base + relativedelta(months=+5)  # 5か月後

        m_plus1 = dt_plus1.strftime("%Y%m")
        m_plus2 = dt_plus2.strftime("%Y%m")
        m_plus5 = dt_plus5.strftime("%Y%m")

        ws.cell(5,6).value = f"{m}月仕入高"
        ws.merge_cells(start_row=5, start_column=6, end_row=5, end_column=8)
        ws.cell(5,6).border = thin_border
        ws.cell(5,7).border = thin_border
        ws.cell(5,8).border = thin_border
        ws.cell(5,6).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(5,9).value = f"{m}月決済高"
        ws.cell(5,10).value = f"{m}月残高"
        ws.cell(5,9).border = thin_border
        ws.cell(5,10).border = thin_border
        ws.cell(6,9).border = thin_border
        ws.cell(6,10).border = thin_border
        ws.cell(6,6).value = f"税込"
        ws.cell(6,7).value = f"税抜"
        ws.cell(6,8).value = f"消費税"
        ws.cell(6,6).border = thin_border
        ws.cell(6,7).border = thin_border
        ws.cell(6,8).border = thin_border

        for row in range(7,66):
            template_brand = ws.cell(row,3).value
            pay_conditions = ws.cell(row,4).value

            ws.cell(row,6).border = thin_border
            ws.cell(row,7).border = thin_border
            ws.cell(row,8).border = thin_border
            ws.cell(row,9).border = thin_border
            ws.cell(row,10).border = thin_border
            if template_brand is None or pay_conditions is None:
                logger.warning(f"テンプレの3列目or4列目が空欄です。行={row}")
                continue

            if "," in template_brand:
                template_brand_list = template_brand.split(",")
            else:
                template_brand_list = [template_brand]
            for tb in template_brand_list:
                for brand_alias in AP_BRAND_MAP:
                    if tb == brand_alias:
                        sheet_alias = AP_BRAND_MAP[brand_alias]
                        purchase_total = val.get(sheet_alias,0)
                        ws.cell(row,6).value = int(purchase_total)
                        ws.cell(row,6).number_format = r'¥#,##0'
                        ws.cell(row,6).border = thin_border
                        ws.cell(row,7).value = int(round(purchase_total / 1.1))
                        ws.cell(row,7).number_format = r'¥#,##0'
                        ws.cell(row,7).border = thin_border
                        ws.cell(row,8).value = int(purchase_total - int(round(purchase_total / 1.1)))
                        ws.cell(row,8).number_format = r'¥#,##0'
                        ws.cell(row,8).border = thin_border
                        if pay_conditions == "前払い":
                            ws.cell(row,9) .value = int(purchase_total) + next_dict[sheet_alias].get(m,0)
                            ws.cell(row,9).number_format = r'¥#,##0'
                            ws.cell(row,9).border = thin_border
                            ws.cell(row,10).value = next_dict[sheet_alias].get("zandaka",0)
                            ws.cell(row,10).number_format = r'¥#,##0'
                            ws.cell(row,10).border = thin_border
                        elif pay_conditions == "月末締め、翌月末払い":
                            next_dict[sheet_alias][m_plus1] += int(purchase_total)
                            next_dict[sheet_alias]["zandaka"] += int(purchase_total)

                            ws.cell(row,9) .value = int(purchase_total) + next_dict[sheet_alias].get(m,0)
                            ws.cell(row,9).number_format = r'¥#,##0'
                            ws.cell(row,9).border = thin_border
                            ws.cell(row,10).value = next_dict[sheet_alias].get("zandaka",0)
                            ws.cell(row,10).number_format = r'¥#,##0'
                            ws.cell(row,10).border = thin_border
                        elif pay_conditions == "月末締め、翌々月の25日払い":
                            next_dict[sheet_alias][m_plus2] += int(purchase_total)
                            next_dict[sheet_alias]["zandaka"] += int(purchase_total)

                            ws.cell(row,9) .value = int(purchase_total) + next_dict[sheet_alias].get(m,0)
                            ws.cell(row,9).number_format = r'¥#,##0'
                            ws.cell(row,9).border = thin_border
                            ws.cell(row,10).value = next_dict[sheet_alias].get("zandaka",0)
                            ws.cell(row,10).number_format = r'¥#,##0'
                            ws.cell(row,10).border = thin_border
                        elif pay_conditions == "月末締め、5か月後払い":
                            next_dict[sheet_alias][m_plus5] += int(purchase_total)
                            next_dict[sheet_alias]["zandaka"] += int(purchase_total)

                            ws.cell(row,9) .value = int(purchase_total) + next_dict[sheet_alias].get(m,0)
                            ws.cell(row,9).number_format = r'¥#,##0'
                            ws.cell(row,9).border = thin_border
                            ws.cell(row,10).value = next_dict[sheet_alias].get("zandaka",0)
                            ws.cell(row,10).number_format = r'¥#,##0'
                            ws.cell(row,10).border = thin_border
            ws.cell(row,5).value = next_dict[sheet_alias].get("zandaka",0)
            ws.cell(row,5).number_format = r'¥#,##0'
            ws.cell(row,5).border = thin_border

        total_f = 0
        total_g = 0
        total_h = 0
        total_i = 0
        total_j = 0
        for total_row in range(7,66):
            if ws.cell(total_row,6).value is not None:
                total_f += int(ws.cell(total_row,6).value)
            if ws.cell(total_row,7).value is not None:
                total_g += int(ws.cell(total_row,7).value)
            if ws.cell(total_row,8).value is not None:
                total_h += int(ws.cell(total_row,8).value)
            if ws.cell(total_row,9).value is not None:
                total_i += int(ws.cell(total_row,9).value)
            if ws.cell(total_row,10).value is not None:
                total_j += int(ws.cell(total_row,10).value)
        ws.cell(66,6).value = int(total_f)
        ws.cell(66,6).number_format = r'¥#,##0'
        ws.cell(66,6).border = thin_border
        ws.cell(66,7).value = int(total_g)
        ws.cell(66,7).number_format = r'¥#,##0'
        ws.cell(66,7).border = thin_border
        ws.cell(66,8).value = int(total_h)
        ws.cell(66,8).number_format = r'¥#,##0'
        ws.cell(66,8).border = thin_border
        ws.cell(66,9).value = int(total_i)
        ws.cell(66,9).number_format = r'¥#,##0'
        ws.cell(66,9).border = thin_border
        ws.cell(66,10).value = int(total_j)
        ws.cell(66,10).number_format = r'¥#,##0'
        ws.cell(66,10).border = thin_border
    
    total_e = 0
    for total_row in range(7,66):
        if ws.cell(total_row,5).value is not None:
            total_e += int(ws.cell(total_row,5).value)
    ws.cell(66,5).value = int(total_e)
    ws.cell(66,5).number_format = r'¥#,##0'
    ws.cell(66,5).border = thin_border

    # --- 保存 & 返却 ---
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = TEMP_PATH / f"AP_{label}_{ts}.xlsx"
    wb.save(out_path)
    return FileResponse(open(out_path, "rb"), as_attachment=True, filename=out_path.name)

def reports_cashflow_export(request):
    from dateutil.relativedelta import relativedelta  # 関数内インポート（新規外部関数は作成しません）

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    cashflow_font = Font(
        name="ＭＳ Ｐゴシック",
        size=22,
        bold=False,
        italic=False,
        color="000000",
    )

    # テンプレ読込
    wb = openpyxl.load_workbook(FORCAST_TEMPLATE_PATH, data_only=False)
    ws = wb["template"]

    # 期間条件
    start_dt, end_dt, fmt, label = _get_report_params_range(request)

    # オーダー取得
    RC_rows = list(
        orders_between_by_batch_created("ROYAL COSMETICS", start_dt, end_dt)
        .filter(batch__buyers="ROYAL COSMETICS")
        .values("Brand_name", "Amount", "Purchase_amount", "batch__created_at")
        .order_by("id")
    )
    nipponika_rows = list(
        orders_between_by_batch_created("NIPPONIKA", start_dt, end_dt)
        .filter(batch__buyers="NIPPONIKATRADING")
        .values("Brand_name", "Amount", "仕入値合計", "batch__created_at")
        .order_by("id")
    )
    yamato_rows = list(
        orders_between_by_batch_created("YAMATO_TOYO", start_dt, end_dt)
        .filter(batch__buyers="YAMATO_TOYO")
        .values("Brand", "Amount_JPY", "総販売価格", "Quantity", "batch__created_at")
        .order_by("id")
    )

    # 月×シート別名で仕入合計
    month_purchase = defaultdict(lambda: defaultdict(int))  # {date(当月1日): {alias: amount}}

    # APテンプレから支払条件（オフセット月）を取得
    ap_wb = openpyxl.load_workbook(AP_TEMPLATE_PATH, data_only=True)
    ap_ws = ap_wb["template"]
    brand_month_offset: Dict[str, int] = {}  # {sheet_alias: month_offset}
    for r in range(7, 67):  # C列=3, D列=4
        names = ap_ws.cell(r, 3).value
        cond = ap_ws.cell(r, 4).value
        if not names or not cond:
            continue
        cond = str(cond).strip()
        if cond == "前払い":
            offset = 0
        elif cond == "月末締め、翌月末払い":
            offset = 1
        elif cond == "月末締め、翌々月の25日払い":
            offset = 2
        elif cond == "月末締め、5か月後払い":
            offset = 5
        else:
            offset = 0  # デフォルト同月

        for nm in (str(names).split(",") if "," in str(names) else [str(names)]):
            nm = nm.strip()
            if nm in AP_BRAND_MAP:
                sheet_alias = AP_BRAND_MAP[nm]  # C列名→シート別名
                brand_month_offset[sheet_alias] = offset

    buyers_month_offset = {
        "ROYAL COSMETICS": 3,
        "NIPPONIKATRADING": 3,
        "YAMATO_TOYO": 4,
    }
    scheduled_in = defaultdict(lambda: defaultdict(int)) 

    # 仕入集計（YAMATOはブランド名→シート別名へ正規化）
    def fold(rows, amount_field, buyer_code):
        for r in rows:
            if amount_field == "Amount_JPY":
                brand = (r["Brand"] or "").strip()
                if not brand:
                    continue
                alias = AP_BRAND_MAP.get(brand, brand)  # YAMATOもAP側の別名に正規化
            else:
                brand = (r["Brand_name"] or "").strip()
                if not brand:
                    continue
                alias = BRAND_TO_SHEET_ALIAS.get(brand)
                if not alias:
                    continue

            dt = r["batch__created_at"]
            if isinstance(dt, str):
                try:
                    dt = dateparser.parse(dt)
                except Exception:
                    continue
            if timezone.is_aware(dt):
                dt = timezone.localtime(dt)

            # 当月1日に正規化
            ym_key = date(dt.year, dt.month, 1)

            # --- 仕入金額の集計（支出側） ---
            amount = r.get(amount_field)
            try:
                amount_i = int(amount or 0)
            except Exception:
                try:
                    amount_i = int(str(amount).replace(",", ""))
                except Exception:
                    amount_i = 0
            month_purchase[ym_key][alias] += amount_i

            # --- 入金予定（当月50%＋オフセット後50%） ---
            # 売上金額（RC/NIPPONIKA=Amount, YAMATO=総販売価格）
            sales_val = r.get("総販売価格") if amount_field == "Amount_JPY" else r.get("Amount")
            try:
                sales_i = int(sales_val or 0)
            except Exception:
                try:
                    sales_i = int(str(sales_val).replace(",", ""))
                except Exception:
                    sales_i = 0

            if sales_i:
                half_now = sales_i // 2                # 奇数円は後払い側へ寄せる
                half_later = sales_i - half_now
                # 当月入金
                scheduled_in[ym_key][buyer_code] += half_now
                # オフセット後入金
                off = buyers_month_offset.get(buyer_code, 0)
                pay_month = ym_key + relativedelta(months=+off)
                scheduled_in[pay_month][buyer_code] += half_later

    fold(RC_rows,        "Purchase_amount", "ROYAL COSMETICS")
    fold(nipponika_rows, "仕入値合計",       "NIPPONIKATRADING")
    fold(yamato_rows,    "Amount_JPY",       "YAMATO_TOYO")

    # 仕入月→支払月へ振替
    scheduled_out = defaultdict(lambda: defaultdict(int))  # {date(支払月1日): {alias: amount}}
    for ym_key, brands in month_purchase.items():
        base_first = date(ym_key.year, ym_key.month, 1)
        for alias, amt in brands.items():
            off = brand_month_offset.get(alias, 0)
            pay_first = base_first + relativedelta(months=+off)
            scheduled_out[pay_first][alias] += int(amt)

    # フォールバック（万一何も出なければ同月出力）
    if not scheduled_out:
        for ym_key, brands in month_purchase.items():
            for alias, amt in brands.items():
                scheduled_out[ym_key][alias] += int(amt)

    # Excel 出力
    cashflow_row = 5
    DATE_COL = 1
    TITLE_COL = 2
    INCOME_COL = 3
    EXPENSE_COL = 4

    all_months = sorted(set(scheduled_out.keys()) | set(scheduled_in.keys()), reverse=False)

    for ym_key in all_months:
        header_date = ym_key  # 当月1日
        # 行の先頭セル（A列）に日付を入れ、以降この月の行が続く
        ws.cell(cashflow_row, DATE_COL).value = header_date
        ws.cell(cashflow_row, DATE_COL).number_format = "yyyy/mm/dd"
        ws.cell(cashflow_row, DATE_COL).font = cashflow_font
        ws.cell(cashflow_row, DATE_COL).border = thin_border

        wrote_any = False

        # 1) 入金の書き出し（バイヤー別）
        if ym_key in scheduled_in:
            for buyer, amt in sorted(scheduled_in[ym_key].items()):
                if int(amt) == 0:
                    continue
                # タイトル
                ws.cell(cashflow_row, TITLE_COL).value = f"{buyer} {ym_key.month:02d}月度 入金予定"
                # 入金列
                ic = ws.cell(cashflow_row, INCOME_COL)
                ic.value = int(amt)
                ic.number_format = r'¥#,##0'

                # 体裁
                for col in (TITLE_COL, INCOME_COL, EXPENSE_COL):
                    ws.cell(cashflow_row, col).font = cashflow_font
                    ws.cell(cashflow_row, col).border = thin_border
                # A列の日付は最初の行にだけ入れ、同月の2行目以降は空白でもOK
                wrote_any = True
                cashflow_row += 1

        # 2) 支出の書き出し（ブランド別）
        if ym_key in scheduled_out:
            for alias, amt in sorted(scheduled_out[ym_key].items()):
                if int(amt) == 0:
                    continue
                # 同月でまだ何も書いていない場合はA列に日付を入れる
                if not wrote_any:
                    ws.cell(cashflow_row, DATE_COL).value = header_date
                    ws.cell(cashflow_row, DATE_COL).number_format = "yyyy/mm/dd"
                    ws.cell(cashflow_row, DATE_COL).font = cashflow_font
                    ws.cell(cashflow_row, DATE_COL).border = thin_border
                    wrote_any = True

                # タイトル
                ws.cell(cashflow_row, TITLE_COL).value = f"{alias} {ym_key.month:02d}月度 前払い予定"
                # 支出列
                ec = ws.cell(cashflow_row, EXPENSE_COL)
                ec.value = int(amt)
                ec.number_format = r'¥#,##0'

                # 体裁
                for col in (TITLE_COL, INCOME_COL, EXPENSE_COL):
                    ws.cell(cashflow_row, col).font = cashflow_font
                    ws.cell(cashflow_row, col).border = thin_border

                cashflow_row += 1

        # 月ごとに1行空ける（その月に何か1件でも書いた場合のみ）
        if wrote_any:
            cashflow_row += 1

    # 保存 & 返却
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = TEMP_PATH / f"CASHFLOW_{label}_{ts}.xlsx"
    wb.save(out_path)
    return FileResponse(open(out_path, "rb"), as_attachment=True, filename=out_path.name)

#以下商品情報管理
VENDOR_CONFIG = {
    "yamato_toyo": {
        "label": "YAMATO/TOYO TRADING社向け",
        "model": YAMATO_TOYO_ProductInfo,  # ← 実在モデルに置き換え済み想定
        "list_columns": [
            {"field": "id",         "label": "ID",           "width": "100px", "mono": True, "link_to_detail": True},
            {"field": "Order_Code",   "label": "Order_Code",     "width": "180px", "mono": True},
            {"field": "Brand", "label": "Brand"},
            {"field": "Item_Name", "label": "Item_Name"},
            {"field": "updated_at", "label": "更新",          "width": "160px", "format": "datetime", "mono": True},
        ],
        "detail_columns": [
            {"field": "Order_Code", "label": "Order_Code", "mono": True},
            {"field": "Brand","label": "Brand"},
            {"field": "Item_Name","label": "Item_Name"},
            {"field": "Unit_price_JPY","label": "Unit_price"},
            {"field": "販売価格","label": "販売価格","mono": True},
            {"field": "updated_at","label": "更新","format": "datetime","mono": True},
            {"field": "RequiredProduct_flg", "label": "現地登録必要商品"},
        ],
    },
    "royal_cosmetics": {
        "label": "ROYAL COSMETICS社向け",
        "model": RY_ProductInfo,
        "list_columns": [
            {"field": "id",         "label": "ID",           "width": "100px", "mono": True, "link_to_detail": True},
            {"field": "Jan_code",   "label": "Jan code",     "width": "180px", "mono": True},
            {"field": "Brand_name", "label": "Brand name"},
            {"field": "Product_name",      "label": "Product_name" ,"width": "180px"},
            {"field": "updated_at", "label": "更新",          "width": "160px", "format": "datetime", "mono": True},
        ],
        "detail_columns": [
            {"field": "Jan_code","label": "Jan code","mono": True},
            {"field": "Brand_name","label": "Brand name"},
            {"field": "Product_name","label": "Product_name" ,"width": "180px"},
            {"field": "Contents","label": "Contents"},
            {"field": "Unit_price","label": "Unit price","mono": True},
            {"field": "Amount","label": "Amount","mono": True},
            {"field": "成分","label": "成分","multiline": True},
            {"field": "updated_at","label": "更新","format": "datetime","mono": True},
            {"field": "RequiredProduct_flg", "label": "現地登録必要商品"},
        ],
    },
    "nipponikatrading": {
        "label": "NIPPONIKATRADING社向け",
        "model": NIPPONIKATRADING_ProductInfo,
        "list_columns": [
            {"field": "id",         "label": "ID",           "width": "100px", "mono": True, "link_to_detail": True},
            {"field": "Jan_code",   "label": "Jan code",     "width": "180px", "mono": True},
            {"field": "Brand_name", "label": "Brand name"},
            {"field": "Description_of_goods","label": "Description of goods"},
            {"field": "Contents",   "label": "Contents",     "width": "120px"},
            {"field": "updated_at", "label": "更新",          "width": "160px", "format": "datetime", "mono": True},
        ],
        "detail_columns": [
            {"field": "HS_CODE","label": "HS CODE","mono": True},
            {"field": "Jan_code","label": "Jan code","mono": True},
            {"field": "Артикул","label": "Артикул"},
            {"field": "Brand_name","label": "Brand name"},
            {"field": "日本語名","label": "日本語名"},
            {"field": "Description_of_goods","label": "Description of goods"},
            {"field": "Наименование_ДС_англ","label": "Наименование ДС англ"},
            {"field": "Наименование_ДС_рус","label": "Наименование ДС рус"},
            {"field": "Contents","label": "Contents"},
            {"field": "LOT","label": "LOT"},
            {"field": "Case_Qty","label": "Case Q'ty"},
            {"field": "ORDER","label": "ORDER","mono": True},
            {"field": "Unit_price","label": "Unit price","mono": True},
            {"field": "Amount","label": "Amount","mono": True},
            {"field": "仕入値段","label": "仕入値段","mono": True},
            {"field": "仕入値合計","label": "仕入値合計","mono": True},
            {"field": "利益","label": "利益","mono": True},
            {"field": "利益率","label": "利益率","mono": True},
            {"field": "合計重量","label": "合計重量","mono": True},
            {"field": "商品サイズ","label": "商品サイズ"},
            {"field": "Unit_NW","label": "Unit N/W(kg)","mono": True},
            {"field": "Total_NW","label": "Total N/W(kg)","mono": True},
            {"field": "成分","label": "成分","multiline": True},
            {"field": "Марка_бренд_ДС","label": "Марка (бренд) ДС"},
            {"field": "Производель_ДС","label": "Производель ДС"},
            {"field": "created_at","label": "作成","format": "datetime","mono": True},
            {"field": "updated_at","label": "更新","format": "datetime","mono": True},
            {"field": "RequiredProduct_flg", "label": "現地登録必要商品"},
        ],
    },
}

VENDOR_CHOICES = [(k, v["label"]) for k, v in VENDOR_CONFIG.items()]
LABEL_MAP = {k: v["label"] for k, v in VENDOR_CONFIG.items()}

def _resolve_vendor(request):
    v = request.GET.get("vendor") or request.POST.get("vendor") or "nipponikatrading"
    return v if v in VENDOR_CONFIG else "nipponikatrading"

def product_import(request):
    # GET はフォーム表示のみ
    if request.method == "GET":
        return render(request, "kseurasia_manage_app/product_import.html", {"form": ProductImportForm()})

    # POST
    form = ProductImportForm(request.POST, request.FILES)
    if not form.is_valid():
        messages.error(request, "ファイルを選択してください。")
        return render(request, "kseurasia_manage_app/product_import.html", {"form": form})

    f = form.cleaned_data["file"]
    vendor = form.cleaned_data["vendor"]
    RequiredColor = "#F3CFA3"

    def _hex6(val: str) -> str:
        s = (val or "").strip().lstrip("#")
        if len(s) == 8:  # ARGB の場合は末尾6桁（RGB）を使用
            s = s[-6:]
        return s.upper()

    REQ_HEX = _hex6(RequiredColor)

    def _cell_bg_hex6(cell):
        """
        セルの塗り色を #RRGGBB（大文字）で返す。取得できなければ None。
        ※ 条件付き書式の最終色は反映されません（通常のセル塗りはOK）。
        """
        col = getattr(cell.fill, "fgColor", None) or getattr(cell.fill, "start_color", None)
        if not col:
            return None
        t = getattr(col, "type", None)
        if t == "rgb" and getattr(col, "rgb", None):
            return _hex6(col.rgb)
        if t == "indexed":
            return (COLOR_INDEX.get(col.indexed, "000000") or "000000").upper()
        # theme 等は簡易対応しない
        return None

    # ワークブック読み込み
    try:
        wb = openpyxl.load_workbook(io.BytesIO(f.read()), data_only=True)
    except Exception as e:
        messages.error(request, f"Excelの読み込みに失敗しました: {e}")
        return render(request, "kseurasia_manage_app/product_import.html", {"form": ProductImportForm()})

    if vendor == "NIPPONIKATRADING社向け":
        sheet_name = "ORDER SHEET"
        header_row = 6
    elif vendor == "ROYAL COSMETICS社向け":
        sheet_name = "ORDER SHEET"
        header_row = 3
    elif vendor == "YAMATO/TOYO TRADING社向け":
        sheet_name = "20250804_更新"
        header_row = 20

    # 固定シート取得
    if sheet_name not in wb.sheetnames:
        messages.error(request, f"シート '{sheet_name}' が見つかりません。存在: {', '.join(wb.sheetnames)}")
        return render(request, "kseurasia_manage_app/product_import.html", {"form": ProductImportForm()})
    ws = wb[sheet_name]

    # --- ヘッダー正規化（ここが重要） ---
    def _norm_header(s: str) -> str:
        s = (s or "").strip()
        s = s.replace("\u00A0", " ").replace("\u3000", " ")  # NBSP/全角空白
        s = s.replace("’", "'").replace("‘", "'").replace("“", '"').replace("”", '"')  # スマートクォート
        while "  " in s:
            s = s.replace("  ", " ")
        return s

    def _nz(v: object) -> str:
        """None/空/空白は空文字に正規化"""
        return str(v).strip() if v is not None else ""

    raw_headers = ["" if c.value is None else str(c.value).strip() for c in ws[header_row]]
    headers = [_norm_header(h) for h in raw_headers]
    valid_idx = [i for i, h in enumerate(headers) if h]

    # デバッグ: 実ヘッダー表示
    messages.info(request, f"[HEADER] {headers[:12]}{' …' if len(headers) > 12 else ''}")

    # モデルに実在するフィールド名セット
    if vendor == "NIPPONIKATRADING社向け":
        model_fields = {
            f.name for f in NIPPONIKATRADING_ProductInfo._meta.get_fields()
            if getattr(f, "concrete", False) and not getattr(f, "many_to_many", False) and not getattr(f, "auto_created", False)
        }
    elif vendor == "ROYAL COSMETICS社向け":
        model_fields = {
            f.name for f in RY_ProductInfo._meta.get_fields()
            if getattr(f, "concrete", False) and not getattr(f, "many_to_many", False) and not getattr(f, "auto_created", False)
        }
    elif vendor == "YAMATO/TOYO TRADING社向け":
        model_fields = {
            f.name for f in YAMATO_TOYO_ProductInfo._meta.get_fields()
            if getattr(f, "concrete", False) and not getattr(f, "many_to_many", False) and not getattr(f, "auto_created", False)
        }

    # シートに実在するヘッダーだけを対象にした有効マップ
    sheet_header_set = set(headers)
    if vendor == "NIPPONIKATRADING社向け":
        effective_map = {
            _norm_header(k): v
            for k, v in NIPPONIKATRADING_HEADER_MAP.items()
            if v and _norm_header(k) in sheet_header_set and v in model_fields
        }
    elif vendor == "ROYAL COSMETICS社向け":
        effective_map = {
            _norm_header(k): v
            for k, v in HEADER_MAP.items()
            if v and _norm_header(k) in sheet_header_set and v in model_fields
        }
    elif vendor == "YAMATO/TOYO TRADING社向け":
        effective_map = {
            _norm_header(k): v
            for k, v in YAMATO_TOYO_HEADER_MAP.items()
            if v and _norm_header(k) in sheet_header_set and v in model_fields
        }

    # ずれの可視化
    if vendor == "NIPPONIKATRADING社向け":
        missing_in_sheet = [k for k in NIPPONIKATRADING_HEADER_MAP.keys() if _norm_header(k) not in sheet_header_set]
        missing_in_model = [v for v in NIPPONIKATRADING_HEADER_MAP.values() if v and v not in model_fields]
    elif vendor == "ROYAL COSMETICS社向け":
        missing_in_sheet = [k for k in HEADER_MAP.keys() if _norm_header(k) not in sheet_header_set]
        missing_in_model = [v for v in HEADER_MAP.values() if v and v not in model_fields]
    elif vendor == "YAMATO/TOYO TRADING社向け":
        missing_in_sheet = [k for k in YAMATO_TOYO_HEADER_MAP.keys() if _norm_header(k) not in sheet_header_set]
        missing_in_model = [v for v in YAMATO_TOYO_HEADER_MAP.values() if v and v not in model_fields]

    if missing_in_sheet:
        messages.warning(request, f"[WARN] Excelに無いヘッダー: {missing_in_sheet[:8]}{' …' if len(missing_in_sheet)>8 else ''}")
    if missing_in_model:
        messages.error(request, f"[ERROR] モデルに無いフィールド: {missing_in_model}")
        return render(request, "kseurasia_manage_app/product_import.html", {"form": ProductImportForm()})

    # --- 行処理（空レコード抑止） ---
    objects = []
    empty_row_skipped = 0
    for r in range(header_row + 1, ws.max_row + 1):
        vals = ["" if c.value is None else str(c.value).strip() for c in ws[r]]
        if not any(vals):  # 完全空行
            continue

        # Excel行辞書（正規化ヘッダー → 値）
        row_dict = {headers[i]: (vals[i] if i < len(vals) else "") for i in valid_idx}

        # モデル用にキー変換（値が空のものは入れない）
        data = {}
        non_empty = 0
        for src, dest in effective_map.items():
            v = row_dict.get(src, "")
            if v == "":
                continue
            data[dest] = v
            non_empty += 1

        if non_empty == 0:  # ← ここで空レコードを明示的に弾く
            empty_row_skipped += 1
            continue

        if vendor == "ROYAL COSMETICS社向け":
            if not _nz(data.get("Brand_name")) or not _nz(data.get("Product_name")):
                continue

        elif vendor == "NIPPONIKATRADING社向け":
            # 商品名は「日本語名」または「Description_of_goods」のどちらか必須
            if not _nz(data.get("Brand_name")) or not (_nz(data.get("日本語名")) or _nz(data.get("Description_of_goods"))):
                continue

        elif vendor == "YAMATO/TOYO TRADING社向け":
            if not _nz(data.get("Brand")) or not _nz(data.get("Item_Name")):
                continue
            data["Brand"] = "UTENA"

        g_hex = _cell_bg_hex6(ws.cell(row=r, column=7))  # G列
        if g_hex and g_hex == REQ_HEX:
            data["RequiredProduct_flg"] = True

        if vendor == "NIPPONIKATRADING社向け":
            objects.append(NIPPONIKATRADING_ProductInfo(**data))
        elif vendor == "ROYAL COSMETICS社向け":
            objects.append(RY_ProductInfo(**data))
        elif vendor == "YAMATO/TOYO TRADING社向け":
            objects.append(YAMATO_TOYO_ProductInfo(**data))

    if not objects:
        messages.warning(request, "取り込める行がありませんでした。")
        return render(request, "kseurasia_manage_app/product_import.html", {"form": ProductImportForm()})

    # 一括削除 → 一括登録（トランザクションで一気に）
    try:
        with transaction.atomic():
            if vendor == "NIPPONIKATRADING社向け":
                deleted, _ = NIPPONIKATRADING_ProductInfo.objects.all().delete()
                NIPPONIKATRADING_ProductInfo.objects.bulk_create(objects, batch_size=500)
            elif vendor == "ROYAL COSMETICS社向け":
                deleted, _ = RY_ProductInfo.objects.all().delete()
                RY_ProductInfo.objects.bulk_create(objects, batch_size=500)
            elif vendor == "YAMATO/TOYO TRADING社向け":
                deleted, _ = YAMATO_TOYO_ProductInfo.objects.all().delete()
                YAMATO_TOYO_ProductInfo.objects.bulk_create(objects, batch_size=500)
    except Exception as e:
        messages.error(request, f"登録中にエラーが発生しました: {e}")
        return render(request, "kseurasia_manage_app/product_import.html", {"form": ProductImportForm()})

    messages.success(request, f"取り込み完了: 旧データ {deleted} 件削除、新規 {len(objects)} 件登録しました。")
    return redirect("product_list")

def product_list(request):
    vendor = _resolve_vendor(request)
    cfg = VENDOR_CONFIG[vendor]
    Model = cfg["model"]

    qs = Model.objects.all().order_by("-updated_at", "-id")
    page = request.GET.get("page", 1)
    page_obj = Paginator(qs, 50).get_page(page)

    return render(request, "kseurasia_manage_app/product_list.html", {
        "page_obj": page_obj,
        "vendor": vendor,
        "vendor_label": cfg["label"],
        "vendor_choices": VENDOR_CHOICES,
        "list_columns": cfg["list_columns"],
    })

def product_detail(request, pk: int):
    # ベンダー解決 → モデル・表示設定を取得
    vendor = _resolve_vendor(request)
    cfg = VENDOR_CONFIG[vendor]
    Model = cfg["model"]
    obj = get_object_or_404(Model, pk=pk)

    if request.method == "POST":
        action = request.POST.get("action", "save")

        # 削除
        if action == "delete":
            obj.delete()
            messages.success(request, f"{cfg['label']} #{pk} を削除しました。")
            return redirect(f"{reverse('product_list')}?vendor={vendor}")

        # ---- ここから保存（更新） ----
        # 編集可能フィールド名をモデルメタから抽出（id/created/updated などは除外）
        editable = []
        for f in Model._meta.get_fields():
            if not getattr(f, "concrete", False):
                continue
            if getattr(f, "auto_created", False):
                continue
            # AutoField を除外
            from django.db import models as _dj_models  # 念のためローカル参照
            if isinstance(f, _dj_models.AutoField):
                continue
            if f.name in ("id", "created_at", "updated_at"):
                continue
            editable.append(f.name)

        # 値の適用：RequiredProduct_flg だけは「必要/不要」→ bool に正規化
        def _to_bool_required(val) -> bool:
            s = str(val).strip().lower()
            true_words  = {"true", "1", "yes", "y", "on", "必要", "ひつよう", "必須", "required"}
            false_words = {"false","0","no",  "n", "off","不要", "ふよう", "not required"}
            if s in true_words:
                return True
            if s in false_words:
                return False
            # 空文字など曖昧値は False に倒す（要件に応じて調整可）
            return False

        for name in editable:
            if name in request.POST:
                raw = request.POST.get(name, "")
                if name == "RequiredProduct_flg":
                    setattr(obj, name, _to_bool_required(raw))
                else:
                    # 空は None にしておきたい場合の吸収
                    setattr(obj, name, raw if raw != "" else None)

        obj.save()
        messages.success(request, "変更を保存しました。")
        # 同ページ再読込（vendor パラメータ維持）
        return redirect(request.path + f"?vendor={vendor}")

    # GET：詳細表示
    return render(request, "kseurasia_manage_app/product_detail.html", {
        "obj": obj,
        "vendor": vendor,
        "vendor_label": cfg["label"],
        "detail_columns": cfg["detail_columns"],
    })

#4レポートダウンロード
def reports_export_bundle(request):
    """
    ?start_month=YYYY-MM&end_month=YYYY-MM を受け取り、
    4つのレポート（売上表/売掛/AP/収支）を生成して1つのZIPで返す。
    """
    # ラベル（期間文字列）だけ先に取得（例: 2025-02_2025-09）
    try:
        _, _, _, label = _get_report_params_range(request)  # 既存ヘルパを使う
    except Exception as e:
        return HttpResponse(f"Bad params: {e}", status=400)

    exporters = [
        ("売上表",      reports_sales_export),
        ("売掛金管理表", reports_ar_export),
        ("買掛金管理表", reports_ap_export),
        ("収支予定表",   reports_cashflow_export),
    ]

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for _title, fn in exporters:
            resp = fn(request)  # 既存の各エクスポート（FileResponse）をそのまま呼ぶ

            # ファイル名を Content-Disposition から抽出（fallbackあり）
            cd = resp.get("Content-Disposition", "")
            m = re.search(r'filename="([^"]+)"', cd)
            filename = m.group(1) if m else "report.xlsx"

            # 本文をバイト列として取得（FileResponseはstreaming_content）
            if hasattr(resp, "streaming_content"):
                content = b"".join(resp.streaming_content)
            else:
                content = resp.content

            zf.writestr(filename, content)

    zip_name = f"Reports_{label}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
    buf.seek(0)
    resp = HttpResponse(buf.getvalue(), content_type="application/zip")
    resp["Content-Disposition"] = f'attachment; filename="{escape_uri_path(zip_name)}"'
    return resp

#以下ランキング
def _parse_yyyy_mm_dd(s: str) -> date:
    try:
        dt = datetime.strptime(s.strip(), "%Y-%m-%d")
        return date(dt.year, dt.month, dt.day)
    except Exception:
        raise Http404("start_date / end_date は 'YYYY-MM-DD' 形式で指定してください。")

def _aware_range_from_dates(start_d: date, end_d: date):
    """
    naive date の [start_d, end_d] を、tz aware の [start_dt, end_dt_next) に変換
    """
    tz = timezone.get_current_timezone()
    start_dt = timezone.make_aware(datetime(start_d.year, start_d.month, start_d.day, 0, 0, 0), tz)
    # end は翌日の0:00を排他的上限に
    end_dt   = timezone.make_aware(datetime(end_d.year, end_d.month, end_d.day, 0, 0, 0), tz) + timedelta(days=1)
    return start_dt, end_dt

def _to_int(val):
    # 文字列の金額/数量をざっくり整数化（空/Noneは0）
    if val is None:
        return 0
    s = str(val).strip().replace(",", "")
    if s == "":
        return 0
    try:
        return int(float(s))
    except Exception:
        return 0

def _collect_rows(start_dt, end_dt, vendor: str):
    """
    期間内の各社の rows を辞書で返す。
    vendor = 'all' / 'ROYAL COSMETICS' / 'NIPPONIKATRADING' / 'YAMATO_TOYO'
    """
    rows = {"ROYAL COSMETICS": [], "NIPPONIKATRADING": [], "YAMATO_TOYO": []}

    # ROYAL
    if vendor in ("all", "ROYAL COSMETICS"):
        rc = (
            OrderContent.objects
            .select_related("batch")
            .filter(batch__created_at__gte=start_dt, batch__created_at__lt=end_dt, batch__buyers="ROYAL COSMETICS")
            .values("Brand_name", "Product_name", "Order", "Amount", "batch__created_at")
            .order_by("id")
        )
        rows["ROYAL COSMETICS"] = list(rc)

    # NIPPONIKATRADING
    if vendor in ("all", "NIPPONIKATRADING"):
        nk = (
            NIPPONIKATRADING_OrderContent.objects
            .select_related("batch")
            .filter(batch__created_at__gte=start_dt, batch__created_at__lt=end_dt, batch__buyers="NIPPONIKATRADING")
            .values("Brand_name", "日本語名", "Description_of_goods", "ORDER", "Amount", "batch__created_at")
            .order_by("id")
        )
        rows["NIPPONIKATRADING"] = list(nk)

    # YAMATO_TOYO
    if vendor in ("all", "YAMATO_TOYO"):
        ym = (
            YAMATO_TOYO_OrderContent.objects
            .select_related("batch")
            .filter(batch__created_at__gte=start_dt, batch__created_at__lt=end_dt, batch__buyers="YAMATO_TOYO")
            .values("Brand", "Item_Name", "Quantity", "総販売価格", "batch__created_at")
            .order_by("id")
        )
        rows["YAMATO_TOYO"] = list(ym)

    return rows

def _resolve_product_link(brand: str, name: str):
    """
    ランキングの商品名/ブランドから、登録商品の (vendor_key, pk) を推定する。
    見つからなければ (None, None) を返す。
    vendor_key は product_detail が参照する ?vendor= の値（'royal'/'nipponika'/'yamato'）に合わせる。
    """
    brand = (brand or "").strip()
    name  = (name or "").strip()
    if not name:
        return (None, None)

    # ROYAL
    try:
        pk = (RY_ProductInfo.objects
              .filter(Q(Product_name__iexact=name) | Q(日本語名__iexact=name))
              .filter(Brand_name__iexact=brand if brand else Q())
              .values_list("id", flat=True).first())
        if pk:
            return ("royal", pk)
    except Exception:
        pass

    # NIPPONIKA
    try:
        pk = (NIPPONIKATRADING_ProductInfo.objects
              .filter(Q(日本語名__iexact=name) | Q(Description_of_goods__iexact=name))
              .filter(Brand_name__iexact=brand if brand else Q())
              .values_list("id", flat=True).first())
        if pk:
            return ("nipponika", pk)
    except Exception:
        pass

    # YAMATO/TOYO （Brand は実運用的に UTENA 固定が多い想定）
    try:
        qs = YAMATO_TOYO_ProductInfo.objects.filter(Item_Name__iexact=name)
        if brand:
            qs = qs.filter(Brand__iexact=brand)
        pk = qs.values_list("id", flat=True).first()
        if pk:
            return ("yamato", pk)
    except Exception:
        pass

    return (None, None)

def _make_rankings(rows_dict, limit: int = None):
    """
    rows_dict = {
      "ROYAL COSMETICS": [ {...}, ... ],
      "NIPPONIKATRADING": [ {...}, ... ],
      "YAMATO_TOYO": [ {...}, ... ],
    }
    を受けて、(buyer, brand, product) 各ランキングを返す
    """
    # buyer 別
    buyer_total = defaultdict(lambda: {"sales": 0, "qty": 0})

    # brand 別（全社横断だが「バイヤー×ブランド」で行を持つ）
    brand_total = defaultdict(lambda: {"sales": 0, "qty": 0})  # key=(buyer, brand)

    # product 別（全社横断だが「バイヤー×ブランド×商品」で行を持つ）
    product_total = defaultdict(lambda: {"sales": 0, "qty": 0})  # key=(buyer, brand, name)

    # ROYAL 集計
    for r in rows_dict.get("ROYAL COSMETICS", []):
        buyer = "ROYAL COSMETICS"
        brand = (r.get("Brand_name") or "").strip()
        pname = (r.get("Product_name") or "").strip()
        qty   = _to_int(r.get("Order"))
        sales = _to_int(r.get("Amount"))

        buyer_total[buyer]["sales"] += sales
        buyer_total[buyer]["qty"]   += qty

        if brand:
            brand_total[(buyer, brand)]["sales"] += sales
            brand_total[(buyer, brand)]["qty"]   += qty
        if pname:
            product_total[(buyer, brand, pname)]["sales"] += sales
            product_total[(buyer, brand, pname)]["qty"]   += qty

    # NIPPONIKA 集計
    for r in rows_dict.get("NIPPONIKATRADING", []):
        buyer = "NIPPONIKATRADING"
        brand = (r.get("Brand_name") or "").strip()
        # 商品名は「日本語名」優先、なければ「Description_of_goods」
        pname = (r.get("日本語名") or "").strip() or (r.get("Description_of_goods") or "").strip()
        qty   = _to_int(r.get("ORDER"))
        sales = _to_int(r.get("Amount"))

        buyer_total[buyer]["sales"] += sales
        buyer_total[buyer]["qty"]   += qty

        if brand:
            brand_total[(buyer, brand)]["sales"] += sales
            brand_total[(buyer, brand)]["qty"]   += qty
        if pname:
            product_total[(buyer, brand, pname)]["sales"] += sales
            product_total[(buyer, brand, pname)]["qty"]   += qty

    # YAMATO 集計（売上は「総販売価格」）
    for r in rows_dict.get("YAMATO_TOYO", []):
        buyer = "YAMATO_TOYO"
        brand = (r.get("Brand") or "").strip()
        pname = (r.get("Item_Name") or "").strip()
        qty   = _to_int(r.get("Quantity"))
        sales = _to_int(r.get("総販売価格"))

        buyer_total[buyer]["sales"] += sales
        buyer_total[buyer]["qty"]   += qty

        if brand:
            brand_total[(buyer, brand)]["sales"] += sales
            brand_total[(buyer, brand)]["qty"]   += qty
        if pname:
            product_total[(buyer, brand, pname)]["sales"] += sales
            product_total[(buyer, brand, pname)]["qty"]   += qty

    # 並び替え（売上降順、同額は数量降順）
    def _sorted_list(items):
        return sorted(items, key=lambda x: (x["sales"], x["qty"]), reverse=True)

    # buyer ランキング（従来通り）
    buyer_rank = _sorted_list([{"key": k, **v} for k, v in buyer_total.items()])

    # brand ランキング（Buyer 列を追加）
    brand_rank = _sorted_list([
        {"buyer": k[0], "brand": k[1], "sales": v["sales"], "qty": v["qty"]}
        for k, v in brand_total.items()
    ])

    # product ランキング（Buyer/Brand 列を追加。リンク解決もここで）
    product_rank = _sorted_list([
        {"buyer": k[0], "brand": k[1], "name": k[2], "sales": v["sales"], "qty": v["qty"]}
        for k, v in product_total.items()
    ])

    # 任意の limit 指定がある場合にスライス（None/0 は全件）
    if limit:
        buyer_rank   = buyer_rank[:limit]
        brand_rank   = brand_rank[:limit]
        product_rank = product_rank[:limit]

    # ▼ 商品ランキングの各行に「登録商品へのリンク情報」を付与
    for r in product_rank:
        vendor_key, pk = _resolve_product_link(r.get("brand"), r.get("name"))
        r["vendor"] = vendor_key  # 'royal'/'nipponika'/'yamato' or None
        r["pk"] = pk              # int or None

    return buyer_rank, brand_rank, product_rank

def ranking_console(request):
    # ▼ パラメータ
    start_s = (request.GET.get("start_date") or "").strip()
    end_s   = (request.GET.get("end_date") or "").strip()
    vendor  = (request.GET.get("vendor") or "all").strip()  # all / ROYAL COSMETICS / NIPPONIKATRADING / YAMATO_TOYO
    # Excel 用の想定で残しておく（画面の表示には使わない）
    try:
        limit = int(request.GET.get("limit") or "50")
    except Exception:
        limit = 50
    # 初期表示のプレビュー件数（各セクションの上位のみ）
    try:
        preview = int(request.GET.get("preview") or "10")
    except Exception:
        preview = 10
    expand = (request.GET.get("expand") or "").strip()  # buyer / brand / product / all / ""

    # 既定：当月1日～本日
    today = timezone.localdate()
    if not start_s:
        start_d = today.replace(day=1)
    else:
        start_d = _parse_yyyy_mm_dd(start_s)

    end_d = _parse_yyyy_mm_dd(end_s) if end_s else today
    start_dt, end_dt = _aware_range_from_dates(start_d, end_d)

    rows_dict = _collect_rows(start_dt, end_dt, vendor)
    # 画面では全件を計算しておき、テンプレでプレビューor全件を切替
    buyer_rank, brand_rank, product_rank = _make_rankings(rows_dict, limit=None)

    expand_all = (expand == "all")
    expanded_buyer = expand_all or (expand == "buyer")
    expanded_brand = expand_all or (expand == "brand")
    expanded_product = expand_all or (expand == "product")

    ctx = {
        "start_date": start_d.strftime("%Y-%m-%d"),
        "end_date": end_d.strftime("%Y-%m-%d"),
        "vendor": vendor,
        "limit": limit,
        "preview": preview,
        "expand": expand,
        "expand_all": expand_all,
        "expanded_buyer": expanded_buyer,
        "expanded_brand": expanded_brand,
        "expanded_product": expanded_product,
        "buyer_rank": buyer_rank,
        "brand_rank": brand_rank,
        "product_rank": product_rank,
    }
    return render(request, "kseurasia_manage_app/rankings.html", ctx)

def rankings_export(request):
    # 画面と同じパラメータを解釈
    start_s = (request.GET.get("start_date") or "").strip()
    end_s   = (request.GET.get("end_date") or "").strip()
    vendor  = (request.GET.get("vendor") or "all").strip()
    try:
        limit = int(request.GET.get("limit") or "1000")  # Excelはデフォルト広め
    except Exception:
        limit = 1000

    today = timezone.localdate()
    start_d = _parse_yyyy_mm_dd(start_s) if start_s else today.replace(day=1)
    end_d   = _parse_yyyy_mm_dd(end_s)   if end_s   else today
    start_dt, end_dt = _aware_range_from_dates(start_d, end_d)

    rows_dict = _collect_rows(start_dt, end_dt, vendor)
    buyer_rank, brand_rank, product_rank = _make_rankings(rows_dict, limit)

    # --- Excel 生成 ---
    wb = openpyxl.Workbook()
    thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                  top=Side(style="thin"), bottom=Side(style="thin"))
    head_font = Font(bold=True)
    align_center = Alignment(horizontal="center", vertical="center")

    def _write_sheet_buyer(items):
        ws = wb.create_sheet("Buyer")
        ws.append(["Rank", "Buyer", "Sales(¥)", "Qty"])
        # ヘッダ装飾
        for c in range(1, 5):
            ws.cell(1, c).font = head_font
            ws.cell(1, c).alignment = align_center
            ws.cell(1, c).border = thin
        # 本文
        for i, row in enumerate(items, start=1):
            ws.append([i, row["key"], int(row["sales"]), int(row["qty"])])
            for c in range(1, 4+1):
                ws.cell(i+1, c).border = thin
        # 幅/書式
        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 24
        ws.column_dimensions["C"].width = 14
        ws.column_dimensions["D"].width = 8
        for r in range(2, len(items)+2):
            ws.cell(r, 3).number_format = r'¥#,##0'
            ws.cell(r, 4).number_format = r'#,##0'
        try:
            ws.header_footer.left_header = f"&L期間: {start_d:%Y-%m-%d} ～ {end_d:%Y-%m-%d}"
            ws.header_footer.right_header = f"&R{timezone.localtime().strftime('%Y-%m-%d %H:%M')}"
        except AttributeError:
            pass

    def _write_sheet_brand(items):
        ws = wb.create_sheet("Brand")
        ws.append(["Rank", "Buyer", "Brand", "Sales(¥)", "Qty"])
        for c in range(1, 5+1):
            ws.cell(1, c).font = head_font
            ws.cell(1, c).alignment = align_center
            ws.cell(1, c).border = thin
        for i, row in enumerate(items, start=1):
            ws.append([i, row["buyer"], row["brand"], int(row["sales"]), int(row["qty"])])
            for c in range(1, 5+1):
                ws.cell(i+1, c).border = thin
        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 28
        ws.column_dimensions["D"].width = 14
        ws.column_dimensions["E"].width = 8
        for r in range(2, len(items)+2):
            ws.cell(r, 4).number_format = r'¥#,##0'
            ws.cell(r, 5).number_format = r'#,##0'
        try:
            ws.header_footer.left_header = f"&L期間: {start_d:%Y-%m-%d} ～ {end_d:%Y-%m-%d}"
            ws.header_footer.right_header = f"&R{timezone.localtime().strftime('%Y-%m-%d %H:%M')}"
        except AttributeError:
            pass

    def _write_sheet_product(items):
        ws = wb.create_sheet("Product")
        ws.append(["Rank", "Buyer", "Brand", "Product", "Sales(¥)", "Qty"])
        for c in range(1, 6+1):
            ws.cell(1, c).font = head_font
            ws.cell(1, c).alignment = align_center
            ws.cell(1, c).border = thin
        for i, row in enumerate(items, start=1):
            ws.append([i, row["buyer"], row["brand"], row["name"], int(row["sales"]), int(row["qty"])])
            for c in range(1, 6+1):
                ws.cell(i+1, c).border = thin
        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 24
        ws.column_dimensions["D"].width = 40
        ws.column_dimensions["E"].width = 14
        ws.column_dimensions["F"].width = 8
        for r in range(2, len(items)+2):
            ws.cell(r, 5).number_format = r'¥#,##0'
            ws.cell(r, 6).number_format = r'#,##0'
        try:
            ws.header_footer.left_header = f"&L期間: {start_d:%Y-%m-%d} ～ {end_d:%Y-%m-%d}"
            ws.header_footer.right_header = f"&R{timezone.localtime().strftime('%Y-%m-%d %H:%M')}"
        except AttributeError:
            pass

    # 先頭の空シートを削除してから追加
    del wb[wb.sheetnames[0]]
    _write_sheet_buyer(buyer_rank)
    _write_sheet_brand(brand_rank)
    _write_sheet_product(product_rank)

    filename = f"rankings_{start_d:%Y%m%d}_{end_d:%Y%m%d}.xlsx"
    from io import BytesIO
    bio = BytesIO()
    wb.save(bio); bio.seek(0)
    return FileResponse(bio, as_attachment=True, filename=filename)