from typing import Dict, List, Any, Optional
from decimal import Decimal, InvalidOperation
import io, re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

def _norm(s: Any) -> str:
    if s is None: return ""
    s = str(s).lower().strip()
    s = s.replace("　","").replace(" ","")
    s = re.sub(r"[\'’`.\-_/]", "", s)
    return s

def _to_qty(v: Any) -> Decimal:
    if v is None: return Decimal(0)
    s = str(v).strip()
    if not s: return Decimal(0)
    if s in {"○","●","◯","x","X"}: return Decimal(1)
    try:
        return Decimal(s.replace(",",""))
    except (InvalidOperation, AttributeError):
        return Decimal(0)

def _to_str(v: Any) -> Optional[str]:
    if v is None: return None
    s = str(v)
    return s if s.strip() else None

# ====== Brand名→テンプレシート名のエイリアス定義（必要に応じて追記） ======
BRAND_TO_SHEET_ALIAS: Dict[str, str] = {
    # 既存サンプル（表記ゆれ/TESTER/PROなどを吸収）
    "cbon": "C'BON",
    "estlabo": "ESTLABO", "estlabopro": "ESTLABO",
    "elegadoll": "ELEGADOLL",
    "evliss": "Evliss", "evlisstester": "Evliss",
    "luxces": "Luxces", "luxcestester": "Luxces",
    "lapidem": "LAPIDEM", "lapidempro": "LAPIDEM",
    "cosmepro": "COSMEPRO", "cosmeprotester": "COSMEPRO",
    "afura": "AFURA", "afuratester": "AFURA",
    "beautyconexion": "Beauty Conexion", "beatyconexion": "Beauty Conexion",
    "cocochi": "COCOCHI　発注書",  # 全角スペースに注意
    "himelabo": "HIMELABO",
    "lejeu": "LEJEU",
    "hanako": "HANAKO",
    "mccoy": "McCoy", "mccoypro": "McCoy",
    "medion": "Dr.Medion", "medionsample": "Dr.Medion",
    "diaasjapan": "Diaas",
    "rosydrop": "ROSY DROP",
    "relent": "リレント通常注文",

    # ★未解決ブランドに対して、寄せ先が決まればここに追記してください
    # "beautygarage": "Rey Beauty",          # ← 例
    # "dimehealthcarepro": "DIME（仮のシート名）",
}

def _find_header_row(ws: Worksheet, search_rows: int = 40) -> Optional[int]:
    """テンプレ側ヘッダー（"Jan code" を含む行）を 1-based で返す"""
    for r in range(1, search_rows+1):
        vals = [str(c.value).strip() if c.value is not None else "" for c in ws[r]]
        if "jancode" in [_norm(x) for x in vals]:
            return r
    return None

def _header_map(ws: Worksheet, header_row: int) -> Dict[str, str]:
    """テンプレのヘッダ行を正規化→列記号にマッピング"""
    m: Dict[str, str] = {}
    for cell in ws[header_row]:
        key = _norm(cell.value)
        if key:
            m[key] = cell.column_letter
    return m

def build_multi_brand_po_from_order_sheet(
    *,
    uploaded_excel_bytes: bytes,
    order_sheet_name: str = "ORDER SHEET",
    order_header_row_zero_based: int = 2,  # 3行目がヘッダー
    brand_column_header: str = "Brand name",
    order_column_header: str = "ORDER",
    blank_line_between_header_and_detail: int = 1,  # ヘッダーの次に空行1行
    default_sheet_if_missing: Optional[str] = None, # 未知ブランドの寄せ先（なければスキップ）
) -> Dict[str, Any]:
    """
    - ORDER SHEET を読み、ORDER>0 の行だけをブランド毎にテンプレ各シートへ転記
    - 書式はテンプレそのまま（openpyxl）
    - シート解決: 正規化一致 → BRAND_TO_SHEET_ALIAS → default_sheet_if_missing
    戻値: {"filename": str, "buffer": BytesIO, "report": {...}}
    """
    bio = io.BytesIO(uploaded_excel_bytes)

    # 入力（ORDER SHEET）
    df = pd.read_excel(io.BytesIO(bio.getvalue()),
                       sheet_name=order_sheet_name,
                       engine="openpyxl",
                       header=order_header_row_zero_based)
    df["__qty__"] = df[order_column_header].apply(_to_qty)
    active = df[df["__qty__"] > 0].copy()
    if active.empty:
        return {"filename": "purchase_orders_by_brand.xlsx",
                "buffer": io.BytesIO(),
                "report": {"written": 0, "brands": {}, "skipped_brands": [], "reason": "no active rows"}}

    # テンプレ（同一ファイル）を openpyxl で開く
    wb = load_workbook(io.BytesIO(bio.getvalue()))
    sheetnames = wb.sheetnames
    norm_sheet_map = {_norm(s): s for s in sheetnames}

    # 元の列名を正規化辞書に（テンプレのヘッダーと突合するため）
    src_header_norm = {_norm(c): c for c in active.columns}

    written_total = 0
    brand_report: Dict[str, Dict[str, Any]] = {}
    skipped_brands: List[str] = []

    # ブランド毎に展開
    for brand, gdf in active.groupby(brand_column_header, dropna=True):
        brand_str = str(brand).strip()
        key = _norm(brand_str)

        # 対象シートの決定
        target_sheet = (
            norm_sheet_map.get(key)
            or BRAND_TO_SHEET_ALIAS.get(key)
            or default_sheet_if_missing
        )

        if not target_sheet or target_sheet not in sheetnames:
            skipped_brands.append(brand_str)
            continue

        ws = wb[target_sheet]
        header_row = _find_header_row(ws)
        if not header_row:
            skipped_brands.append(brand_str)
            continue

        tmpl_map = _header_map(ws, header_row)
        # ヘッダー一致（テンプレ ∩ 入力）
        common_keys = sorted(set(tmpl_map.keys()) & set(src_header_norm.keys()))
        out_row = header_row + 1 + blank_line_between_header_and_detail

        written_rows = 0
        for _, row in gdf.iterrows():
            # 念のため二重チェック
            if _to_qty(row.get(order_column_header)) <= 0:
                continue
            for k in common_keys:
                src_col = src_header_norm[k]
                val = row.get(src_col)
                if k == _norm(order_column_header):
                    try:
                        ws[f"{tmpl_map[k]}{out_row}"].value = float(_to_qty(val))
                    except Exception:
                        ws[f"{tmpl_map[k]}{out_row}"].value = _to_str(val)
                else:
                    ws[f"{tmpl_map[k]}{out_row}"].value = _to_str(val)
            out_row += 1
            written_rows += 1

        brand_report[brand_str] = {"sheet": target_sheet, "rows": written_rows}
        written_total += written_rows

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return {
        "filename": "purchase_orders_by_brand.xlsx",
        "buffer": buf,
        "report": {"written": written_total,
                   "brands": brand_report,
                   "skipped_brands": skipped_brands},
    }
